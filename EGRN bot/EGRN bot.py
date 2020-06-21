import re, time

from recognize import recognize
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import *
from datetime import datetime

EGRN = None
wb = None
ws = None
tout = 290 ## пауза между запросами в сеукндах
  
##------------------------------------------------------------##
def EGRNinit():
    global EGRN
    EGRN = webdriver.Chrome()
    time.sleep(1)
    EGRN.set_page_load_timeout(15)
    try:
        EGRN.get("https://rosreestr.ru/wps/portal/p/cc_present/ir_egrn")
        EGRN.implicitly_wait(20)
    except:
        print("Сайт Росреестра не работает")
##------------------------------------------------------------##
def Wait(t):
    print(f"     Пауза между запросами {t} секунд [----------]")
    print(f"                             ожидание [", end = "")
    t10 = t // 10
    for t in range(0, 10):
        time.sleep(t10)
        print("X", end = "")
    time.sleep(t % 10)
    print("]")
##------------------------------------------------------------##
def RecognizeCaptcha(c):
    global EGRN
    img_base64 = EGRN.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", c)
    captcha = recognize(img_base64)
    return captcha
  
##------------------------------------------------------------##
def GetInfo():
    global EGRN
    global wb
    global ws
    global tout

    while EGRN.find_element_by_css_selector("div.blockGrey").is_displayed():
        time.sleep(1)

    KN = re.search(r"\b\d{2}:\d{2}:\d{1,7}:\d{1,}\b", EGRN.find_element_by_css_selector("div.header3").get_attribute("innerText"))[0]
    print(f"{KN} - подготовка запроса")
    vapp = EGRN.find_element_by_css_selector("div.v-app")

    Radio = vapp.find_elements_by_xpath("//input[@type='radio']")[1]
    CaptchaField = vapp.find_elements_by_xpath("//input[@type='text']")[1]
    CaptchaReload = vapp.find_element_by_xpath("//span[contains(@class,'v-button-caption') and contains(text(),'Другую картинку')]")
    SendButton = vapp.find_element_by_xpath("//span[contains(@class,'v-button-caption') and contains(text(),'Отправить запрос')]")

    EGRN.execute_script("arguments[0].scrollIntoView();", Radio)
    Radio.click()
##    print("Radio.click()", end = " ")
    while True:
        captcha = ""
        CaptchaImage = vapp.find_element_by_xpath("//img[@style='width: 180px; height: 50px;']")
        captcha = RecognizeCaptcha(CaptchaImage)
        if captcha != "" and captcha != "44444" :
            break
        print("     reloading captcha image")
        CaptchaReload.click()
        time.sleep(10)
        
    time.sleep(1)
    while True:
        while True:
            try:
                CaptchaField.click()
                break
            except:
                time.sleep(1)
                print("CaptchaField.click() wait")
            
##        print("CaptchaField.click()", end = " ")
        time.sleep(1)
        CaptchaField.send_keys(captcha)
        time.sleep(1)
        SendButton.click()
##        print("wait for popup", end = " ")
        PopUp = vapp.find_elements_by_xpath("//div[@class='v-window']")
        OK = False
        while True:
##            print("+", end = " ")
            t1 = time.time()
            while len(PopUp) == 0 and time.time() - t1 < 60: 
                print(".", end = " ")
                PopUp = vapp.find_elements_by_xpath("//div[@class='v-window']")
            if len(PopUp) != 0:
                OkButton = PopUp[0].find_elements_by_xpath("//span[contains(@class,'v-button-caption') and contains(text(),'Продолжить работу')]")[0]
                NZ = ""
                NZ = re.search(r"\d{2}\-\d{9}", PopUp[0].get_attribute("innerText"))[0]
                if NZ != "":
                    OK = True
                    break
            else:
                break
        if OK:
            break

    OkButton.click()
##    print("OkButton.click()")
    print(f"{KN} :: {NZ} - запрос выполнен " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    row = ws.max_row + 1
    ws.cell(row = row, column = 1).value = KN
    ws.cell(row = row, column = 2).value = NZ
    wb.save(filename = 'rq.xlsx')
    Wait(tout)
##------------------------------------------------------------##
def GetAll(t = 300):
    global EGRN
    global tout
    tout = t
    AllKN = []
    AllON = []
    i = 1
    DataTable = EGRN.find_element_by_css_selector("table.v-table-table")
    AllON = DataTable.find_elements_by_css_selector("tr > td:first-child")
    for ON in AllON:
        AllKN.append(ON.get_attribute("innerText"))
    for KN in AllKN:
        DataTable = EGRN.find_element_by_css_selector("table.v-table-table")
        AllON = []
        AllON = DataTable.find_elements_by_css_selector("tr > td:first-child")
        for ON in AllON:
            if KN == ON.get_attribute("innerText"):
                EGRN.execute_script("arguments[0].scrollIntoView();", ON)
                print(f"     Объект {i} из {len(AllON)} :: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " :: ", end = "")
                ON.click()
                time.sleep(2)
                GetInfo()
                i = i + 1
                break
    print("Всё готово!")

##------------------------------------------------------------##
EGRNinit()

try:
    wb = load_workbook(filename = 'rq.xlsx')
except:
    wb = Workbook()
ws = wb.worksheets[0]
if ws.cell(row = 1, column = 1).value == None:
    ws.cell(row = 1, column = 1).value = "Кадастровый номер"
    ws.cell(row = 1, column = 2).value = "Номер запроса"
    wb.save(filename = 'rq.xlsx')




##<div class="v-loading-indicator-wait" style="position: absolute; display: block;"></div>
