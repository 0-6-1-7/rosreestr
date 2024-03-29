import re, time

from datetime import datetime
from os.path import exists

from openpyxl import Workbook
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


EGRN = None
status = 0 # состояние сайта
row = 0
wb = None
ws = None
RQ = {} # отчёт по статусам запросов
p = 0 # счётчик строк на странице 0..25
PageRefreshRetriesMax = 10 # количество попыток перезагрузки страницы авторизации
default_implicitly_wait = 40

##------------------------------------------------------------##
def EGRNinit():
    global EGRN
    SiteRestartRetriesMax = 10 # количество попыток перезапуска сайта
    SiteRestartRetriesCounter = 0
    SiteStatusOK = False
# На случай если мы попали сюда после какой-нибудь ошибки - нужно начать заново
    if EGRN != None:
        EGRN.close()
        EGRN.quit()
    while True:
        try:
            chrome_options = Options()
            chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            EGRN = webdriver.Chrome(options = chrome_options)
            EGRN.set_page_load_timeout(40)
            EGRN.get("https://rosreestr.gov.ru/wps/portal/p/cc_present/ir_egrn")
            DemoKey = WebDriverWait(EGRN, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'6F9619FF-8B86-D011-B42D-00CF4FC964FF')]")))
            print("Сайт Росреестра работает, страница авторизации загружена нормально.")
            EGRN.implicitly_wait(default_implicitly_wait)
            return("SiteOK")
        except:
            SiteRestartRetriesCounter = SiteRestartRetriesCounter + 1
            print(f"Страница не загружена. Перезагрузка, попытка № {SiteRestartRetriesCounter}")
        EGRN.close()
        EGRN.quit()
        if SiteRestartRetriesCounter > SiteRestartRetriesMax:
            print("\n\n\n\nСайт Росреестра не работает. Мы пытались...")
            return("SiteFailed")

##------------------------------------------------------------##
def EGRNauth():
    global AuthKey
    global EGRN
    Status = None
    try:
        f = open("auth.txt", "r")
        AuthKey = f.readline()[:36]
        f.close()
        if re.search("\w{8}-\w{4}-\w{4}-\w{4}-\w{12}", AuthKey) != None: Status = "GotAuthKey"
        else: return("NoAuthKey")
    except: return("NoAuthKey")
        
    if Status == "GotAuthKey":
        vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")
        AuthKeyFields = vapp.find_elements(By.XPATH, ".//input[@type='text']")
        AuthKeyFields[0].send_keys(AuthKey)
        t0 = time.time()
        while True:
            if re.search(AuthKeyFields[4].get_attribute("value"), AuthKey) != None:
                Status = "KeyEntered"
                break
            time.sleep(5)
            if time.time() - t0 > 60:
                print("За 60 секунд ключ не принят.")
                return("AuthTimeOut")
    if Status == "KeyEntered":
        try:
            SendButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Войти')]")
            SendButton.click()
            return("AuthOK")
        except: return("AuthFailed")
    else: return("AuthFailed")
        
##------------------------------------------------------------##

def EGRNRequestPage():
    global EGRN
    vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")
    t0 = time.time()
    while True:
        try:
            sButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Мои заявки')]")
            break
        except:
            pass
        if time.time() - t0 > 60:
# За 60 секунд страница не загрузилась полностью
            return("RequestPageFailed")
    sButton.click()
    try:
        InfoImage = vapp.find_element(By.XPATH, ".//img[contains(@src, 'btn_first_na.gif')]")
        return("RequestPageOK")
    except:
        return("RequestPageFailed")
    
####------------------------------------------------------------##
def init_pyxl():
    global wb
    global ws
    try: wb = load_workbook(filename = 'dn.xlsx')
    except: wb = Workbook()
    ws = wb.worksheets[0]
    if ws.cell(row = 1, column = 1).value == None:
        ws.cell(row = 1, column = 1).value = "Номер запроса"
        ws.cell(row = 1, column = 2).value = "Дата запроса"
        ws.cell(row = 1, column = 3).value = "Статус запроса"
        wb.save(filename = 'dn.xlsx')

##------------------------------------------------------------##
def init_pyxlx(fn):
    global wb
    global ws
    if not fn.endswith(".xlsx"): fn += ".xlsx"
    if not exists(fn):
        print("Нет такого файла, дальнейшая работа невозможна")
        return "FileNotFound"

    try:
        wb = load_workbook(fn)
        ws = wb.worksheets[0]
        ws.cell(row = 2, column = 5).value = "Статус"
        wb.save(fn)
    except:
        print("Проблема с доступом к файлу, дальнейшая работа невозможна")
    return "OK"

##------------------------------------------------------------##
def CheckDNfile():
    global wb
    global ws
    wb = Workbook()
    ws = wb.worksheets[0]
    if ws.cell(row = 1, column = 1).value == None:
        ws.cell(row = 1, column = 1).value = "Номер запроса"
        ws.cell(row = 1, column = 2).value = "Дата запроса"
        ws.cell(row = 1, column = 3).value = "Статус запроса"
    try:
        wb.save(filename = 'dn.xlsx')
        return("FileDNOK")
    except:
        print("\n\n\n\n\nОшибка при сохранении файла dn.xlsx - вероятно, он открыт в Excel")
        return("FileDNFailed")

##------------------------------------------------------------##
def dn(d1, d2, download):
    global row
    global wb
    global ws
    global EGRN
    global RQ
    global p
    p = 0
    table = EGRN.find_element(By.CSS_SELECTOR, "table.v-table-table")
    for tr in table.find_elements(By.CSS_SELECTOR, "tr"):
        NZ = tr.find_elements(By.CSS_SELECTOR, "td")[0].get_attribute("innerText")
        found = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(row = r, column = 1).value == NZ:
                found = True
                row = r
                break

        DZ = tr.find_elements(By.CSS_SELECTOR, "td")[1].get_attribute("innerText")
        dz = datetime.strptime(DZ[:10], "%d.%m.%Y")
        SZ = tr.find_elements(By.CSS_SELECTOR, "td")[2].get_attribute("innerText")
        LZ = tr.find_elements(By.CSS_SELECTOR, "td")[3].find_elements(By.CSS_SELECTOR, "a")
        if (d1 <= dz <= d2):
            p = p + 1
            s = f"{NZ} {DZ} {SZ}"
            if len(LZ) > 0 and (not found or ws.cell(row = r, column = 3).value != "Завершена"):
                if download:
                    LZ[0].click()
                    s = s + " :: download"
                    RQ["Завершена и загружена"] = RQ.get("Завершена и загружена", 0) + 1
                    time.sleep(1)
##                else:
##                    RQ["Завершена"] = RQ.get("Завершена", 0) + 1

            if not found:
                row = ws.max_row + 1

            ws.cell(row = row, column = 1).value = NZ
            ws.cell(row = row, column = 2).value = DZ
            ws.cell(row = row, column = 3).value = SZ
            print(s)
            RQ[SZ] = RQ.get(SZ, 0) + 1
            RQ["_____Всего"] = RQ.get("_____Всего", 0) + 1
        
    wb.save(filename = 'dn.xlsx')
    return (d1 <= dz <= d2) or (d1 <= d2 <= dz)

##------------------------------------------------------------##
def dnd(d, p1=1, p2=200):
    dndv(d, p1, p2, download=True)
    
def dnv(d, p1=1, p2=200):
    dndv(d, p1, p2, download=False)

def dndv(d, p1, p2, download):
    global EGRN
    global RQ
    global p
    Status = CheckDNfile()
    if Status != "FileDNOK":
        return
    EGRN.implicitly_wait(0)
    RQ = {}
    init_pyxl()
    table = EGRN.find_element(By.CSS_SELECTOR, "table.v-table-table")
    pm = EGRN.find_elements(By.CSS_SELECTOR, "div.v-horizontallayout.v-horizontallayout-pageManagement.pageManagement > div > div")
    first_page = pm[1].find_element(By.XPATH, "div/div")
    is_first_page = first_page.get_attribute("aria-pressed") != "false"
    if not is_first_page:
        first_page.click()
    while not is_first_page:
        time.sleep(1)
        is_first_page = first_page.get_attribute("aria-pressed") != "false"
    print(".......................................Первая страница ОК")
    try:
        activepage = int(pm[3].find_element(By.CSS_SELECTOR, "div.v-horizontallayout div.v-label.v-label-undef-w").get_attribute("innerText"))
    except:
        activepage = 1
    next_page = pm[4].find_element(By.CSS_SELECTOR, "div")
    dd = d.split("-")
    if not d:
        dd = "01.01.1000-31.12.9999".split("-")
    elif len(dd) == 1:
        dd.append(dd[0])
    dd[0] = datetime.strptime(dd[0], "%d.%m.%Y")
    dd[1] = datetime.strptime(dd[1], "%d.%m.%Y")
    dd.sort()
##    print(dd[0], dd[1])

    ## skip to initial page if it is not the first one
    while activepage < p1:
        old_activepage = activepage
        next_page.click()
        while old_activepage == activepage:
            print(f'page {activepage} in range({p1}, {p2 + 1})')
            time.sleep(4)
            try:
                pm = EGRN.find_elements(By.CSS_SELECTOR, "div.v-horizontallayout.v-horizontallayout-pageManagement.pageManagement > div > div")
                activepage = int(pm[3].find_element(By.CSS_SELECTOR, "div.v-horizontallayout div.v-label.v-label-undef-w").get_attribute("innerText"))
                next_page = pm[4].find_element(By.CSS_SELECTOR, "div")
                status = True
            except:
                activepage = 9999
                status = False

    status = True
    while dn(dd[0], dd[1], download) and status and activepage in range(p1, p2):
        print(f".......................................Обработана страница {activepage} :: строк {p}")
        ## check if the current page is the final one
        if len(next_page.find_elements(By.CSS_SELECTOR, "div.v-button.v-disabled.v-button-link.link")) > 0:
            break
        old_activepage = activepage
        next_page.click()
        while old_activepage == activepage:
            time.sleep(4)
            try:
                pm = EGRN.find_elements(By.CSS_SELECTOR, "div.v-horizontallayout.v-horizontallayout-pageManagement.pageManagement > div > div")
                activepage = int(pm[3].find_element(By.CSS_SELECTOR, "div.v-horizontallayout div.v-label.v-label-undef-w").get_attribute("innerText"))
                next_page = pm[4].find_element(By.CSS_SELECTOR, "div")
                status = True
            except:
                activepage = 9999
                status = False

    print("===============================")
    print("Всё готово")
    for sz in sorted(RQ):
        print("%s: %s" % (sz, RQ[sz]))
    EGRN.implicitly_wait(default_implicitly_wait)
        
##------------------------------------------------------------##
def dndx(fn = 'rq'):
    global EGRN
    global RQ
    global p
    RQ = {}
    EGRN.implicitly_wait(0)
    if init_pyxlx(fn) != "OK": return
    ffield = ffield = EGRN.find_elements(By.XPATH, ".//input[@type='text' and contains(@class,'v-textfield')]")[0]
    fbutton = EGRN.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Обновить')]")

    for row in ws.iter_rows(min_row = 3):
        rqn = row[2].value
        rqs = row[4].value
        print("Строка {0:4d} из {1:4d} :: запрос {2} :: статус ".
            format(row[0].row - 2, ws.max_row - 2, rqn), end = "")
        if row[4].value == "Завершена":
                print("Завершена, результат был загружен ранее")
                RQ["Завершена, результат был загружен ранее"] = RQ.get("Завершена, результат был загружен ранее", 0) + 1
        else:
            if re.search(r"80-\d{0}", f"{rqn}") != None:
                ffield.clear()
                ffield.send_keys(rqn)
                fbutton.click()
                time.sleep(1)
                try:
                    td = WebDriverWait(EGRN, 15).until(EC.presence_of_element_located((By.XPATH, f".//div[contains(text(), '{rqn}')]")))
                except:
                    print("Ошибка при поиске результата запроса")
                    break
                table = EGRN.find_element(By.CSS_SELECTOR, "table.v-table-table")
                td = table.find_element(By.XPATH, f".//div[contains(text(), '{rqn}')]")
                if td != None:
                    tr = table.find_elements(By.CSS_SELECTOR, "tr")[0]
                    SZ = tr.find_elements(By.CSS_SELECTOR, "td")[2].get_attribute("innerText")
                    LZ = tr.find_elements(By.CSS_SELECTOR, "td")[3].find_elements(By.CSS_SELECTOR, "a")
                    print(SZ, end = "")
                    RQ[SZ] = RQ.get(SZ, 0) + 1
                    if len(LZ) > 0:
                        if row[4].value != "Завершена":
                            LZ[0].click()
                            print(", результат загружен")
                        else:
                            print(", результат был загружен ранее")
                    else:
                        print()
                    row[4].value = SZ

                else:
                    print("запрос по файлу выполнен, но не найден")
            else:
                print("запрос по файлу не выполнялся")
            if (row[0].row - 2) % 10 == 0:
                try:
                    wb.save(fn + ".xlsx")
                    print("Файл сохранён")
                except:
                    print("Ошибка при сохранении файла")

    try:
        wb.save(fn + ".xlsx")
    except:
        print("Ошибка при сохранении файла")
        
    print("\n\nВсе запросы из файла проверены")
    for sz in sorted(RQ): print("%s: %s" % (sz, RQ[sz]))
        
##------------------------------------------------------------##
def h():
    help()
    
def help():
    print("Доступные команды: ")
    print("1. dndx(\"ИмяФайлаСЗапросами\") - проверка запросов по списку, отработанному ботом, \n\tи загрузка результатов")
    print("\tdndx() - проверка файла rq.xlsx")
    print("\nили\n")
    print("2. dnv(\"Д1-Д2\", С1, С2) - список всех отправленных запросов будет сохранён в файл dn.xlsx")
    print("3. dnd(\"Д1-Д2\", С1, С2) - то же, что и предыдущая команда, плюс загрузка результатов")
    print("где:")
    print("\tД1-Д2 - диапазон дат \"01.03.2022-15.03.2022\" или одна дата \"27.03.2022\" или \"\"")
    print("\tС1 и С2 - диапазон номеров страниц (можно не указывать), по умолчанию - от 1 до 200")
    print("например:")
    print("\tdnv(\"21.04.2022\") - список всех запросов за дату, в пределах страниц 1-200")
    print("\tdnd(\"\", 17) - загрузка всех запросов со страницы 17 (и до 200) без уточнения даты")
    print("\tdnv(\"01.02.2021-28.02.2021\", 1, 9999) - список всех запросов за февраль 2021 года")
    print("чем дальше дата от текущей и чем больше номера страниц, тем дольше ждать перебора страниц")
    print("-" * 90)



##------------------------------------------------------------##
while True:
    Status = None
    Status = EGRNinit()
    if Status == "SiteOK":
        Status = EGRNauth()
    if Status == "AuthOK":
        Status = EGRNRequestPage()
    if Status == "NoAuthKey":
        break
    if Status == "RequestPageOK":
        break
print("Всё готово для работы")
print("\nПодсказка: введите help() или h()\n\n")
help()

