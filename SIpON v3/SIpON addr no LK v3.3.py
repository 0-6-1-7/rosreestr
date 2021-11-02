

import os, re, sys, time
import PIL

from base64 import b64decode
from io import BytesIO
from PIL import Image, ImageTk
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
    
RR = None
RunFromIDLE = None
AppWindow = None
FileKN = None
Progress = None
Status = None
lbl_captcha = None
txt_captcha = None

##------------------------------------------------------------##
##------------------------------------------------------------##
##------------------------------------------------------------##
def GetFileKNToProceed():
    global AppWindow
    FileKNfilename = askopenfilename(parent = AppWindow, title = "Выберите файл для обработки", filetypes = [("Файлы Excel", "*.xlsx")], multiple = False)
    if FileKNfilename == "": printStatus("Файл не выбран"); return
    progress = CheckFileKNToProceed(FileKNfilename)
    FileKN.set(FileKNfilename)
    printProgress(progress)
    printStatus("Файл выбран")

##------------------------------------------------------------##
def CheckFileKNToProceed(wbName):
    wb = load_workbook(filename = wbName)
    ws = wb.worksheets[0]
    rmax = ws.max_row - 1
    r = 2
    x = ws.cell(row = r , column = 2).value
    while x != None:
        r = r + 1
        x = ws.cell(row = r , column = 2).value
    return f"Всего в файле строк: {rmax}, из них обработаны: {r - 2}"
    
##------------------------------------------------------------##
def resize(event):
    global AppWindow
    SaveWindowGeometry(AppWindow.geometry())
    
##------------------------------------------------------------##
def SaveWindowGeometry(g):
    global AppWindow
    with open("app.ini", "w") as conf:
        conf.write(g) 

##------------------------------------------------------------##
def main():
    global AppWindow, FileKN, Progress, Status, lbl_captcha, txt_captcha
    
    AppWindow = Tk()
    FileKN = StringVar()
    Progress = StringVar()
    Status = StringVar()
    AppWindow.title("Росреестр - справочная информация по объектам недвижимости в режиме online")
    if os.path.isfile("app.ini"): 
        with open("app.ini", "r") as conf: 
            AppWindow.geometry(conf.read())
    else:
        AppWindow.geometry('550x250')
    AppWindow.resizable(1, 1)
    AppWindow.call('wm', 'attributes', '.', '-topmost', True)
    AppWindow.bind("<Configure>", resize)

    
    lbl_status = Label(AppWindow, textvariable = Status)
    lbl_status.grid(row = 1, column = 1, columnspan = 3)
    
    btn_1 = Button(AppWindow, text = "Начать работу", command = RRinit, width = 25).grid(row = 2, column = 1, sticky = "we")

    btn_2 = Button(AppWindow, text = "Выбрать файл: ", command = GetFileKNToProceed).grid(row = 3, column = 1, sticky = "we")
    txtFileKN = Label(AppWindow, textvariable = FileKN).grid(row = 3, column = 2, columnspan = 3, sticky = "w")
    lbl_progress = Label(AppWindow, textvariable = Progress).grid(row = 4, column = 2, columnspan = 3, sticky = "we")

    CaptchaValue = StringVar()
    txt_captcha = Entry(AppWindow, textvariable = CaptchaValue)
    txt_captcha.grid(row = 5, column = 1, sticky = "we")

    nocaptchaimage = PIL.Image.open(BytesIO(b64decode("iVBORw0KGgoAAAANSUhEUgAAAMgAAAAyCAMAAAAuj2TTAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAADUExURQAAAKd6PdoAAAABdFJOUwBA5thmAAAACXBIWXMAAA7DAAAOwwHHb6hkAAAAIElEQVRoQ+3BAQEAAACCIP+vbkhAAAAAAAAAAAAAcKgGJ0IAAT6TWzgAAAAASUVORK5CYII=")))
    img_nocaptcha = ImageTk.PhotoImage(nocaptchaimage)
    lbl_captcha = Label(AppWindow, image = img_nocaptcha)
    lbl_captcha.grid(row = 5, column = 2, sticky = "we")

    btn_captcha = Button(AppWindow, text = "Обновить капчу", command = RRrefresh).grid(row = 5, column = 3, sticky = "we")
    
    btn_3 = Button(AppWindow, text = "Обработка файла", command = RRgo, width = 15).grid(row = 6, column = 1, sticky = "we")
    btn_5 = Button(AppWindow, text = "Выход", command = RRquit).grid(row = 8, column = 3, sticky = "we")

    lbl_about = Label(AppWindow, text = "Поиск по адресу :: версия 3.3").grid(row = 9, column = 1, columnspan = 3, sticky = "w")

    Grid.columnconfigure(AppWindow, 0, minsize = 25)
    Grid.columnconfigure(AppWindow, 1, minsize = 150)
    Grid.columnconfigure(AppWindow, 2, minsize = 200)
    Grid.columnconfigure(AppWindow, 3, minsize = 150)
    Grid.columnconfigure(AppWindow, 4, minsize = 25, weight=1) 
    AppWindow.mainloop()
  

##------------------------------------------------------------##
def RRinit():
    global RR, RunFromIDLE, lbl_captcha, captchaimage
    printStatus("Запускается Chrome. Дождитесь появления капчи.")
    chrome_options = Options()
    if not RunFromIDLE: chrome_options.add_argument("--headless")
    chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\rosreestr")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    try:
        RR = webdriver.Chrome(options = chrome_options)
    except:
        sta = "ERROR_INIT_CHROME"; printStatus(sta)
        return sta
    RR.set_page_load_timeout(180)
    RR.get("https://lk.rosreestr.ru/eservices/real-estate-objects-online")
    wait = WebDriverWait(RR, 15)
    try:
        sipono = wait.until(EC.visibility_of_element_located((By.ID, "sipono-selector")))
    except:
        sta = "ERROR_INIT_START_PAGE_NOT_LOADED"; printStatus(sta)
        return sta

## первая капча никогда не срабатывает
    reloadlink = RR.find_element(By.XPATH, "//span[@class='rros-ui-lib-captcha-content-reload-btn-name']")
    reloadlink.click()
    time.sleep(1)
    
    img = ImageTk.PhotoImage(GetCaptcha())
    lbl_captcha.configure(image = img)
    lbl_captcha.image = img
    AppWindow.update()
    printStatus("Начальная страница загружена, капча получена")

##------------------------------------------------------------##
def RRquit():
    global RR
    printStatus("Программа закрывается, подождите...")
    try: RR.quit()
    except: pass
    exit()

##------------------------------------------------------------##
def printProgress(msg):
    global AppWindow, Progress
    print(msg)
    Progress.set(msg)
    AppWindow.update()

##------------------------------------------------------------##
def printStatus(msg):
    global AppWindow, Status
    txt = msg
    if msg[:5] != "ERROR": txt = msg
    if msg == "ERROR_INIT_CHROME": txt = "Ошибка загрузки Chrome - он уже запущен или другая серъёзная ошибка"
    if msg == "ERROR_INIT_START_PAGE_NOT_LOADED": txt = "Ошибка загрузки начальной страницы"
    if msg == "ERROR_CAPTCHA_WRONG": txt = "Введена неправильная капча"
    if msg == "ERROR_UNKNOWN_PAGE": txt = "Непонятная страница"
    if msg == "ERROR_CAPTCHA_NOT_ENTERED": txt = "Капча не введена"
    if msg == "ERROR_KN_FOUND_NOT_SHOWN": txt = "КН найден, но не отображается"
    if msg == "ERROR_OBJECT_NOT_SHOWN": txt = "Объект найден, но страница свойств не открылась"
    if msg == "ERROR_CAN_NOT_CLEAR_KN_FIELD": txt = "Не получается очистить поле ввода КН"
    if msg == "ERROR_CAN_NOT_CLEAR_ADDR_FIELD": txt = "Не получается очистить поле ввода КН"
    if msg == "ERROR_ADDR_NOT_FOUND": txt = "Адрес не найден"

    print(txt)
    Status.set(txt)
    AppWindow.update()

##------------------------------------------------------------##
def RRgo():
    global RR, FileKN, Progress, txt_captcha
    printStatus("...идёт обработка файла...")
    t0 = time.monotonic()
    if RR == None: printStatus("Сайт ещё не загружен"); return
## капчу надо вводить два раза два раза
    captchainput = RR.find_element(By.CSS_SELECTOR, "input.rros-ui-lib-captcha-input")
    while captchainput.get_attribute("value") != "": captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
    captchainput.send_keys(txt_captcha.get())
    while captchainput.get_attribute("value") != "": captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
    captchainput.send_keys(txt_captcha.get())

    try: #wrap for any return
    ## начало обработки файла
        wbName = FileKN.get()
        try:
            wb = load_workbook(filename = wbName)
            ws = wb.worksheets[0]
            rmax = ws.max_row - 1
        except:
            printStatus("Пробема с файлом: невозможно открыть и т.д."); return
        if rmax < 1: printStatus("Пустой файл"); return
    ## пройти мимо обработанных строк
        r = 2
        while True:
            x = ws.cell(row = r , column = 2).value
            if x == None: break
            r = r + 1
        if r - 2 == rmax:  printStatus("Файл уже полностью обработан"); return

    ## добавить заголовок, если в файле его нет 
        if r == 2:
            h = "Строка поиска\tНайдено кадастровых номеров".split("\t")
            for i in range(0, len(h)): ws.cell(row = 1, column = i + 1).value = h[i]

    ## добавить лист с результататми текущего поиска
        ws2 = wb.create_sheet("Sheet KNaddr")
        r2 = ws2.max_row + 1
        h = "Строка поиска\tКадастровый номер\tАдрес".split("\t")
        for i in range(0, len(h)): ws2.cell(row = 1, column = i + 1).value = h[i]

    ## обработка списка
        while True:
            addr = ws.cell(row = r , column = 1).value
            if addr == None: printStatus(f"Файл полностью обработан за {round(time.monotonic() - t0, 2)} сек."); return
            t = GetInfoAddr(addr)            
            if t == "ERROR_ADDR_NOT_FOUND": t = "Адрес не найден или ошибка при поиске"
            if t[:5] == "ERROR": wb.save(wbName); printStatus(t); return
            kns = t.split("\t")
            knsl = len(kns)
            ws.cell(row = r, column = 2).value = knsl
            for i in range(0, knsl):
                ka = kns[i].split("\n")
                if ka[0] == "Информация не найдена":
                    ws.cell(row = r, column = 2).value = "Информация не найдена"
                else:
                    ws2.cell(row = r2, column = 1).value = addr
                    ws2.cell(row = r2, column = 2).value = ka[1]
                    ws2.cell(row = r2, column = 3).value = ka[0]
                    r2 = r2 + 1
            if (r - 1) % 10 == 0: wb.save(wbName)
            r = r + 1
            printProgress(f"Всего в файле строк: {rmax}, из них обработаны: {min(rmax, r - 2)} за {round(time.monotonic() - t0, 2)} сек.")
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
    finally:
        printProgress("")
        wb.save(wbName)
    

##------------------------------------------------------------##
def CaptchaOK():
    global RR
    try: return RR.find_element(By.CSS_SELECTOR, "span.rros-ui-lib-message--error") == None
    except: return True

##------------------------------------------------------------##
def RRrefresh():
    global RR, lbl_captcha, captchaimage, XFactor
    try:
        reloadlink = RR.find_element(By.XPATH, "//span[@class='rros-ui-lib-captcha-content-reload-btn-name']")
        reloadlink.click()
        time.sleep(1)
        captchaimage = GetCaptcha()
        img = ImageTk.PhotoImage(captchaimage)
        lbl_captcha.configure(image = img)
        lbl_captcha.image = img
        txt_captcha.delete(0, END)
    except:
        return

##------------------------------------------------------------##
def GetCaptcha():
    global RR, Status
    try:
        captcha = RR.find_element(By.CSS_SELECTOR, "img.rros-ui-lib-captcha-content-img")
        img_base64 = RR.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 200; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", captcha)
        return HiCaptcha(PIL.Image.open(BytesIO(b64decode(img_base64))))
    except:
        printStatus("Ошибка получения капчи")
        return None

def HiCaptcha(img):
    mask = PIL.Image.open(BytesIO(b64decode("iVBORw0KGgoAAAANSUhEUgAAAMgAAAAyCAAAAAA8Oss9AAADzElEQVR4nNVZW7bcIAyzyNx9df9ranE/kgA2wnEyt6dzf3KYiWxsIYU88EsAKRAAuBoBEBQ5BnFAERdwnL3KO42SsCIigsUBbtROSgjDDIsDZtgOuJ6+wVDGID+K+kvCWH9JmMzYgIay5C7uLwljDNxgi87AYUVsUKQNV0coIdYfgbG8rLVrWMmRsqojqY04IKIha5FTWv/dIjWgIUVvuaMN3IUxBjisrMuN8nZYmXvUhTYeWERNwH5Sf8959W2L4JXXxlzupUVAAvA154WFebZW049dFaYNW+m7FqlOJRRW37TIfvn13RqioHPhLaA1X9eLvhkGRCmsLMtd5rWwQrRxTnWUC6nVJpu4q7KZ/vSADdqvIvtPON6+ZReBvEgkBKLo5WqrcmWR4paaaH9rJbAqTd4nFhl29rFbFUAan2bTjC0iJ+frXUTHub7pRksGj5hwqMjJs+t0wd0xAp19tIjgH9xoCfY7/v7/ceFH70/PwZhWh1ksDcpmp9o4t5MDpu9ZJLhaHCOwOqgi2gqT/sgML5sXdnTXIn1nD9dQIHUUyB1tOJUch7qA1YcWaR650P6+dH4WJqGwtQ7bzElto+LKjfKarkriuuFmj2B89kphPe/VlTDDVpmD/FKvNLqSkFqYKMo6IKLhjkXwst1iapQJJJLQvD2AWGRNA2ErRUOJSWGrmYQxBiwsUW6U13ZVVuG2v+hGK9MfgYlggjEakmwtntkTGk1qIw6IaLhlkXNnv9ZGYJF6RcOcV5M05Okt5pfes0j18kxapEaU6SOL9Muvm+Asd7+WLixirxTgsxNtbLrfaSlji2wIGdGU8Vfbbnu5O9vV93chIVe4g1WBfC2XmeQl2ab+yqJbNS2XyCLTAZDj2ZdbZAtoaLvptZhsV8X86nO6Z5EacUcXnTyN7bA/IQ2Yml/w5rsq8/+mXDnzE4t4WM4i2y3tc1idsf4NH+223S1JCIv6O2E6BkQ0hDDygL58nLXrfdsirL/lyRUNcV7bn32is9H1fKNV+xsDwh1/VaW+hCNg2j0Fotr18tAiKOYXWrkQwXaON0FokR5zFIQxo+VuEGkrEgrxz1q3nAS+K6OXtdBoqI2SDeiwbQh/tIuMN41DtxDW93dYJAGTGZuwSNvZ2/++K1tHEsYYuIS9ZZFDsvEuwhZ9hi37IzCWl7V2gy3ynT2h0aQ24oCIhtsWYd/ZF4fPtsj0nZ1pAz7oCsYYuMHWE4vQF2K2259hEfLM/jMtQr6zLw4fbhH/nZ1pA1PQB1rEf2e3pJByP9Ui9pk9J+WPtIj8BQNuUvD5KM4GAAAAAElFTkSuQmCC")))
    img = img.convert("L")
    i = img.load()
    m = mask.load()
    width, height = mask.size
    for y in range(height):
        for x in range(width):
            if i[x,y] < 20: i[x,y] = 255 
            elif abs(i[x,y] - m[x,y]) < 10: i[x,y] = 255
            if i[x,y] < 250: i[x,y] = 0
    return img
    
##------------------------------------------------------------##
def GetInfoAddr(addr):
    global RR
    if not CaptchaOK(): return "ERROR_CAPTCHA_WRONG"
    wait = WebDriverWait(RR, 5)
    try:
        sipono = wait.until(EC.visibility_of_element_located((By.ID, "sipono-selector")))
        ADDRinput = sipono.find_element(By.CSS_SELECTOR, "input")
    except:
        return "ERROR_UNKNOWN_PAGE"
    if ADDRinput.get_attribute("disabled"): return "ERROR_CAPTCHA_NOT_ENTERED"
    n = 1
    while ADDRinput.get_attribute("value") != "":
        ADDRinput.send_keys(Keys.END)
        ADDRinput.send_keys(Keys.BACKSPACE * (len(ADDRinput.get_attribute("value"))))
        n = n + 1
        if n > 10: return "ERROR_CAN_NOT_CLEAR_ADDR_FIELD"
## подождать пока спрячется список объектов (от предыдущего поиска?)
    ADDRinput.send_keys(Keys.ESCAPE)
    n = 1
    tmp = wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__menu")))
    ADDRinput.send_keys(addr)
    time.sleep(1)
    info = ""

    wait = WebDriverWait(RR, 10)
    n = 1
    try:
        loo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__menu")))
##        if tmp: loo = find_element(By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__menu"); break
    except:
        if not CaptchaOK(): return "ERROR_CAPTCHA_WRONG"
        return "ERROR_ADDR_NOT_FOUND"
    opts = loo.find_elements(By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__option")
    for opt in opts:
        info = info + ("" if len(info) == 0 else "\t") + opt.get_attribute("innerText")
    if len(opts) == 0: return f"Информация не найдена\n"
    return info

##########################################################################
RunFromIDLE = "idlelib" in sys.modules
main()
