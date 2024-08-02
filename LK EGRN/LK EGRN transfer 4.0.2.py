import os, sys

from configparser import ConfigParser

from re import match, search, sub

from datetime import datetime

from time import monotonic, sleep

from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import *

from openpyxl import Workbook
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

RR = None
RUN_FROM_IDLE = None
APP_WINDOW = None
FILE_NAME = None
PROGRESS = None
PROGRESSBAR = None
STATUS = None
WORK_PAUSED = None
SEARCH_TYPE = "RQ"
SEARCH_BUTTON_CAPTION = None
AUTH_TIMEOUT = 120 ## таймаут ожидания авторизации
##
##wb = None
##ws = None

# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
def get_file_to_proceed():
    filename = askopenfilename(parent=APP_WINDOW,
                               title="Выберите файл для обработки",
                               filetypes=[("Файлы Excel", "*.xlsx")], multiple=False)
    if filename == "":
        FILE_NAME.set("Файл не выбран")
        print_status("Файл не выбран")
        return
    s = filename
    r = s.rfind("/")
    if r > 0: s = s[:r] + "\n" + s[r + 1:]
    FILE_NAME.set(s)
    PROGRESS = CheckRQfile(filename)
    print(f"PROGRESS = {PROGRESS}")
    if PROGRESS[2] is None:
        print_status("Файл выбран", False)
        print_progress(total=PROGRESS[0], done=PROGRESS[1], message=PROGRESS[2], win_update=False)
    else:
        print_status(PROGRESS[2], False)
        print_progress(total=0, done=0, message="", win_update=False)
    if SEARCH_TYPE == "RQ":
        SEARCH_BUTTON_CAPTION.set("Запросы выписок")
    else:
        SEARCH_BUTTON_CAPTION.set("...")
    APP_WINDOW.update()

# ------------------------------------------------------------ #
def CheckRQfile(wbName):
##    global wb, ws
##    global KNs, Reg
##    global KNsToDo
##    global defaultBatchMax, defaultBatchSize, BatchMax

    RQStatus = None
    ## Проверим наличие и доступность файла
    try:
        wb = load_workbook(filename = wbName)
        ws = wb.worksheets[0]
    except:
        return [None, None, "RQ_NO_FILE"]
    try:
        wb.save(filename = wbName)
    except:
        return [None, None, "RQ_FILE_SAVE_ERROR"]

    ## Проверим структуру файла
    Reg = ws.cell(row = 1, column = 1).value # Регион: во-первых, должен быть не пустой. Во-вторых - совпадать со списком
    if Reg == None:
        return "RQ_NO_REGION"
    else:
        RegOK = False
        for RegRow in wb["Список регионов"].iter_rows(1): RegOK = RegOK or RegRow[0].value == Reg
        if not RegOK:
            return [None, None, "RQ_WRONG_REGION"]

    ## Проверим количество строк в файле
    if ws.max_row < 3:
        return [None, None, "RQ_NO_DATA"]

    ## Проверим все КН в списке на правильность по маске
    for row in ws.iter_rows(min_row = 3):
        kn = row[0].value
        if search(r"\[\d\d:\d\d:\d{1,7}:\d{1,}\]", f"[{kn}]") != None:
            row[1].value = "да"
        else:
            row[1].value = "нет"
            RQStatus = "RQ_INCORRECT_DATA"

    ## Проверим  список на дубликаты КН
    KNList = []
    DupKNList = []
    for row in ws.iter_rows(min_row = 3):
        if row[0].value in KNList:
            DupKNList.append(row[0].row)
            row[1].value = "дубль"
        else: KNList.append(row[0].value)
    if DupKNList != []:
        print(f"Найдены дубли кадастровых номеров")
        RQStatus = "RQ_DUPLICATE_DATA"

    ## Сохраним файл с результатами проверки
    wb.save(filename = wbName)

    ## Возврат если к этому моменту выявили ошибку
    if RQStatus is not None:
        return [None, None, RQStatus]

    ## Посчитаем обработанные и необработанные КН
    rqs = 0
    for row in ws.iter_rows(min_row = 3):
        if row[3].value: rqs += 1
    return [ws.max_row - 2, rqs, None]
##--- end of CheckRQfile ##------------------------------------------------------------##

# ------------------------------------------------------------ #
def app_resize(event):
    with open("app.ini", "w") as conf:
        conf.write(APP_WINDOW.geometry())

# ------------------------------------------------------------ #
def print_progress(total=0, done=0, done_batch=0, duration=0, message=None, win_update=True):
    if message is not None:
        PROGRESS.set(message)
        print(message)
        APP_WINDOW.update()
        return
    msg = f"Всего в файле строк: {total}, из них обработано: {done}"
    if done_batch > 0:
        msg = msg + f"\nв т.ч. {done_batch}" + \
            (f" за {round(duration, 2)} сек." if duration > 0 else "") + \
            (f" ~{int(3600 / (duration / done_batch))} строк в час" if duration > 0 else "")
    print(msg)
    PROGRESS.set(msg)
    if msg != "":
        PROGRESSBAR['value'] = int(done / total * 100)
        PROGRESSBAR.grid()
    else:
        PROGRESSBAR.grid_remove()
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def print_status(msg, win_update=True):
    txt = msg
    if msg[:5] == "FATAL":
        if msg == "FATAL_ERROR_PAGE_UNKNOWN":
            txt = "Непонятная страница"
        if msg == "FATAL_ERROR_PAGE_NOT_FULLY_LOADED":
            txt = "Страница загружена не полностью"
        if msg == "FATAL_ERROR_INIT_CHROME":
            txt = "Ошибка загрузки Chrome - он уже запущен или другая серъёзная ошибка"
        if msg == "FATAL_ERROR_CHROME":
            txt = "Chrome не запущен или закрыт"
        if msg == "FATAL_ERROR_INIT_START_PAGE_NOT_LOADED":
            txt = "Ошибка загрузки начальной страницы"
    elif msg[:5] == "ERROR":
        if msg == "ERROR_TIMEOUT":
            txt = "За 60 секунд не активировалась кнопка поиска"
        if msg == "ERROR_KN_FOUND_NOT_SHOWN":
            txt = "КН найден, но не отображается"
        if msg == "ERROR_SEARCH_FAILED":
            txt = "Поиск не завершился каким-либо результатом"
        if msg == "ERROR_OBJECT_NOT_SHOWN":
            txt = "Объект найден, но страница свойств не открылась"
        if msg == "ERROR_CAN_NOT_CLEAR_KN_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_CAN_NOT_CLEAR_ADDR_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_ADDR_NOT_FOUND":
            txt = "Адрес не найден"
        if msg == "FATAL_ERROR_SEARCH_BUTTON_IS_NOT_ACTIVE":
            txt = "Кнопка поиска на форме не активна"
    elif msg[:5] == "NOTIF":
        txt = msg.split("$")[1]
    elif msg[:2] == "RQ":
        if msg == "RQ_NO_FILE":
            txt = "Нет файла rq.xlsx"
        if msg == "RQ_FILE_SAVE_ERROR":
            txt = "Ошибка при сохранении файла rq.xlsx - вероятно, он открыт в Excel"
        if msg == "RQ_NO_REGION":
            txt = "В файле не указан регион"
        if msg == "RQ_WRONG_REGION":
            txt = "В файле неправильно указан регион"
        if msg == "RQ_NO_DATA":
            txt = "В файле слишком мало данных"
        if msg == "RQ_INCORRECT_DATA":
            txt = "Неправильный кадастровый номер"
        if msg == "RQ_DUPLICATE_DATA":
            txt = "Найдены дубли кадастровых номеров"
    elif msg[:4] == "ESIA":
        if msg == "ESIA_CANT_READ_ESIA_FILE":
            txt = "Ошибка при чтении файоа с паролем от ЛК"
        if msg == "ESIA_PAGE_HAS_NO_FORM":
            txt = "Не загружена страница авторизации"
        if msg == "ESIA_INFO_NO_PERSONAL_CABINET":
            txt = "Нет страницы личного кабинета"
        if msg == "ESIA_NO_ROLE_SELECTED":
            txt = "Не выбрана нужная роль ЛК"
    else: txt = msg
    print(txt)
    STATUS.set(txt)
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def main():
    global APP_WINDOW, FILE_NAME, PROGRESS, STATUS, SEARCH_BUTTON_CAPTION, PROGRESSBAR
    APP_WINDOW = Tk()
    FILE_NAME = StringVar()
    PROGRESS = StringVar()
    STATUS = StringVar()
    SEARCH_BUTTON_CAPTION = StringVar(value="...")

    APP_WINDOW.title("Росреестр - Запрос посредством доступа к ФГИС ЕГРН")
    if os.path.isfile("app.ini"):
        with open("app.ini", "r") as conf:
            APP_WINDOW.geometry(conf.read())
    else:
        APP_WINDOW.geometry("600x300")
    APP_WINDOW.resizable(1, 1)
    APP_WINDOW.call("wm", "attributes", ".", "-topmost", True)
    APP_WINDOW.bind("<Configure>", app_resize)

    lbl_status = Label(APP_WINDOW, textvariable=STATUS)
    lbl_status.grid(row=1, column=1, columnspan=3)

    btn_1 = Button(APP_WINDOW, text="Запустить Chrome", command=RRinit, width=25)
    btn_1.grid(row=2, column=1, sticky="we")

    btn_2 = Button(APP_WINDOW, text="Выбрать файл: ", command=get_file_to_proceed)
    btn_2.grid(row=3, column=1, sticky="we")

    txt_filename = Label(APP_WINDOW, textvariable=FILE_NAME)
    txt_filename.grid(row=3, column=2, columnspan=3, sticky="w")

    lbl_progress = Label(APP_WINDOW, textvariable=PROGRESS)
    lbl_progress.grid(row=4, column=2, columnspan=3, sticky="we")

    PROGRESSBAR = Progressbar(APP_WINDOW, orient='horizontal', mode='determinate')
    PROGRESSBAR.grid(row=5, column=2, columnspan=2, sticky="we")
    PROGRESSBAR['value'] = 0
    PROGRESSBAR.grid_remove()

    btn_3 = Button(APP_WINDOW, textvariable=SEARCH_BUTTON_CAPTION, command=RRgo, width=15)
    btn_3.grid(row=7, column=1, sticky="we")

    btn_4 = Button(APP_WINDOW, text="Пауза", command=RRpause, width=15)
    btn_4.grid(row=8, column=1, sticky="we")

    btn_5 = Button(APP_WINDOW, text="Выход", command=RRquit)
    btn_5.grid(row=9, column=3, sticky="we")

    lbl_about = Label(APP_WINDOW, text="Запрос посредством доступа к ФГИС ЕГРН :: версия 0.1")
    lbl_about.grid(row=10, column=1, columnspan=3, sticky="w")

    Grid.columnconfigure(APP_WINDOW, 0, minsize=25)
    Grid.columnconfigure(APP_WINDOW, 1, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 2, minsize=200)
    Grid.columnconfigure(APP_WINDOW, 3, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 4, minsize=25, weight=1)
    APP_WINDOW.mainloop()


# ------------------------------------------------------------ #
def get_with_ESIA_auth(RR, url):
    auth = ConfigParser()
    try:
        auth.read("esia.ini", 'UTF-8')
        esia_login = auth.get("ESIA", "login")
        esia_password = auth.get("ESIA", "password")
        esia_role_name = auth.get("ESIA", "name", fallback=None)
    except:
        return "ESIA_CANT_READ_ESIA_FILE"

    print_status("Выполняется авторизация в ЕСИА")
    RR.get(url)
    sleep(1)

##    t0 = monotonic()
##    while url.lower() in RR.current_url.lower():
##        if monotonic() - t0 > AUTH_TIMEOUT:
##            return "FATAL_ESIA_ERROR"
##        sleep(1)
##
##    while True:
##        if "esia.gosuslugi.ru/login/" in RR.current_url.lower():
##            break
##        sleep(1)

    t0 = monotonic()
    while "esia.gosuslugi.ru/login/" not in RR.current_url.lower():
        if monotonic() - t0 > AUTH_TIMEOUT:
            return "FATAL_ESIA_ERROR"
        # если авторизация уже действует (случайный перезапуск?), просто выходим
        if "lk.rosreestr.ru/eservices" in RR.current_url.lower():
            return "OK"
        sleep(1)

    try:
        form_container = RR.find_element(By.CSS_SELECTOR, "div.form-container")
        form_login = form_container.find_element(By.ID, "login")
        form_password = form_container.find_element(By.ID, "password")
        form_button = form_container.find_element(By.XPATH, "//button[contains(text(),'Войти')]")
    except:
        return "ESIA_PAGE_HAS_NO_FORM"

    form_login.send_keys(esia_login)
    form_password.send_keys(esia_password)
    form_button.click()
    sleep(1)

    # ожидание загрузки страницы авторизации может быть очень долгим
    while True:
        if url.lower() in RR.current_url.lower():
            break
        sleep(1)

    try:
        wait = WebDriverWait(RR, AUTH_TIMEOUT)
        personal_cabinet_root = wait.until(
            EC.presence_of_element_located((By.ID, "personal-cabinet-root")))
    except:
        return "ESIA_INFO_NO_PERSONAL_CABINET"

    # если нужно, выбираем роль пользователя
    if esia_role_name is not None:
        time0 = monotonic()
        while True:
            # не дождались выбора роли, на выход
            if monotonic() - time0 > AUTH_TIMEOUT:
                return "ESIA_NO_ROLE_SELECTED"
            try:
    ##            roles_selector = personal_cabinet_root.find_element(
    ##                By.CSS_SELECTOR, "div.roles-selector")
                role = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//*[contains(text(),'" + esia_role_name + "')]")))
                role.click()
                break
            except:
                sleep(1)

    return "OK"


# ------------------------------------------------------------ #
def RRinit():
    global RR
    try:
        print_status("Запускается Chrome.")
        if RR is None:
            chrome_options = Options()
    ##        chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\rosreestr")
            chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            RR = webdriver.Chrome(options=chrome_options)
            RR.set_page_load_timeout(180)

        status = get_with_ESIA_auth(RR, "https://lk.rosreestr.ru/login")
        if status == "OK":
            print_status("Загружается личный кабинет")
            ## переход на нужную страницу ЛК
            url = "https://lk.rosreestr.ru/request-access-egrn"
            RR.get(url)
            
            ## ждём загрузки страницы ЕГРН
            t0 = monotonic()
            while True:
                ## не дождались появления нужной страницы, на выход с ошибкой
                if monotonic() - t0 > AUTH_TIMEOUT:
                    print_status("Ошибка: страница ЕГРН не загружена")
                    return
                try:
                    titles = RR.find_elements(By.CSS_SELECTOR, "div.egrn-list__title")
                    egrn_ready = False
                    for t in titles:
                        egrn_ready = egrn_ready or ("Поиск объектов" in t.get_attribute("innerText"))
                    if egrn_ready:
                        break
                except:
                    sleep(1)

            # переход на страницу поиска объекта
            url = "https://lk.rosreestr.ru/request-access-egrn/property-search"
            RR.get(url)

            ## ждём загрузки страницы поиска
            t0 = monotonic()
            while True:
                ## не дождались появления нужной страницы, на выход с ошибкой
                if monotonic() - t0 > AUTH_TIMEOUT:
                    print_status("Ошибка: страница поиска не загружена")
                    return

                ## ждём появления строки поиска как признак полной загрузки страницы
                try:
                    search_input = RR.find_elements(By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__input")
                    break
                except:
                    sleep(1)

            print_status("Личный кабинет готов к работе")

        else:
            print_status(status)

    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")

# ------------------------------------------------------------ #
def RRquit():
    global RR
    print_status("Программа закрывается, подождите...")
    try: RR.quit()
    except: pass
    exit()

# ------------------------------------------------------------ #
def RRpause():
    global WORK_PAUSED
    WORK_PAUSED = True
    print_status("...пауза...")

# ------------------------------------------------------------ #
def RRgo():
    print_status("...идёт обработка файла...")

    if RR is None:
        print_status("Сайт ещё не загружен")
        return

    ## ожидание загрузки страницы - через долгую проверку авторизации
    t0 = monotonic()
    wait = WebDriverWait(RR, 5)

    while True:
        ## не дождались появления нужной страницы, на выход с ошибкой
        if monotonic() - t0 > AUTH_TIMEOUT:
            print_status("FATAL_ERROR_INIT_START_PAGE_NOT_LOADED")
            return

        url = "https://lk.rosreestr.ru/request-access-egrn/property-search"
        if not url.lower() in RR.current_url.lower():
            RR.get(url)

        ## ждём появления строки поиска
        try:
            search_input = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__input")))
        except:
            search_input = None
        if search_input is not None:
            ## дождались, можно продолжать
            break 

    return_status = None    
    while True:
        return_status = RRgoRQ()

        if return_status in "FATAL;OK;ERROR":
            break

# ------------------------------------------------------------ #
def RRgoRQ():
    global WORK_PAUSED
    prev_info = None
    WORK_PAUSED = False
    t0 = monotonic(); backup_time_mark = t0; save_time_mark = t0

## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Проблема с файлом: невозможно открыть и т.д."); return

    rows_total = ws.max_row - 2
    if rows_total < 1: print_status("Пустой файл"); return

## пропустить обработанные строки
    rows_done = 0
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=3).value is not None:
            rows_done += 1
        else:
            break            
    
    if rows_done == rows_total:  print_status("Файл уже полностью обработан"); return "OK"


    try:  # глобальная обработка исключений
        result = None
        row = rows_done + 3

        timeout_counter = 0 # начальная очка для отсчёта таймаута
## основной цикл обработки
        while row <= ws.max_row:
            if WORK_PAUSED: return

## отработать таймаут 5 секунд
            while monotonic() - timeout_counter < 5:
                    sleep(1)
            timeout_counter = monotonic()
            
            prev_result = result
            kn = ws.cell(row=row , column=1).value

            try:
                clear_button = RR.find_element(By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__clear-indicator")
                clear_button.click()
            except:
                pass
            
            search_input = RR.find_element(By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__input>input")
            search_input.send_keys(kn)

            # ожидание появления таблицы с результатами поиска
            t0 = monotonic()
            while True:
                if monotonic() - t0 > AUTH_TIMEOUT:
                    print_status("Ошибка: поиск не завершился за 1 минуту")
                    return "ERROR"
                try:
                    result = RR.find_element(By.CSS_SELECTOR, "div.realestateobjects-wrapper__results")
                    if prev_result != result:
                        prev_result = result
                        break
                except:
                    sleep(1)

            # ожидание нужного КН или его отсутствия в результатах поиска
            t0 = monotonic()
            while True:
                if monotonic() - t0 > AUTH_TIMEOUT:
                    print_status("Ошибка: поиск не завершился за 1 минуту")
                    return "ERROR"
                try:
                     results_count_wrapper = RR.find_element(By.CSS_SELECTOR, "span.realestateobjects-wrapper__results__total-count")
                     results_count_text = results_count_wrapper.get_attribute("innerText")
                     results_count = int(results_count_text[results_count_text.rfind(" "):])
                     break
                except:
                    sleep(1)

            if results_count == 0:
                print_status("Ошибка: объект не найден по кадастровому номеру")
                return "ERROR"
            elif results_count > 1:
                print_status("Ошибка: найдено более одного объекта")
                return "ERROR"

            # проверим результат поиска
            if kn not in result.get_attribute("innerText"):
                print_status("Ошибка найден иной объект")
                return "ERROR"

## отработать таймаут 5 секунд
##            t0 = monotonic()
            while monotonic() - timeout_counter < 5:
                    sleep(1)
                    print("sleep(1)")
            timeout_counter = monotonic()
            
            # странное поведение - вводим ";" после КН - открывается карточка объекта
            search_input.send_keys(";")

             # ожидание появления окна с информацией об объекте
            t0 = monotonic()
            while True:
                if monotonic() - t0 > AUTH_TIMEOUT:
                    print_status("Ошибка: объект найден, но не отображается")
                    return "ERROR"
                try:
                    rom = RR.find_element(By.CSS_SELECTOR, "div.realestate-object-modal")
                    # кнопка "Назад" как признак окончания загрузки окна с информацией
                    back_button = rom.find_element(By.CSS_SELECTOR, "button.rros-ui-lib-button--reverse")
                    break
                except:
                    sleep(1)

################################################################################
################################################################################
            menu = RR.find_element(By.CSS_SELECTOR,
                "div.rros-ui-lib-dropdown-menu > div > button")
            menu.click()
            sleep(1)
            menu_item = RR.find_element(By.CSS_SELECTOR,
                "div.rros-ui-lib-dropdown-menu__menu-el:nth-child(2) > div > div > div")
            menu_item.click()
################################################################################
################################################################################

            # ожидание результата отправки
            # ожидание может быть долгим, особенно в случае ошибки
            t0 = monotonic()
            while True:
                if monotonic() - t0 > 120:
                    try:
                        notification = RR.find_element(By.CSS_SELECTOR, "div.notifications > div")
                        notification_text = notification.get_attribute("innerText")
                        print_status("Сообщение системы: " +  notification_text)
                        return "ERROR"
                    except:
                        print_status("Ошибка: запрос не завершился ни успехом, ни ошибкой")
                        return "ERROR"
                try:
                    success = RR.find_element(By.ID, "success-view")
                    break
                except:
                    sleep(1)

            # сохранение результата
            ws.cell(row=row, column=3).value = "запрос отправлен"
            ws.cell(row=row, column=4).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb.save(filename = wbName)
            
            # возврат на предыдущую страницу                    
            RR.back()

            rows_done += 1
            print_progress(total=rows_total,
                          done=rows_done)
            row = row + 1

        print_status("Файл полностью обработан")
        return "OK"
            
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    finally:
        if not WORK_PAUSED: print_progress(message="")
        

# ------------------------------------------------------------ #
def check_notifications(RR):
    wait_1 = WebDriverWait(RR, 1)
    try: notifications = wait_1.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.notifications")))
    except: return

    errors = notifications.find_elements(By.CSS_SELECTOR, "div.rros-ui-lib-errors > div")
    notifications_text.get_attribute("innerText")
    
    if errors:
        print_status(notifications_text.replace("\n", "/"))

# ------------------------------------------------------------ #

##########################################################################
RUN_FROM_IDLE = "idlelib" in sys.modules
main()
