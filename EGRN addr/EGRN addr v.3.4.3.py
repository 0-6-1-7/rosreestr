import datetime, os, re, sys, time

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

EGRN = None
AUTH_KEY = ""  # ключ доступа
RUN_FROM_IDLE = None  # True если скрипт запущен из IDLE
PREV_DATA = [None] * 7


##--- end of header ---##------------------------------------------------------------##

def DisplayErrorMessage(text):
    ts = text.split("\n")
    maxlen = 40
    for t in ts: maxlen = max(maxlen, len(t))
    print("╓─" + "─" * maxlen + "─╖")
    print("║ " + "{val:{wid}}".format(wid=maxlen, val=ts[0]) + " ║")
    if len(ts) > 1: print("╟─" + "─" * maxlen + "─╢")
    for t in range(1, len(ts)): print("║ " + "{val:{wid}}".format(wid=maxlen, val=ts[t]) + " ║")
    print("╙─" + "─" * maxlen + "─╜")


##--- end of DisplayErrorMessage ---##------------------------------------------------------------##


def wb_save_30_sec(wb, wbName, save_time_mark):
    if time.monotonic() - save_time_mark > 30:
        wb.save(wbName)
        return time.monotonic()
    else:
        return save_time_mark


# ------------------------------------------------------------ #


def EGRN_init(msg=""):
    global EGRN
    SiteRestartRetriesMax = 10  # количество попыток перезапуска сайта
    SiteRestartRetriesCounter = 0
    plt = 20  # page load timeout

    if msg != "": print(msg, "\n")
    # На случай если мы попали сюда после какой-нибудь ошибки - нужно начать заново
    if EGRN != None:
        EGRN.close(); EGRN.quit()
    while True:
        try:
            chrome_options = Options()
            chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            EGRN = webdriver.Chrome(options=chrome_options)
            EGRN.implicitly_wait(3)
            DemoKey = None
            while True:
                EGRN.set_page_load_timeout(plt)
                try: EGRN.get("https://rosreestr.ru/wps/portal/p/cc_present/ir_egrn")
                except: EGRN.execute_script("window.stop();")
                try: DemoKey = WebDriverWait(EGRN, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'6F9619FF-8B86-D011-B42D-00CF4FC964FF')]")))
                except: plt = plt + 10
                if DemoKey != None: print("Сайт Росреестра работает, страница авторизации загружена нормально."); return("SiteOK")
                if plt > 30: break # слишком больщой таймаут; скорее всего, сайт не работает
                print("Сайт работает медленнее, чем хотелось бы...")            
        except: pass
        SiteRestartRetriesCounter = SiteRestartRetriesCounter + 1
        print(f"Страница не загружена. Перезагрузка, попытка № {SiteRestartRetriesCounter}")
        EGRN.close(); EGRN.quit()
        if SiteRestartRetriesCounter > SiteRestartRetriesMax:
            print("\n\n\n\nСайт Росреестра не работает. Мы пытались...")
            return("SiteFailed")


##--- end of EGRNinit ##------------------------------------------------------------##
        
def EGRN_auth():
    global AUTH_KEY
    AuthStatus = None
    try:
        f = open("auth.txt", "r")
        AUTH_KEY = f.readline()[:36]
        f.close()
        if re.search("\w{8}-\w{4}-\w{4}-\w{4}-\w{12}", AUTH_KEY) != None: AuthStatus = "GotAuthKey"
        else: return("NoAuthKey")
    except: return("NoAuthKey")
        
    if AuthStatus == "GotAuthKey":
        vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")
        AuthKeyFields = vapp.find_elements(By.XPATH, ".//input[@type='text']")
        AuthKeyFields[0].send_keys(AUTH_KEY)
        t0 = time.monotonic()
        while True:
            if re.search(AuthKeyFields[4].get_attribute("value"), AUTH_KEY) != None:
                AuthStatus = "KeyEntered"
                break
            if time.monotonic() - t0 > 60:
                print("За 60 секунд ключ не принят.")
                return("AuthTimeOut")
    if AuthStatus == "KeyEntered":
        SendButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Войти')]")
        SendButton.click()
        print("Авторизация выполнена.")
        return("AuthOK")
##--- end of EGRNauth ##------------------------------------------------------------##

def EGRN_search_page():
    vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")
    t0 = time.monotonic()
    while True:
        try:
            button_search = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Поиск объектов недвижимости')]")
            break
        except: pass
        if time.monotonic() - t0 > 60: return("SearchPageFailed") # За 60 секунд страница не загрузилась полностью
    button_search.click()
    try:
        InfoImage = vapp.find_element(By.XPATH, ".//img[contains(@src, 'i_icon.gif')]")
        print("Страница поиска готова.")
        return("SearchPageOK")
    except: return("SearchPageFailed")
##--- end of EGRNSearchPage ##------------------------------------------------------------##

def field_filterselect(field, field_data, wait_while_disabled=False):
    if not field_data and field.get_attribute("disabled"): return "OK"  # недоступное поле - уже пустое
        
    EGRN.execute_script(f"arguments[0].value='';", field)
    wait = WebDriverWait(EGRN, 30)

    loops = 0
    while wait_while_disabled:
        if not field.get_attribute("disabled"): break
        time.sleep(1); loops += 1
        if loops > 10: return "SearchError"

    field.click()
    if field_data: field.send_keys(field_data)
    else: field.send_keys(Keys.BACKSPACE)

    try: suggestpopup = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.v-filterselect-suggestpopup")))
    except: return "SearchError"

    loops = 0
    while True:
        suggest = suggestpopup.find_element(By.CSS_SELECTOR, "tr")
        suggest_text = suggest.get_attribute("innerText")
        if field_data.upper() in suggest_text.upper(): break  # строка найдена
        time.sleep(1); loops += 1
        if loops > 10: return "SearchError"

    suggest.click()

    return "OK"
##--- end of field_filterselect ##------------------------------------------------------------##

def field_text(field, field_data):
    field.clear()
    if not field_data: return "OK"
    field.send_keys(field_data)
    return "OK"

def strip(s):
    if s is None: return ""
    return s.strip()

def EGRN_search(data):
    global PREV_DATA
    try:
        vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")

        # Проверим, не висит ли эта ошибка с прошлого поиска
        NotificationWarningText = None
        try:
            NotificationWarning = EGRN.find_element(By.CLASS_NAME, "v-Notification-warning")
            NotificationWarningText = NotificationWarning.get_attribute("innerText")
            if "Не найдены данные, удовлетворяющие Вашему запросу." in NotificationWarningText:
                NotificationWarning.click()
                while NotificationWarningText is not None:
                    try:
                        NotificationWarning = EGRN.find_element(By.CLASS_NAME, "v-Notification-warning")
                        NotificationWarningText = NotificationWarning.get_attribute("innerText")
                        time.sleep(1)
                        print("wait...")
                    except:
                        NotificationWarningText = None

        except: pass

        button_change = None
        try:
            button_change = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Изменить')]")
        except: pass
        if button_change is not None:
            button_change.click()
        button_search = None
        try:
            
            f_f_x = vapp.find_elements(By.XPATH, ".//input[contains(@class,'v-filterselect-input')]")
            f1 = f_f_x[0]; f1_d = data[0] # Регион
            f2 = f_f_x[1]; f2_d = data[1] # Район
            f3 = f_f_x[2]; f3_d = data[2] # Населенный пункт
            f4 = f_f_x[3]; f4_d = data[3] # Улица
            f_t_x = vapp.find_elements(By.XPATH, ".//input[contains(@class,'v-textfield')]")
            f5 = f_t_x[1]; f5_d = data[4] # Название улицы
            f6 = f_t_x[2]; f6_d = data[5] # Дом
            f7 = f_t_x[3]; f7_d = data[6] # Квартира

            if f1_d != PREV_DATA[0]:
                sta = field_filterselect(f1, f1_d)
                if sta == "OK": PREV_DATA[0] = f1_d
                else: return sta
            if f2_d != PREV_DATA[1]:
                sta = field_filterselect(f2, f2_d, wait_while_disabled=True)
                if sta == "OK": PREV_DATA[1] = f2_d
                else: return sta
            if f3_d != PREV_DATA[2]:
                sta = field_filterselect(f3, f3_d, wait_while_disabled=True)
                if sta == "OK": PREV_DATA[2] = f3_d
                else: return sta
            if f4_d != PREV_DATA[3]:
                sta = field_filterselect(f4, f4_d)
                if sta == "OK": PREV_DATA[3] = f4_d
                else: return sta
            if f5_d != PREV_DATA[4]:
                sta = field_text(f5, f5_d)
                if sta == "OK": PREV_DATA[4] = f5_d
                else: return sta
            if f6_d != PREV_DATA[5]:
                sta = field_text(f6, f6_d)
                if sta == "OK": PREV_DATA[5] = f6_d
                else: return sta
            if f7_d != PREV_DATA[6]:
                sta = field_text(f7, f7_d)
                if sta == "OK": PREV_DATA[6] = f7_d
                else: return sta

            button_search = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Найти')]")
            print("Поиск подготовлен")
            button_search.click()
            print("Поиск запущен")

        except:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(f"При выполнении возникло исключение {exc_type}\n"
                f"\tописание:\t{exc_obj}\n"
                f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")

            pass # ошибка при заполнении форма поиска, начать поиск заново
            return "SearchError"    

        EGRN.implicitly_wait(1)
        print("Ожидание реультата поиска")
        t0 = time.monotonic()
        jjj = 0
        while True:
            jjj = jjj + 1
            data_table = None
            try: data_table = vapp.find_element(By.CSS_SELECTOR, "table.v-table-table")
            except: pass
            NotificationWarning = None
            try:
                NotificationWarningText = EGRN.find_element(By.CLASS_NAME, "v-Notification-warning").get_attribute("innerText")
                print(NotificationWarningText)
                if "Не найдены данные, удовлетворяющие Вашему запросу." in NotificationWarningText:
                    DisplayErrorMessage(NotificationWarningText)
                    return "Информация не найдена"
            except: pass
            if data_table == None and NotificationWarning == None:
                time.sleep(1)
                if time.monotonic() - t0 > 60:
                    print("За 60 секунд не дождались нужной страницы. Вероятно, ошибка сайта.")
                    return "SearchTimeoutError"
            elif data_table != None:
                time.sleep(1)
                t = EGRN.find_element(By.CSS_SELECTOR, "div.blockSmall").get_attribute("innerText")
                if t == "Найдено более 200 записей. Необходимо уточнить параметры запроса.":
                    print("Найдено более 200 объектов")
                    return "Найдено более 200 объектов"
                results = []
                print("Поиск удачный")
                for row in data_table.find_elements(By.CSS_SELECTOR, "tbody > tr"):
                    result = [row.find_elements(By.CSS_SELECTOR, "td")[i].get_attribute("innerText") for i in range(7)]
                    S = result[3].split(" ")
                    try:
                        result[3] = float(S[0])
                    except:
                        result[3] = S[0]
                    try:
                        result.insert(4, S[1])
                    except:
                        print("\n\n\nОшибка при вычислении площади\n\n\n")
                        result.insert(4, "")
                        print(f"S = {S}")
                        pass
                    results.append(result)
                
                return results
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
        return "ERROR"

##--- end of EGRNSearch ##------------------------------------------------------------##

def text_of_cell(c):
    if c is None:
        return ""
    return strip(f"{c}")


def EGRN_go_search(f):
    t0 = time.monotonic(); backup_time_mark = t0; save_time_mark = t0
    
    if EGRN is None:
        print("Сайт ещё не загружен");
        if EGRN_preload() not in "OK":
            print("Ошибка при загрузке сайта")
            return

    if f == "RESTART":
        print("Перезагрузка после фатальной ошибки");
        EGRN_preload()
        return
        
    color = {True: "00CCFFCC", False: "00CCCCFF"}; color_index = True

## начало обработки файла
    wbName = f
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print("Пробема с файлом: невозможно открыть и т.д."); return

    rows_total = ws.max_row - 1
    if rows_total < 1: print("Пустой файл"); return

## лист со списком адресов
## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=10).value is None: row = row - 1; break
    rows_done = row - 1
    if rows_done == rows_total:  print("Файл уже полностью обработан"); return

## лист с результататми текущего поиска
    try: ws2 = wb["KNaddr#results"]
    except: ws2 = wb.create_sheet("KNaddr#results")
    color_index = ws2.cell(row=ws2.max_row, column=1).fill.fgColor.rgb != "FFCCFFCC"
    row2 = ws2.max_row + 1
    h = "ЛС (id)\t"\
        "Кадастровый номер\t"\
        "Полный адрес объекта\t"\
        "Тип объекта\t"\
        "Площадь\t"\
        "Ед. изм. площади\t"\
        "Категория ЗУ\t"\
        "Вид разрешенного использования ЗУ\t"\
        "Назначение здания или помещения".split("\t")
    for i in range(0, len(h)): ws2.cell(row=1, column=i + 1).value = h[i]
    try:  # глобальная обработка исключений
## основной цикл обработки
        row1 = row + 1
        while True:
            print(f"\nСтрока {row1 - 1} из {rows_total}: ")
            data = [text_of_cell(ws.cell(row=row1,column=i).value) for i in range(3,10)]
            if data[0] == "":
                print(f"Файл {f} полностью обработан за {round(time.monotonic() - t0, 2)} сек.")
                wb.save(wbName)
                return "DONE"
            if ws.cell(row=row1, column=10).value is not None:
                print("\tзапрос был выполнен ранее")
                row1 += 1
                continue
                
            objects_info = EGRN_search(data)
            if objects_info == "ERROR": return "ERROR"

## на выходе либо ошибка
            if objects_info[:5] in ("ERROR", "FATAL", "NOTIF"):
                save_time_mark = wb_save_30_sec(wb, wbName, save_time_mark); print(objects_info)
                return

## либо массив с результатами
## каждый элемент массива - массив из 8 элементов
            if objects_info in ("Информация не найдена", "Найдено более 200 объектов"):
                ws.cell(row=row1, column=10).value = objects_info
            elif objects_info in ("SearchTimeoutError", "SearchError"):
                print("Какая-то ошибка при заполнении формы поиска")
                print("Повтор поиска")
                row -= 1
                return "ERROR"
            else:
                ws.cell(row=row1, column=10).value = len(objects_info)
                for obj in objects_info: # остальные данные
                    print(f"\tobj = {obj}")
                    ws2.cell(row=row2, column=1).value = ws.cell(row=row1, column=1).value # ЛС (id)
                    for i in range(8):
                        ws2.cell(row=row2, column=2 + i).value = obj[i]
                        for col in range(1, 10):
                            ws2.cell(row=row2, column=col).fill = PatternFill("solid", fgColor=color[color_index])
                    row2 += 1
            
            color_index = not color_index
            
            if (row1 - 1) % 10 == 0: save_time_mark = wb_save_30_sec(wb, wbName, save_time_mark)
            if time.monotonic() - backup_time_mark > 600:  # backup every 10 minutes
                try: wb.save(wbName + ".backup")
                except: pass
                finally: backup_time_mark = time.monotonic()

            row1 += 1

    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"111При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
        try: wb.save(wbName)
        except: pass
        finally: backup_time_mark = time.monotonic()
    
##------------------------------------------------------------##
def EGRN_preload():
    while True:
        Status = None
        while True:
            Status = EGRN_init("Начало работы с сайтом")
            if Status == "SiteOK": Status = EGRN_auth()
            if Status == "AuthOK": Status = EGRN_search_page()
            if Status == "NoAuhtKey": print("Нет ключа доступа. Дальнейшая работа без ключа невозможна."); quit()
            if Status == "SearchPageOK": return "OK"
            if Status == "SearchError": print("Перезапуск для обработки очередной части списка"); break
            if Status == "AllDone": break
            if Status == "SearchKNsNotFound":
                print("Ни один из запрошенных КН не найден во ФГИС ЕГРН.")
                print("Перезапуск для обработки очередной части списка")
                break
            else: print("Перезапуск после фатальной ошибки"); break
        else: print("\n")
        if Status == "AllDone": break




def go(f):
    global PREV_DATA
    if not f.endswith(".xlsx"):
        f = f + ".xlsx"
    try:
        while True:
            result = EGRN_go_search(f)
            if result == "DONE":
                break
            else:
                PREV_DATA = [None] * 7
                EGRN_go_search("RESTART")

    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\t\tстрока:\t{exc_tb.tb_lineno}")

    print("Всё готово, можно закрыть окно")

## main ##########################
RUN_FROM_IDLE = "idlelib" in sys.modules

if not RUN_FROM_IDLE:
    if len(sys.argv) == 2: # запуск с параметром (drag'n'drop в проводнике), только 1 файл
        f = sys.argv[1]
        go(f)
    else:
        print("Для запуска поиска перетащите в проводнике файл со списком адресов на скрипт.")
        time.sleep(30)
else:
    print("Для запуска поиска:")
    print("\tgo(\"ИмяФайла\") - расширение .xlsx указывать не нужно")
