from gatherInfo2 import Convert

import os, re, sys, time
import PIL

from base64 import b64decode
from io import BytesIO
from PIL import Image, ImageTk
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
    
RR = None
sessionid = ""
RunFromIDLE = None
AppWindow = None
FileKN = None
FileProgressTxt = None
Status = None
lbl_captcha = None
txt_captcha = None


##captchaimage = None
##------------------------------------------------------------##
##------------------------------------------------------------##
##------------------------------------------------------------##
def GetFileKNToProceed():
    global AppWindow
    AppWindow.update()
    FileKNfilename = askopenfilename(parent = AppWindow, title = "Выберите файл для обработки", filetypes = [("Файлы Excel", "*.xlsx")], multiple = False)
    print(FileKNfilename)
    progress = CheckFileKNToProceed(FileKNfilename)
    FileKN.set(FileKNfilename)
    FileProgressTxt.set(progress)

##------------------------------------------------------------##
def CheckFileKNToProceed(wbName):
    wb = load_workbook(filename = wbName)
    ws = wb.worksheets[0]
    rmax = ws.max_row - 1
    r = 2
    kn = ws.cell(row = r , column = 2).value
    while kn != None:
        r = r + 1
        kn = ws.cell(row = r , column = 2).value
    return f"Всего в файле строк: {rmax}, из них обработаны: {r - 2}"
    
##------------------------------------------------------------##
def SaveWindowGeometry(event):
    global AppWindow
    with open("app.ini", "w") as conf:
        conf.write(AppWindow.geometry()) 

##------------------------------------------------------------##
def main():
##    global AppWindow, FileKN, FileProgressTxt, Status, lbl_captcha
    global AppWindow, FileKN, FileProgressTxt, Status, lbl_captcha, txt_captcha
    
    AppWindow = Tk()
    FileKN = StringVar()
    FileProgressTxt = StringVar()
    Status = StringVar()
    AppWindow.title("Росреестр - справочная информация по объектам недвижимости в режиме online")
    if os.path.isfile("app.ini"): 
        with open("app.ini", "r") as conf: 
            AppWindow.geometry(conf.read())
    else:
        AppWindow.geometry('600x250')
##    AppWindow.geometry('600x250')
    AppWindow.resizable(1, 1)
    AppWindow.call('wm', 'attributes', '.', '-topmost', True)
    AppWindow.bind("<Configure>",SaveWindowGeometry)

    
    lbl_status = Label(AppWindow, textvariable = Status)
    lbl_status.grid(row = 1, column = 1, columnspan = 3)
    
    btn_1 = Button(AppWindow, text = "Начать работу", command = RRinit).grid(row = 2, column = 1, sticky = "w")

    btn_2 = Button(AppWindow, text = "Выбрать файл: ", command = GetFileKNToProceed).grid(row = 3, column = 1, sticky = "w")
    txtFileKN = Label(AppWindow, textvariable = FileKN).grid(row = 3, column = 2, columnspan = 3, sticky = "w")
    lbl_progress = Label(AppWindow, textvariable = FileProgressTxt).grid(row = 4, column = 2, columnspan = 3, sticky = "w")

    CaptchaValue = StringVar()
    txt_captcha = Entry(AppWindow, textvariable = CaptchaValue)
    txt_captcha.grid(row = 5, column = 1, sticky = "w")

    nocaptcha = PIL.Image.open(BytesIO(b64decode("iVBORw0KGgoAAAANSUhEUgAAAMgAAAAyCAMAAAAuj2TTAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAADUExURQAAAKd6PdoAAAABdFJOUwBA5thmAAAACXBIWXMAAA7DAAAOwwHHb6hkAAAAIElEQVRoQ+3BAQEAAACCIP+vbkhAAAAAAAAAAAAAcKgGJ0IAAT6TWzgAAAAASUVORK5CYII=")))
    img_nocaptcha = ImageTk.PhotoImage(nocaptcha)
    lbl_captcha = Label(AppWindow, image = img_nocaptcha)
    lbl_captcha.grid(row = 5, column = 2, sticky = "w")

    btn_captcha = Button(AppWindow, text = "Обновить капчу", command = RRrefresh).grid(row = 5, column = 3, sticky = "w")
    
    btn_3 = Button(AppWindow, text = "Обработка файла", command = RRgo).grid(row = 6, column = 1, sticky = "w")
    btn_4 = Button(AppWindow, text = "Остановить обработку", command = RRquit).grid(row = 7, column = 1, sticky = "w")
    btn_5 = Button(AppWindow, text = "Выход", command = RRquit).grid(row = 8, column = 3, sticky = "e")


    Grid.columnconfigure(AppWindow, 0, minsize = 25)
    Grid.columnconfigure(AppWindow, 1, minsize = 150)
    Grid.columnconfigure(AppWindow, 2, minsize = 200)
    Grid.columnconfigure(AppWindow, 4, minsize = 50, weight=1)
    Grid.columnconfigure(AppWindow, 5, minsize = 25)
    AppWindow.mainloop()
  

##------------------------------------------------------------##
def RRinit():
    global RR, Status, RunFromIDLE, lbl_captcha

    Status.set("")
    chrome_options = Options()
    if not RunFromIDLE: chrome_options.add_argument("--headless")
    chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\rosreestr")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    try:
        RR = webdriver.Chrome(options = chrome_options)
    except:
        print("Ошибка запуска Chrome, возможно, он уже запущен")
        Status.set("Ошибка запуска Chrome, возможно, он уже запущен")
        return "error 1"
    RR.set_page_load_timeout(180)
    RR.get("https://lk.rosreestr.ru/eservices/real-estate-objects-online")
    wait = WebDriverWait(RR, 15)
    try:
        sipono = wait.until(EC.visibility_of_element_located((By.ID, "sipono-selector")))
    except:
        print("Ошибка загрузки начальной страницы")
        Status.set("Ошибка загрузки начальной страницы")
        return "error 1"

## первая капча никогда не срабатывает
    reloadlink = RR.find_element(By.XPATH, "//span[@class='rros-ui-lib-captcha-content-reload-btn-name']")
    reloadlink.click()
    time.sleep(1)
    
    img = ImageTk.PhotoImage(GetCaptcha())
    lbl_captcha.configure(image = img)
    lbl_captcha.image = img
    print("Начальная страница загружена, капча получена")
    Status.set("Начальная страница загружена, капча получена")

##------------------------------------------------------------##
def RRquit():
    try: RR.quit()
    except: pass
    exit()

##------------------------------------------------------------##
def RRgo():
    global RR, Status, FileKN, FileProgressTxt, txt_captcha
    
    captchainput = RR.find_element(By.CSS_SELECTOR, "input.rros-ui-lib-captcha-input")
    while captchainput.get_attribute("value") != "": captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
    captchainput.send_keys(txt_captcha.get())
    while captchainput.get_attribute("value") != "": captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
    captchainput.send_keys(txt_captcha.get())
####################
    wbName = FileKN.get()
    wb = load_workbook(filename = wbName)
    ws = wb.worksheets[0]
    rmax = ws.max_row - 1
    r = 2
    kn = ws.cell(row = r , column = 2).value
    while kn != None:
        r = r + 1
        kn = ws.cell(row = r , column = 2).value
####################
    if r == 2:
        t = "Кадастровый номер\tСтатус объекта\tДата обновления информации\tВид объекта недвижимости\tПлощадь\tЕд. изм.\tАдрес\tНазначение или вид разрешенного использования\tПрава и ограничения"
        tl = t.split("\t")
        for l in range(0, len(tl)): ws.cell(row = 1, column = l + 1).value = tl[l]
    kn = ws.cell(row = r , column = 1).value
    while kn != None:
        t0 = time.monotonic()
        print(f"-- [ {r - 1} из {rmax} ] -- [ {kn} ] --")
        t = GetInfo(kn)
        if t == "ERROR_CAPTCHA":
            wb.save(wbName)
            print("Ошибка капчи")
            Status.set("Ошибка капчи")
            break
        if t == "ERROR_NOT_FOUND":
            t = "КН не найден или ошибка при поиске"
        if t[:5] == "ERROR":
            wb.save(wbName)
            print(f"Ошибка {t}")
            Status.set(f"Ошибка {t}")
            break
        
        t = kn + "\t" + t
        tl = t.split("\t")
        for l in range(0, len(tl)):
            if l == 4 and len(tl[l]) > 0: ws.cell(row = r, column = l + 1).value = float(tl[l])
            else: ws.cell(row = r, column = l + 1).value = tl[l]
        if (r - 1) % 10 == 0: wb.save(wbName)
        r = r + 1
        kn = ws.cell(row = r, column = 1).value
        print(f"Всего в файле строк: {rmax}, из них обработаны: {r - 2}")
        FileProgressTxt.set(f"Всего в файле строк: {rmax}, из них обработаны: {r - 2}")
        Status.set("...идёт обработка файла...")
        print(f"Время обрабоки строки: {round(time.monotonic() - t0, 2)} сек.")
        AppWindow.update()

    wb.save(wbName)
    print(f"\n\n\n\n\nВсё готово! ---------- [ {time.strftime('%H:%M:%S', time.localtime())} ] ")
    print()
    

##------------------------------------------------------------##
def CaptchaOK():
    global RR
    return RR.find_element(By.XPATH, "//span[@class='rros-ui-lib-message--error']") == None

##------------------------------------------------------------##
def RRrefresh():
    global RR, lbl_captcha
    print("RRcaptcha")
    try:
        reloadlink = RR.find_element(By.XPATH, "//span[@class='rros-ui-lib-captcha-content-reload-btn-name']")
        reloadlink.click()
        time.sleep(1)
        img = ImageTk.PhotoImage(GetCaptcha())
        lbl_captcha.configure(image = img)
        lbl_captcha.image = img
        txt_captcha.delete(0, END)
    except:
        return
##------------------------------------------------------------##
def GetCaptcha():
    global RR, Status
    try:
        captcha = RR.find_element(By.CSS_SELECTOR, "img.rros-ui-lib-captcha-content-img")
        img_base64 = RR.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 200; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", captcha)
        return PIL.Image.open(BytesIO(b64decode(img_base64)))
    except:
        Status.set("Ошибка получения капчи")
        print("Ошибка получения капчи")
        return None
    
##------------------------------------------------------------##
def GetInfo(KN):
    global RR
    if not CaptchaOK: return "ERROR_CAPTCHA"
    wait = WebDriverWait(RR, 5)
    while True:
        try:
            sipono = wait.until(EC.visibility_of_element_located((By.ID, "sipono-selector")))
            KNinput = sipono.find_element(By.CSS_SELECTOR, "input")
        except:
            return "ERROR1"
        if KNinput.get_attribute("disabled"): return "ERROR_CAPTCHA"
        sipono.click()
        while KNinput.get_attribute("value") != "": KNinput.send_keys(Keys.BACKSPACE * len(KNinput.get_attribute("value")))
        if KNinput.get_attribute("value") == "": break
        time.sleep(1)
    KNinput.send_keys(KN)
    time.sleep(1)
    
    wait = WebDriverWait(RR, 1)
    n = 1
    while True:
        try:
            loo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__menu")))
            time.sleep(1)
            if KN in loo.get_attribute("innerText"):
                break
            else:
                print(KN)
                print(loo.get_attribute("innerText"))
                n = n + 1
                print(f"попытка {n}")
        except:
            n = n + 1
            pass
        if n > 4:
            print("не удалось получить нужный объект")
            return "ERROR_NOT_FOUND"
        time.sleep(1)
    loo.click()

    wait = WebDriverWait(RR, 5)
    try:
        result = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.realestateobjects-wrapper__results__cadNumber")))
        result.click()
    except:
        return "ERROR4"

    try:
        reverselink = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "button.rros-ui-lib-button--reverse")))
        objectinfo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.realestate-object-modal        ")))
    except:
        return "ERROR5"

    info = objectinfo.get_attribute("innerText")
    info = Convert(info)
    RR.execute_script("arguments[0].click();", reverselink)
    return info

##########################################################################
RunFromIDLE = "idlelib" in sys.modules
main()
