from gatherInfo2 import Convert

import os, re, sys, time
import inspect

from PIL import Image, ImageTk
from tkinter import *
from tkinter.ttk import *

from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
    
RR = None
sessionid = ""
RunFromIDLE = None
window = None
##------------------------------------------------------------##
##------------------------------------------------------------##
##------------------------------------------------------------##
def save_size(event):
    global window
    with open("app.conf", "w") as conf:
        conf.write(window.geometry()) 

##------------------------------------------------------------##
def main():
    global window
    window = Tk()
    window.title("Росреестр - справочная информация по объектам недвижимости в режиме online")
    if os.path.isfile("app.conf"): 
        with open("app.conf", "r") as conf: 
            window.geometry(conf.read())
    else:
        window.geometry('200x200')

    
    window.resizable(0, 0)
##  window.call('wm', 'iconphoto', window._w, ImageTk.PhotoImage(file='logoP.png'))
    window.call('wm', 'attributes', '.', '-topmost', True)

    go1 = Button(window, text = "Авторизация", command = RRinit).grid(column = 1, row = 1)
    go2 = Button(window, text = "Работа", command = RRgo).grid(column = 1, row = 2)
    go2 = Button(window, text = "Выход", command = RRquit).grid(column = 1, row = 4)
    window.bind("<Configure>",save_size)

    window.mainloop()
  

##------------------------------------------------------------##
def RRinit():
    global RR, RunFromIDLE
    chrome_options = Options()
##    if not RunFromIDLE: chrome_options.add_argument("--headless")
    chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\rosreestr")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    RR = webdriver.Chrome(options = chrome_options)
    RR.set_page_load_timeout(180)
    print("Авторизация")    
    RR.get("https://lk.rosreestr.ru/login")
##    time.sleep(10)

##------------------------------------------------------------##
def RRquit():
    try: RR.quit()
    except: pass
    exit()
##------------------------------------------------------------##
def RRgo():
    global RR
        
    if Step1() == "error": return
    wbName = 'KN.xlsx'
    if len(sys.argv) > 1: wbName = sys.argv[1]
    else: wbName = 'KN.xlsx'

    p = os.path.split(os.path.dirname(os.path.abspath(__file__)))[1] + "\\" + wbName

    from openpyxl import load_workbook
    wb = load_workbook(filename = wbName)
    ws1 = wb.worksheets[0]
    ws2 = wb.create_sheet("Sheet KN")

    r = 2 # строка в исходном списке
    rmax = ws1.max_row - 1 # всего строк
    t = "Кадастровый номер\tДата обновления информации\tВид объекта недвижимости\tСтатус объекта\tПлощадь\tЕд. изм.\tАдрес\tНазначение или вид разрешенного использования\tПрава и ограничения"
    tl = t.split("\t")
    for l in range(0, len(tl)): ws2.cell(row = 1, column = l + 1).value = tl[l]
    kn = ws1.cell(row = r , column = 1).value

    while kn != None:
        t0 = time.monotonic()
        print(f"-- [ {r - 1} из {rmax} ] -- [ {kn} ] --")
        t = kn + "\t" + GetInfo(kn)
        tl = t.split("\t")
        for l in range(0, len(tl)):
            if l == 4 and len(tl[l]) > 0: ws2.cell(row = r, column = l + 1).value = float(tl[l])
            else: ws2.cell(row = r, column = l + 1).value = tl[l]
        if (r - 1) % 10 == 0: wb.save(wbName)
        r = r + 1
        kn = ws1.cell(row = r, column = 1).value
        print(f"Время обрабоки строки: {round(time.monotonic() - t0, 2)} сек.")

    wb.save(wbName)
    print(f"\n\n\n\n\nВсё готово! ----- [ {p} ] ----- [ {time.strftime('%H:%M:%S', time.localtime())} ] ")
    print()
    

def Step1():
    global RR
    print("Начало работы с сайтом")
    wait = WebDriverWait(RR, 5)

    try:
        sidebarmenu = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "ul.sidebar__list > li:nth-child(2)")))
    except:
        return "error"
    if re.match(".*sidebar__item-active.*", sidebarmenu.get_attribute("class")) == None:
        sidebarmenuitem = sidebarmenu.find_element(By.CSS_SELECTOR, "a")
        RR.execute_script("arguments[0].click();", sidebarmenuitem)

    try:
        eservice = wait.until(EC.presence_of_element_located((By.XPATH, "//A[@data-test-id='real-estate-objects-online']")))
        RR.execute_script("arguments[0].click();", eservice)
    except:
        pass
    return "OK"

##------------------------------------------------------------##
def GetInfo(KN):
    global RR
    wait = WebDriverWait(RR, 5)
    while True:
        try:
            sipono = wait.until(EC.visibility_of_element_located((By.ID, "sipono-selector")))
            KNinput = sipono.find_element(By.CSS_SELECTOR, "input")
        except:
            return "error 1"
        
        try:
            clearbutton = sipono.visibility_of_element_located(By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__clear-indicator")
            clearbutton.click()
        except:
            while KNinput.get_attribute("value") != "": KNinput.send_keys(Keys.BACKSPACE * len(KNinput.get_attribute("value")))
            pass
        KNinput.clear()
        if KNinput.get_attribute("value") == "": break
        time.sleep(1)
    KNinput.send_keys(KN)
    time.sleep(1)
    
    wait = WebDriverWait(RR, 15)
    while True:
        try:
            loo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.rros-ui-lib-dropdown__menu")))
            if KN in loo.get_attribute("innerText"):
                break
            else:
                print("Ошибка, нужно ввести КН ещё раз")
                time.sleep(1)
        except:
            return "error 3"
    loo.click()

    wait = WebDriverWait(RR, 5)
    try:
        result = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.realestateobjects-wrapper__results__cadNumber")))
        result.click()
    except:
        return "error 4"

    try:
        reverselink = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "button.rros-ui-lib-button--reverse")))
        objectinfo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.realestate-object-modal        ")))
    except:
        return "error 5"

    info = objectinfo.get_attribute("innerText")
    info = Convert(info)
    RR.execute_script("arguments[0].click();", reverselink)
    return info













##########################################################################
RunFromIDLE = "idlelib" in sys.modules
main()
