from gatherInfo import Convert
from recognize import recognize

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import openpyxl, os, time

SIpON = None
status = 0 # состояние сайта
  
##------------------------------------------------------------##
def SIpONinit():
  global SIpON
  chrome_options = Options()  
  chrome_options.add_argument("--headless")
  SIpON = webdriver.Chrome(options = chrome_options)
  SIpON.implicitly_wait(2)
  SIpON.set_page_load_timeout(30)
  try:
    SIpON.get("https://rosreestr.gov.ru/wps/portal/online_request")
  except:
    print("Сайт Росреестра не работает")
    return 999

##------------------------------------------------------------##
def SIpONrestart(t):
  global SIpON
  if SIpON != None:
    SIpON.close()
    SIpON.quit()
  print(f"Restart")
  SIpONinit()

##------------------------------------------------------------##
def RecognizeCaptcha(c):
  global SIpON
  #code from https://gist.github.com/spirkaa/4c3b8ad8fd34324bd307#gistcomment-3157744
  img_base64 = SIpON.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", c)
  captcha = recognize(img_base64)
  return captcha
  
##------------------------------------------------------------##
def GetInfo(KN):
  global SIpON
  i = 1
  status = 1
  while status != 0 and i < 10:
    status = 0
    if i > 1:
      SIpONrestart(i)
    if status == 0:
      try:
        SetSearchCritVisible = SIpON.find_element_by_css_selector("a[href*='SetSearchCritVisible']")
        SetSearchCritVisible.click()
      except:
        pass
      try:
        captchaImage2 = SIpON.find_element_by_id("captchaImage2")
      except:
        status = 999 # капчи нет, вероятно, сайт не работает
    if status == 0:    
      try:
        captcha = RecognizeCaptcha(captchaImage2)
      except:
        status = 999 # капча не проявилась
        pass
    if status == 0:
      try:
        SIpON.execute_script("arguments[0].click();", SIpON.find_element_by_id("cad-number"))
        cadnum = SIpON.find_element_by_css_selector("input[type='text'][name='cad_num']")
        SIpON.execute_script("arguments[0].click();", cadnum)
        cadnum.clear()
        cadnum.send_keys(KN)
        captchaText = SIpON.find_element_by_css_selector("input[type='text'][name='captchaText']")
        SIpON.execute_script("arguments[0].click();", captchaText)
        captchaText.send_keys(captcha)
        SIpON.execute_script("arguments[0].click();", SIpON.find_element_by_id("submit-button"))
      except:
        status = 999
    if status == 0:
      try:
        link = SIpON.find_element_by_css_selector("a[href*='object_data_id']").get_attribute("href")
      except:
        status = 998
    if status == 998:
      try:
        infomsg = SIpON.find_element_by_css_selector("td.infomsg1")
        if infomsg != None:
          return f"{KN}\tОбъект не найден"
      except:
        status = 999
    if status == 0:
      mainwindow = SIpON.current_window_handle
      SIpON.execute_script("window.open('" + link + "');")
      SIpON.switch_to.window(SIpON.window_handles[1])  
      try:
        r_enc = SIpON.find_element_by_id("r_enc")
        if "none" in r_enc.get_attribute("style"):
          sw_r_enc = SIpON.find_element_by_id("sw_r_enc")
          SIpON.execute_script("arguments[0].scrollIntoView();", sw_r_enc)
          sw_r_enc.click()
      except:
        pass
      try:
        s_notes = SIpON.find_element_by_id("s_notes")
        if "none" in s_notes.get_attribute("style"):
          sw_s_notes = SIpON.find_element_by_id("sw_s_notes")
          sw_s_notes.click()
      except:
        pass
      try:
        content = SIpON.find_element_by_css_selector("div.portlet-body").get_attribute("innerText")
        t = Convert(content)
      except:
        status = 999
      finally:
        SIpON.close()
        SIpON.switch_to.window(mainwindow)  
    if status == 999:
      i = i + 1
  else:
    if status == 999:
      return f"{KN}\tНе удалось выполнить за 10 попыток"
    elif status == 0:
      return t # OK

##------------------------------------------------------------##
SIpONinit()

p = os.path.split(os.path.dirname(os.path.abspath(__file__)))[1]

from openpyxl import load_workbook
wb = load_workbook(filename = 'KN.xlsx')
ws1 = wb.worksheets[0]
ws2 = wb.create_sheet("Sheet KN")

r = 2 # строка в исходном списке
rmax = ws1.max_row - 1 # всего строк
t = "Лицевой счёт\tКадастровый номер\tКатегория объекта\tУчтённый\tАдрес\tТип объекта\tФорма собств.\tДата последнего перехода права собств."
tl = t.split("\t")
for l in range(0, len(tl)):
  ws2.cell(row = 1, column = l + 1).value = tl[l]
cc = ws1.cell(row = r , column = 1).value
ls = ws1.cell(row = r , column = 2).value

while cc != None:
  t = GetInfo(cc)
  print(f"----- [ {r - 1} из {rmax} ] ----- [ {p} ] ----------------------------------------------")
  print(t)
  tl = t.split("\t")
  ws2.cell(row = r, column = 1).value = ls
  for l in range(0, len(tl)):
    ws2.cell(row = r, column = l + 2).value = tl[l]
##    print(f"{l}: {tl[l - 1]}")
  if (r - 1) % 10 == 0:
    wb.save("KN.xlsx")
  r = r + 1
  cc = ws1.cell(row = r, column = 1).value
  ls = ws1.cell(row = r, column = 2).value

wb.save("KN.xlsx")
print(f"\n\n\n\n\nВсё готово! ----- [ {p} ] ----- [ {time.strftime('%H:%M:%S', time.localtime())} ] ")
print()
