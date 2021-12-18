from get_info import get_info_kn, get_info_addr
from clear_address import clear_address_for_search

import os, re, sys, time
import PIL

from base64 import b64decode
from io import BytesIO
from PIL import Image, ImageTk
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

RR = None
RUN_FROM_IDLE = None
APP_WINDOW = None
FILE_NAME = None
PROGRESS = None
PROGRESSBAR = None
STATUS = None
WORK_PAUSED = None
SEARCH_TYPE = None
SEARCH_BUTTON_CAPTION = None


# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
def get_file_to_proceed():
    filename = askopenfilename(parent=APP_WINDOW,
                               title="Выберите файл для обработки",
                               filetypes=[("Файлы Excel", "*.xlsx")], multiple=False)
    if filename == "":
        FILE_NAME.set("Файл не выбран")
        print_status("Файл не выбран")
        return
    s = filename
    r = s.rfind("/")
    if r > 0: s = s[:r] + "\n" + s[r + 1:]
    FILE_NAME.set(s)
    PROGRESS = check_file_to_proceed(filename)
    print_progress(total=PROGRESS[0], done=PROGRESS[1], message=PROGRESS[2], win_update=False)
    print_status("Файл выбран", False)
    if SEARCH_TYPE == "addr":
        SEARCH_BUTTON_CAPTION.set("Поиск по адресам")
    elif SEARCH_TYPE == "KN":
        SEARCH_BUTTON_CAPTION.set("Поиск по кад. номерам")
    else:
        SEARCH_BUTTON_CAPTION.set("...")
    APP_WINDOW.update()

# ------------------------------------------------------------ #
def check_file_to_proceed(wbName):
    global SEARCH_TYPE
    try:
        wb = load_workbook(filename=wbName)
        ws = wb.worksheets[0]
        rmax = ws.max_row
    except:
        return [None, None, "Проблема с файлом"]

## пустой файл
    if rmax == 1: SEARCH_TYPE = None; return [None, None, "Файл пустой"]

## в файле есть данные, определить тип файла - адреса или кадастровые номера
    try: v = re.match(r"^\d{1,2}:\d{1,2}:\d{6,7}:\d{1,10}$", ws.cell(row=2, column=2).value)
    except: SEARCH_TYPE = None; return [None, None, "Проблема с файлом"]

    if v is None: SEARCH_TYPE = "addr"
    else: SEARCH_TYPE = "KN"

## количество обработанных строк
    for row in range(2, rmax + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break

    if row == rmax: SEARCH_TYPE = None; return [None, None, "Файл полностью обработан"]
    else: return [rmax - 1, row - 1, None]

# ------------------------------------------------------------ #
def app_resize(event):
    with open("app.ini", "w") as conf:
        conf.write(APP_WINDOW.geometry())

# ------------------------------------------------------------ #
def print_progress(total=0, done=0, done_batch=0, duration=0, message=None, win_update=True):
    if message is not None:
        PROGRESS.set(message)
        print(message)
        APP_WINDOW.update()
        return
    msg = f"Всего в файле строк: {total}, из них обработано: {done}"
    if done_batch > 0:
        msg = msg + f"\nв т.ч. {done_batch}" + \
            (f" за {round(duration, 2)} сек." if duration > 0 else "") + \
            (f" ~{int(3600 / (duration / done_batch))} строк в час" if duration > 0 else "")
    print(msg)
    PROGRESS.set(msg)
    if msg != "":
        PROGRESSBAR['value'] = int(done / total * 100)
        PROGRESSBAR.grid()
    else:
        PROGRESSBAR.grid_remove()
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def print_status(msg, win_update=True):
    txt = msg
    if msg[:5] == "FATAL":
        if msg == "FATAL_ERROR_UNKNOWN_PAGE":
            txt = "Непонятная страница"
        if msg == "FATAL_ERROR_INIT_CHROME":
            txt = "Ошибка загрузки Chrome - он уже запущен или другая серъёзная ошибка"
        if msg == "FATAL_ERROR_CHROME":
            txt = "Chrome не запущен или закрыт"
        if msg == "FATAL_ERROR_INIT_START_PAGE_NOT_LOADED":
            txt = "Ошибка загрузки начальной страницы"
    elif msg[:5] == "ERROR":
        if msg == "ERROR_KN_FOUND_NOT_SHOWN":
            txt = "КН найден, но не отображается"
        if msg == "ERROR_OBJECT_NOT_SHOWN":
            txt = "Объект найден, но страница свойств не открылась"
        if msg == "ERROR_CAN_NOT_CLEAR_KN_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_CAN_NOT_CLEAR_ADDR_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_ADDR_NOT_FOUND":
            txt = "Адрес не найден"
    elif msg[:5] == "NOTIF":
        txt = msg.split("$")[1]
    else: txt = msg
    print(txt)
    STATUS.set(txt)
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def main():
    global APP_WINDOW, FILE_NAME, PROGRESS, STATUS, SEARCH_BUTTON_CAPTION, PROGRESSBAR
    APP_WINDOW = Tk()
    FILE_NAME = StringVar()
    PROGRESS = StringVar()
    STATUS = StringVar()
    SEARCH_BUTTON_CAPTION = StringVar(value="...")

    APP_WINDOW.title("Росреестр - справочная информация по объектам недвижимости в режиме online")
    if os.path.isfile("app.ini"):
        with open("app.ini", "r") as conf:
            APP_WINDOW.geometry(conf.read())
    else:
        APP_WINDOW.geometry("600x300")
    APP_WINDOW.resizable(1, 1)
    APP_WINDOW.call("wm", "attributes", ".", "-topmost", True)
    APP_WINDOW.bind("<Configure>", app_resize)

    lbl_status = Label(APP_WINDOW, textvariable=STATUS)
    lbl_status.grid(row=1, column=1, columnspan=3)

    btn_1 = Button(APP_WINDOW, text="Запустить Chrome", command=RRinit, width=25)
    btn_1.grid(row=2, column=1, sticky="we")

    btn_2 = Button(APP_WINDOW, text="Выбрать файл: ", command=get_file_to_proceed)
    btn_2.grid(row=3, column=1, sticky="we")

    txt_filename = Label(APP_WINDOW, textvariable=FILE_NAME)
    txt_filename.grid(row=3, column=2, columnspan=3, sticky="w")

    lbl_progress = Label(APP_WINDOW, textvariable=PROGRESS)
    lbl_progress.grid(row=4, column=2, columnspan=3, sticky="we")

    PROGRESSBAR = Progressbar(APP_WINDOW, orient='horizontal', mode='determinate')
    PROGRESSBAR.grid(row=5, column=2, columnspan=2, sticky="we")
    PROGRESSBAR['value'] = 0
    PROGRESSBAR.grid_remove()

    btn_3 = Button(APP_WINDOW, textvariable=SEARCH_BUTTON_CAPTION, command=RRgo, width=15)
    btn_3.grid(row=7, column=1, sticky="we")

    btn_4 = Button(APP_WINDOW, text="Пауза", command=RRpause, width=15)
    btn_4.grid(row=8, column=1, sticky="we")

    btn_5 = Button(APP_WINDOW, text="Выход", command=RRquit)
    btn_5.grid(row=9, column=3, sticky="we")

    lbl_about = Label(APP_WINDOW, text="Поиск по адресам и кадастровым номерам :: версия 4.0")
    lbl_about.grid(row=10, column=1, columnspan=3, sticky="w")

    Grid.columnconfigure(APP_WINDOW, 0, minsize=25)
    Grid.columnconfigure(APP_WINDOW, 1, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 2, minsize=200)
    Grid.columnconfigure(APP_WINDOW, 3, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 4, minsize=25, weight=1)
    APP_WINDOW.mainloop()

# ------------------------------------------------------------ #
def RRinit():
    global RR
    try:
        print_status("Запускается Chrome.")
        chrome_options = Options()
        chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\rosreestr")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        RR = webdriver.Chrome(options=chrome_options)
        RR.set_page_load_timeout(180)
        RR.get("https://lk.rosreestr.ru/login")
        print_status("Начальная страница загружена. Авторизуйтесь в ЛК.")
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")

# ------------------------------------------------------------ #
def RRquit():
    global RR
    print_status("Программа закрывается, подождите...")
    try: RR.close()
    except: pass
    try: RR.quit()
    except: pass
    exit()

# ------------------------------------------------------------ #
def RRpause():
    global WORK_PAUSED
    WORK_PAUSED = True
    print_status("...пауза...")

# ------------------------------------------------------------ #
def RRgo():
    print_status("...идёт обработка файла...")
    if RR is None: print_status("Сайт ещё не загружен"); return
    if not "real-estate-objects-online".lower() in RR.current_url.lower():
        print(1)
        RR.get("https://lk.rosreestr.ru/eservices/real-estate-objects-online")
    else:
        print(2)
    if SEARCH_TYPE == "KN": RRgoKN()
    elif SEARCH_TYPE == "addr": RRgoAddr()

# ------------------------------------------------------------ #
def RRgoKN():
    global WORK_PAUSED
    prev_info = None
    WORK_PAUSED = False
    t0 = time.monotonic(); tb = t0
    color = {True: "00CCFFCC", False: "00CCCCFF"}; color_index = True

## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Проблема с файлом: невозможно открыть и т.д."); return

    rows_total=ws.max_row - 1
    if rows_total < 1: print_status("Пустой файл"); return

## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break
    rows_done=row - 1
    if rows_done == rows_total:  print_status("Файл уже полностью обработан"); return

## добавить заголовок, если в файле его нет
    if row < 2:
        h = "Статус объекта\t"\
            "Дата обновления информации\t"\
            "Вид объекта недвижимости\t"\
            "Площадь\t"\
            "Ед. изм.\t"\
            "Адрес\t"\
            "Назначение или вид разрешенного использования\t"\
            "Дата последенего перехода права".split("\t")
        for i in range(0, len(h)): ws.cell(row=1, column=i + 3).value = h[i]

    try:  # глобальная обработка исключений
## основной цикл обработки
        row = row + 1
        while True:
            if WORK_PAUSED: return
            if ws.cell(row=row, column=2).value is None:
                print_status(f"Файл полностью обработан за {round(time.monotonic() - t0, 2)} сек.")
                print_progress(message="")
                break
## обработка порциями по несколько строк
            Multi = 5
            Dict_of_KNs = dict()
            if row + Multi - 1 > rows_total + 1: Multi = rows_total + 1 - row + 1
            range_of_KNs = ws[f"B{row}": f"J{row + Multi - 1}"]
            for r in range_of_KNs:
                kn = r[0].value
                if kn is None: continue
                kn = kn.replace(" ", "")
                r[0].value = kn
                if kn == "": continue
                if re.match(r"^\d{1,2}:\d{1,2}:\d{6,7}:\d{1,10}$", kn) is None:
                    r[1].value = "это не кадастровый номер"
                else:
                    Dict_of_KNs[kn] = ""
            if len(Dict_of_KNs) == 0: print_status("Файл заполнен каким-то мусором"); return
            t = get_info_kn(RR, Dict_of_KNs)

## на выходе либо ошибка
            if t[:5] == "ERROR": wb.save(wbName); print_status(t); return
            if t[:5] == "FATAL": wb.save(wbName); print_status(t); return

## либо словарь с результатами
            for r in range_of_KNs:
                kn = r[0].value
                if kn is None: continue
                if r[1].value is not None: continue  # это не кадастровый номер (уже записано в файле)
                if Dict_of_KNs[kn] == "": r[1].value = "Объект по кадастровому номеру не найден"; continue
                t = Dict_of_KNs[kn].split("\t")
                for i in range(0, 8):
                    if i == 3:
                        try: r[i + 1].value = float(t[i])
                        except: r[i + 1].value = t[i]
                    else: r[i + 1].value = t[i]

            if time.monotonic() - tb > 600:  # backup every 10 minutes
                try: wb.save(wbName + ".backup")
                except: pass
                finally: tb = time.monotonic()

            print_progress(total=rows_total,
                          done=row - 1 + Multi - 1,
                          done_batch=row - 1 + Multi - 1 - rows_done,
                          duration=time.monotonic() - t0)
            row = row + Multi

        wb.save(wbName)
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    finally:
        if not WORK_PAUSED: print_progress(message="")
        wb.save(wbName)

# ------------------------------------------------------------ #
def RRgoAddr():
    global WORK_PAUSED
    prev_addr = None
    WORK_PAUSED = False
    t0 = time.monotonic(); tb = t0
    if RR is None: print_status("Сайт ещё не загружен"); return
    color = {True: "00CCFFCC", False: "00CCCCFF"}; color_index = True


## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Пробема с файлом: невозможно открыть и т.д."); return

    rows_total=ws.max_row - 1
    if rows_total < 1: print_status("Пустой файл"); return

## лист со списком адресов
## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break
    rows_done=row - 1
    if rows_done == rows_total:  print_status("Файл уже полностью обработан"); return

## добавить заголовок, если в файле его нет
    if rows_done < 0:
        h = "ЛС (может быть пустым)\t"\
            "Адрес\t"\
            "Найдено кадастровых номеров".split("\t")
        for i in range(0, len(h)): ws.cell(row=1, column=i + 1).value = h[i]

## лист с результататми текущего поиска
    try: ws2 = wb["KNaddr#results"]
    except: ws2 = wb.create_sheet("KNaddr#results")
    color_index = ws2.cell(row=ws2.max_row, column=1).fill.fgColor.rgb != "FFCCFFCC"
    row2 = ws2.max_row + 1
    h = "ЛС\t"\
        "Исходный адрес\t"\
        "Строка поиска\t"\
        "Кадастровый номер\t"\
        "Статус объекта\t"\
        "Дата обновления информации\t"\
        "Вид объекта недвижимости\t"\
        "Площадь\t"\
        "Ед. изм.\t"\
        "Адрес\t"\
        "Назначение или вид разрешенного использования\t"\
        "Дата последенего перехода права".split("\t")
    for i in range(0, len(h)): ws2.cell(row=1, column=i + 1).value = h[i]

    try:  # глобальная обработка исключений
## основной цикл обработки
        row1 = row + 1
        while True:
            if WORK_PAUSED: return

            addr_original = ws.cell(row=row1 , column=2).value
            if addr_original is None:
                print_status(f"Файл полностью обработан за {round(time.monotonic() - t0, 2)} сек.")
                print_progress(message="")
                return
            addr = re.sub(r"^\d{6}[\ ,]*", "", addr_original).strip()
            addr = clear_address_for_search(addr)
            if addr == prev_addr:
                t = "Повтор адреса"
            else:
                D = dict()
##                print(get_info_addr(RR, addr, D))
                t = get_info_addr(RR, addr, D)
            prev_addr = addr

## на выходе либо ошибка
            if t[:5] == "ERROR": wb.save(wbName); print_status(t); return
            if t[:5] == "FATAL": wb.save(wbName); print_status(t); return
            if t[:5] == "NOTIF": wb.save(wbName); print_status(t); return

## либо массив с результатами
            if t in ("Информация не найдена", "Повтор адреса"):
                ws.cell(row=row1, column=3).value = t
            else:
                ws.cell(row=row1, column=3).value = len(D)
                for key, value in D.items():
                    ws2.cell(row=row2, column=1).value = ws.cell(row=row1, column=1).value
                    ws2.cell(row=row2, column=2).value = addr_original
                    ws2.cell(row=row2, column=3).value = addr

                    t = value.split("\t")
                    ws2.cell(row=row2, column=4).value = key
                    for i in range(0, 8):
                        if i == 3:
                            try: j = float(t[i])
                            except: j = t[i]
                        else: j = t[i]
                        ws2.cell(row=row2, column=5 + i).value = j

                    for col in range(1, 13):
                        ws2.cell(row=row2, column=col).fill = PatternFill("solid", fgColor=color[color_index])

                    row2 = row2 + 1

                color_index = not color_index

            if (row - 1) % 10 == 0: wb.save(wbName)
            if time.monotonic() - tb > 600:  # backup every 10 minutes
                try: wb.save(wbName + ".backup")
                except: pass
                finally: tb = time.monotonic()

            print_progress(total=rows_total,
                          done=row1 - 1,
                          done_batch=row1 - 1 - rows_done,
                          duration=time.monotonic() - t0)
            row1 = row1 + 1

    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    finally:
        if not WORK_PAUSED: print_progress(message="")
        wb.save(wbName)

##########################################################################
RUN_FROM_IDLE = "idlelib" in sys.modules
main()
