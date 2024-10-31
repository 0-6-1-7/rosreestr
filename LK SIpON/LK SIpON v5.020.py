from get_info_4 import get_info_kn, get_info_addr, get_header, clear_notifications
from clear_address import clear_address_for_search

import os, sys

from configparser import ConfigParser

from re import match, sub

from time import monotonic, sleep

from math import floor

from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import *

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

RR = None
RUN_FROM_IDLE = None
APP_WINDOW = None
FILE_NAME = None
PROGRESS = None
PROGRESSBAR = None
STATUS = None
WORK_STOPPED = None
SEARCH_TYPE = None
SEARCH_BUTTON_CAPTION = None
TIMEOUT_VALUE = None
AUTH_TIMEOUT = 120 ## таймаут ожидания авторизации


# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
def get_file_to_proceed():
    filename = askopenfilename(parent=APP_WINDOW,
                               title="Выберите файл для обработки",
                               filetypes=[("Файлы Excel", "*.xlsx")], multiple=False)
    if filename == "":
        FILE_NAME.set("Файл не выбран")
        print_status("Файл не выбран")
        return
    s = filename
    r = s.rfind("/")
    if r > 0: s = s[:r] + "\n" + s[r + 1:]
    FILE_NAME.set(s)
    PROGRESS = check_file_to_proceed(filename)
    print_progress(total=PROGRESS[0], done=PROGRESS[1], message=PROGRESS[2], win_update=False)
    print_status("Файл выбран", False)
    if SEARCH_TYPE == "addr":
        SEARCH_BUTTON_CAPTION.set("Поиск по адресам")
    elif SEARCH_TYPE == "KN":
        SEARCH_BUTTON_CAPTION.set("Поиск по кад. номерам")
    else:
        SEARCH_BUTTON_CAPTION.set("...")
    APP_WINDOW.update()

# ------------------------------------------------------------ #
def check_file_to_proceed(wbName):
    global SEARCH_TYPE
    try:
        wb = load_workbook(filename=wbName)
        ws = wb.worksheets[0]
        rmax = ws.max_row
    except:
        return [None, None, "Проблема с файлом"]

## пустой файл
    if rmax == 1: SEARCH_TYPE = None; return [None, None, "Файл пустой"]

## в файле есть данные, определить тип файла - адреса или кадастровые номера
    try: v = match(r"^\d{1,2}:\d{1,2}:\d{6,7}:\d{1,10}$", ws.cell(row=2, column=2).value)
    except: SEARCH_TYPE = None; return [None, None, "Проблема с файлом"]

    if v is None: SEARCH_TYPE = "addr"
    else: SEARCH_TYPE = "KN"

## количество обработанных строк
    for row in range(2, rmax + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break

    if row == rmax: SEARCH_TYPE = None; return [None, None, "Файл полностью обработан"]
    else: return [rmax - 1, row - 1, None]

# ------------------------------------------------------------ #
def app_resize(event):
    with open("app.ini", "w") as conf:
        conf.write(APP_WINDOW.geometry())

# ------------------------------------------------------------ #
def print_progress(total=0, done=0, done_batch=0, duration=0, message=None, win_update=True):
    if message is not None:
        PROGRESS.set(message)
        print(message)
        APP_WINDOW.update()
        return
    msg = f"Всего в файле строк: {total}, из них обработано: {done}"
    if done_batch > 0:
        msg = msg + f"\nв т.ч. {done_batch}" + \
            (f" за {int(duration)} сек." if duration > 0 else "") + \
            (f" ~{int(3600 / (duration / done_batch))} строк в час" if duration > 0 else "")
    print(msg)
    PROGRESS.set(msg)
    if msg != "":
        PROGRESSBAR['value'] = int(done / total * 100)
        PROGRESSBAR.grid()
    else:
        PROGRESSBAR.grid_remove()
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def print_status(msg, win_update=True):
    txt = msg
    if msg[:5] == "FATAL":
        if msg == "FATAL_ERROR_PAGE_UNKNOWN":
            txt = "Непонятная страница"
        if msg == "FATAL_ERROR_PAGE_NOT_FULLY_LOADED":
            txt = "Страница загружена не полностью"
        if msg == "FATAL_ERROR_INIT_CHROME":
            txt = "Ошибка загрузки Chrome - он уже запущен или другая серъёзная ошибка"
        if msg == "FATAL_ERROR_CHROME":
            txt = "Chrome не запущен или закрыт"
        if msg == "FATAL_ERROR_INIT_START_PAGE_NOT_LOADED":
            txt = "Ошибка загрузки начальной страницы"
    elif msg[:5] == "ERROR":
        if msg == "ERROR_TIMEOUT":
            txt = "За 60 секунд не активировалась кнопка поиска"
        if msg == "ERROR_KN_FOUND_NOT_SHOWN":
            txt = "КН найден, но не отображается"
        if msg == "ERROR_SEARCH_FAILED":
            txt = "Поиск не завершился каким-либо результатом"
        if msg == "ERROR_OBJECT_NOT_SHOWN":
            txt = "Объект найден, но страница свойств не открылась"
        if msg == "ERROR_CAN_NOT_CLEAR_KN_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_CAN_NOT_CLEAR_ADDR_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_ADDR_NOT_FOUND":
            txt = "Адрес не найден"
        if msg == "FATAL_ERROR_SEARCH_BUTTON_IS_NOT_ACTIVE":
            txt = "Кнопка поиска на форме не активна"
    elif msg[:5] == "NOTIF":
        txt = msg.split("$")[1]
    else: txt = msg
    print(txt)
    STATUS.set(txt)
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def main():
    global APP_WINDOW, FILE_NAME, PROGRESS, STATUS, SEARCH_BUTTON_CAPTION, PROGRESSBAR, TIMEOUT_VALUE
    APP_WINDOW = Tk()
    FILE_NAME = StringVar()
    TIMEOUT_VALUE = StringVar()
    TIMEOUT_VALUE.set(5)
    PROGRESS = StringVar()
    STATUS = StringVar()
    SEARCH_BUTTON_CAPTION = StringVar(value="...")

    APP_WINDOW.title("Росреестр - справочная информация по объектам недвижимости в режиме online")
    if os.path.isfile("app.ini"):
        with open("app.ini", "r") as conf:
            APP_WINDOW.geometry(conf.read())
    else:
        APP_WINDOW.geometry("600x300")
    APP_WINDOW.resizable(1, 1)
    APP_WINDOW.call("wm", "attributes", ".", "-topmost", True)
    APP_WINDOW.bind("<Configure>", app_resize)

    lbl_status = Label(APP_WINDOW, textvariable=STATUS)
    lbl_status.grid(row=1, column=1, columnspan=3)

    btn_1 = Button(APP_WINDOW, text="Запустить Chrome", command=RRinit, width=25)
    btn_1.grid(row=2, column=1, sticky="we")

    btn_2 = Button(APP_WINDOW, text="Выбрать файл: ", command=get_file_to_proceed)
    btn_2.grid(row=3, column=1, sticky="we")

    txt_filename = Label(APP_WINDOW, textvariable=FILE_NAME)
    txt_filename.grid(row=3, column=2, columnspan=3, sticky="w")

    lbl_progress = Label(APP_WINDOW, textvariable=PROGRESS)
    lbl_progress.grid(row=4, column=2, columnspan=3, sticky="we")

    PROGRESSBAR = Progressbar(APP_WINDOW, orient='horizontal', mode='determinate')
    PROGRESSBAR.grid(row=5, column=2, columnspan=2, sticky="we")
    PROGRESSBAR['value'] = 0
    PROGRESSBAR.grid_remove()

    btn_3 = Button(APP_WINDOW, textvariable=SEARCH_BUTTON_CAPTION, command=RRgo, width=15)
    btn_3.grid(row=7, column=1, sticky="we")

# значение таймаута
    lbl_timeout = Label(APP_WINDOW, text='Таймаут: ', width=15)
    lbl_timeout.grid(row=7, column=3, sticky="e")
    ent_timeout = Entry(APP_WINDOW, textvariable=TIMEOUT_VALUE, width=5)
    ent_timeout.grid(row=7, column=3, sticky="e")

    btn_4 = Button(APP_WINDOW, text="Стоп", command=RRstop, width=15)
    btn_4.grid(row=8, column=1, sticky="we")

    btn_5 = Button(APP_WINDOW, text="Выход", command=RRquit)
    btn_5.grid(row=9, column=3, sticky="we")

    lbl_about = Label(APP_WINDOW, text="Поиск по адресам и кадастровым номерам :: версия 5.020")
    lbl_about.grid(row=10, column=1, columnspan=3, sticky="w")

    Grid.columnconfigure(APP_WINDOW, 0, minsize=25)
    Grid.columnconfigure(APP_WINDOW, 1, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 2, minsize=200)
    Grid.columnconfigure(APP_WINDOW, 3, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 4, minsize=25, weight=1)
    APP_WINDOW.mainloop()


# ------------------------------------------------------------ #
def get_with_ESIA_auth(RR, url):
    auth = ConfigParser()
    auth.read("esia.ini", 'UTF-8')
    esia_login = auth.get("ESIA", "login", fallback=None)
    esia_password = auth.get("ESIA", "password", fallback=None)
    esia_role_name = auth.get("ESIA", "name", fallback=None)
    esia_role_index = auth.get("ESIA", "index", fallback=None)


    print_status("Выполняется авторизация в ЕСИА")
    RR.get(url)
    sleep(1)

    t0 = monotonic()
    while True:
        print_status(f"Выполняется авторизация в ЕСИА, прошло {floor(monotonic() - t0)} сек. из {AUTH_TIMEOUT}")
        if monotonic() - t0 > AUTH_TIMEOUT:
            return "ESIA_ERROR_TIMEOUT"
        
        current_url = RR.current_url.lower()

        if "lk.rosreestr.ru/" in current_url and "lk.rosreestr.ru/login" not in current_url:
            return "OK"

        if "esia.gosuslugi.ru/login/" in current_url:
            try:
                form_container = RR.find_element(By.CSS_SELECTOR, "div.form-container")
                form_login = form_container.find_element(By.ID, "login")
                form_password = form_container.find_element(By.ID, "password")
                form_button = form_container.find_element(By.XPATH, "//button[contains(text(),'Войти')]")

                if esia_login is not None:
                    form_login.send_keys(esia_login)
                if esia_password is not None:
                    form_password.send_keys(esia_password)
                if esia_login is not None and esia_password is not None:
                    form_button.click()
                    sleep(1)
            except:
                None

        if "lk.rosreestr.ru/login" in current_url:
            if esia_role_name is not None or esia_role_index is not None:
                try:
                    wait = WebDriverWait(RR, AUTH_TIMEOUT)
                    personal_cabinet_root = wait.until(
                        EC.presence_of_element_located((By.ID, "personal-cabinet-root")))
                    buttons = personal_cabinet_root.find_elements(By.TAG_NAME, "button")
                    if esia_role_name is not None:
                        for button in buttons:
                            if button.get_attribute("innerText").lower() == esia_role_name.lower():
                                button.click()
                                break
                    elif esia_role_index is not None:
                        button = buttons[int(esia_role_index) - 1]
                        button.click()
                except:
                    None
        sleep(1)
    return "OK"

# ------------------------------------------------------------ #
def RRinit():
    global RR
    try:
        print_status("Запускается Chrome.")
        chrome_options = Options()
##        chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\rosreestr")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        RR = webdriver.Chrome(options=chrome_options)
        RR.set_page_load_timeout(180)

        status = get_with_ESIA_auth(RR, "https://lk.rosreestr.ru/login")
        if status == "OK":
            print_status("Личный кабинет готов к работе")
        else:
            print_status(status)
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")

# ------------------------------------------------------------ #
def RRquit():
    global RR
    print_status("Программа закрывается, подождите...")
    try: RR.close()
    except: pass
    try: RR.quit()
    except: pass
    exit()

# ------------------------------------------------------------ #
def RRstop():
    global WORK_STOPPED
    WORK_STOPPED = True
    print_status("...СТОП...")

# ------------------------------------------------------------ #
def RRgo():
    print_status("...идёт обработка файла...")

    if RR is None:
        print_status("Сайт ещё не загружен")
        return

    ## переход на нужную страницу ЛК
    if not "real-estate-objects-online".lower() in RR.current_url.lower():
        RR.get("https://lk.rosreestr.ru/eservices/real-estate-objects-online")

    ## ожидание загрузки страницы - через долгую проверку авторизации
    t0 = monotonic()
    wait = WebDriverWait(RR, 5)

    while True:
        ## не дождались появления нужной страницы, на выход с ошибкой
        if monotonic() - t0 > AUTH_TIMEOUT:
            print_status("FATAL_ERROR_INIT_START_PAGE_NOT_LOADED")
            return

        if "real-estate-objects-online".lower() in RR.current_url.lower():
            ## ждём появления кнопки поиска
            try: search_button = wait.until(EC.visibility_of_element_located((By.ID, "realestateobjects-search")))
            except: search_button = None
            if search_button is not None:
                ## дождались, можно продолжать
                break 

    return_status = None    
    while True:
        if SEARCH_TYPE == "KN": return_status = RRgoKN()
        elif SEARCH_TYPE == "addr": return_status = RRgoAddr()

        if return_status is None:
            continue
        if return_status in "FATAL":
            break
        if return_status in "OK":
            return
        if return_status in "STOP":
            break
        RR.refresh

# ------------------------------------------------------------ #
def wb_save_30_sec(wb, wbName, save_time_mark):
    if monotonic() - save_time_mark > 60:
        wb.save(wbName)
        print("saved")
        return monotonic()
    else:
        print("waiting...")
        return save_time_mark
        
# ------------------------------------------------------------ #
def range_name(r1, c1, r2, c2):
    return chr(ord("A") - 1 + c1) + str(r1) + ":" + chr(ord("A") - 1 + c2) + str(r2)

# ------------------------------------------------------------ #
def get_timeout():
    global TIMEOUT_VALUE
    timeout_value = 0
    try:
        timeout_value = int(TIMEOUT_VALUE.get())
        if timeout_value < 5:
            timeout_value = 5
            TIMEOUT_VALUE.set(timeout_value)
        if timeout_value > 15:
            timeout_value = 15
            TIMEOUT_VALUE.set(timeout_value)
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    return timeout_value

# ------------------------------------------------------------ #
def RRgoKN():
    global WORK_STOPPED
    prev_info = None
    WORK_STOPPED = False
    t0 = monotonic(); backup_time_mark = t0; save_time_mark = t0

## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Проблема с файлом: невозможно открыть и т.д."); return

    rows_total=ws.max_row - 1
    if rows_total < 1: print_status("Пустой файл"); return

## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break
    rows_done=row - 1
    if rows_done == rows_total:
        print_status("Файл уже полностью обработан")
        return "OK"

## добавить заголовок, если в файле его нет
    header = get_header().split("\t")
    if row < 2:
        for i in range(0, len(header)): ws.cell(row=1, column=i + 3).value = header[i]

    try:  # глобальная обработка исключений
## основной цикл обработки
        row = row + 1

        ## нововведение от 21.02.23 - таймаут 5 секунд
        t1 = 0
        while True:
            if WORK_STOPPED: return "STOP"
            if ws.cell(row=row, column=2).value is None:
                print_status(f"Файл полностью обработан за {round(monotonic() - t0, 2)} сек.")
                print_progress(message="")
                wb.save(wbName)
                return "OK"

            kn = ws.cell(row=row , column=2).value
            kn = kn.replace(" ", "")
            if kn == "": continue
            print(f"kn = {kn}")
            if match(r"^\d{1,2}:\d{1,2}:\d{6,7}:\d{1,10}$", kn) is None:
                object_info = "это не кадастровый номер"
            else:
                print(get_timeout())
                while monotonic() - t1 < get_timeout():
                    sleep(1)
                result = get_info_kn(RR, kn, t1, get_timeout())

                if type(result) == type("string"):
                    print("==================")
                    print(result)
                    print("==================")
                    t1 = monotonic()
                else:
                    t1 = result['timeout_start']

## на выходе 
            ## либо строка с ошибкой
            if type(result) == type("string"):
                print("==================")
                print(result)
                print("==================")
                object_info = result
            else:
                object_info = result['object_info']                
            if object_info is not None:
                if object_info[:5] in "ERROR;FATAL":
                    wb.save(wbName)
                    return "STOP"
    ##                check_notifications(RR)
##                if object_info[:5] in "ERROR":
##                    return "REFRESH"
##                if object_info[:5] in "FATAL":
##                    return "STOP"

            ## либо строка с результатами, разделитель \t
            t = object_info.split("\t")
            for i in range(0, len(t)):
                if i == 3:
                    try:
                        ws.cell(row=row, column=i + 3).value = float(t[i])
                    except:
                        ws.cell(row=row, column=i + 3).value = t[i]
                else:
                    ws.cell(row=row, column=i + 3).value = t[i]

            if monotonic() - backup_time_mark > 600:  # backup every 10 minutes
                try: wb.save(wbName + ".backup")
                except: pass
                finally: backup_time_mark = monotonic()

            print_progress(total=rows_total,
                          done=row - 1,
                          done_batch=row - 1 - rows_done,
                          duration=monotonic() - t0)
            row = row + 1

            save_time_mark = wb_save_30_sec(wb, wbName, save_time_mark)

        return "OK"
            
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    finally:
        if not WORK_STOPPED: print_progress(message="")
        

# ------------------------------------------------------------ #
def RRgoAddr():
    global WORK_STOPPED
    prev_addr = None
    WORK_STOPPED = False
    t0 = monotonic(); backup_time_mark = t0; save_time_mark = t0
    
    if RR is None: print_status("Сайт ещё не загружен"); return
    color = {True: "00CCFFCC", False: "00CCCCFF"}; color_index = True

## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Пробема с файлом: невозможно открыть и т.д."); return

    rows_total=ws.max_row - 1
    if rows_total < 1: print_status("Пустой файл"); return

## лист со списком адресов
## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break
    rows_done=row - 1
    if rows_done == rows_total:
        print_status("Файл уже полностью обработан")
        return "OK"

## добавить заголовок, если в файле его нет
    if rows_done < 0:
        h = "ЛС (может быть пустым)\t"\
            "Адрес\t"\
            "Найдено кадастровых номеров".split("\t")
        for i in range(0, len(h)): ws.cell(row=1, column=i + 1).value = h[i]

## лист с результататми текущего поиска
    try: ws2 = wb["KNaddr#results"]
    except: ws2 = wb.create_sheet("KNaddr#results")
    color_index = ws2.cell(row=ws2.max_row, column=1).fill.fgColor.rgb != "FFCCFFCC"
    row2 = ws2.max_row + 1
    h = "ЛС\t"\
        "Исходный адрес\t"\
        "Строка поиска\t"\
        "Кадастровый номер\t"\
        "Статус объекта\t"\
        "Дата обновления информации\t"\
        "Вид объекта недвижимости\t"\
        "Площадь\t"\
        "Ед. изм.\t"\
        "Адрес\t"\
        "Назначение или вид разрешенного использования\t"\
        "Дата последнего перехода права".split("\t")
    for i in range(0, len(h)): ws2.cell(row=1, column=i + 1).value = h[i]

    try:  # глобальная обработка исключений
## основной цикл обработки
        row1 = row + 1
        while True:
            if WORK_STOPPED: return "STOP"

            addr_original = ws.cell(row=row1 , column=2).value
            if addr_original is None:
                print_status(f"Файл полностью обработан за {round(monotonic() - t0, 2)} сек.")
                print_progress(message="")
                wb.save(wbName)
                return "OK"
            addr = sub(r"^\d{6}[\ ,]*", "", addr_original).strip()
            addr = clear_address_for_search(addr)
            if addr == prev_addr:
                t = "Повтор адреса"
            else:
                D = dict()
##                print(get_info_addr(RR, addr, D))
                t = get_info_addr(RR, addr, D)
            prev_addr = addr

## на выходе либо ошибка
            if t[:5] in ("ERROR", "FATAL", "NOTIF"):
                save_time_mark = wb_save_30_sec(wb, wbName, save_time_mark); print_status(t)
                return

## либо массив с результатами
            ws.cell(row=row1, column=4).value = addr ## сохраним обработанную строку, по которой выполнялся поиск
            if t.strip() in ("Информация не найдена", "Повтор адреса"):
                ws.cell(row=row1, column=3).value = t
            else:
                ws.cell(row=row1, column=3).value = len(D)
                for key, value in D.items():
                    ws2.cell(row=row2, column=1).value = ws.cell(row=row1, column=1).value
                    ws2.cell(row=row2, column=2).value = addr_original
                    ws2.cell(row=row2, column=3).value = addr

                    t = value.split("\t")
## в некоторых случаях информация о КН отсутсвует (например, 34:00:000000:97839)
## в этом случае возвращается пустая строка
## нужно иметь это в виду
                    ws2.cell(row=row2, column=4).value = key
                    if not t[0] == "":
                        for i in range(0, 8):
                            if i == 3:
                                try: j = float(t[i])
                                except: j = t[i]
                            else: j = t[i]
                            ws2.cell(row=row2, column=5 + i).value = j
                    else:
                        ws2.cell(row=row2, column=5).value = "Росреестр: объект не найден"
                        for i in range(1, 8):
                            ws2.cell(row=row2, column=5 + i).value = "-"

                    for col in range(1, 13):
                        ws2.cell(row=row2, column=col).fill = PatternFill("solid", fgColor=color[color_index])

                    row2 = row2 + 1

                color_index = not color_index

            if (row1 - 1) % 10 == 0: save_time_mark = wb_save_30_sec(wb, wbName, save_time_mark)
            if monotonic() - backup_time_mark > 600:  # backup every 10 minutes
                try: wb.save(wbName + ".backup")
                except: pass
                finally: backup_time_mark = monotonic()

            print_progress(total=rows_total,
                          done=row1 - 1,
                          done_batch=row1 - 1 - rows_done,
                          duration=monotonic() - t0)
            row1 = row1 + 1

    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    finally:
        if not WORK_STOPPED: print_progress(message="")
        wb.save(wbName)

##########################################################################
RUN_FROM_IDLE = "idlelib" in sys.modules
main()
