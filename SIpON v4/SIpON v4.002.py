from get_info import get_info_kn, get_info_addr
from clear_address import clear_address_for_search

import os, re, sys, time
import PIL

from base64 import b64decode
from io import BytesIO
from PIL import Image, ImageTk
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

RR = None
RUN_FROM_IDLE = None
APP_WINDOW = None
FILE_NAME = None
PROGRESS = None
PROGRESSBAR = None
STATUS = None
LBL_CAPTCHA = None
TXT_CAPTCHA = None
WORK_PAUSED = None
SEARCH_TYPE = None
SEARCH_BUTTON_CAPTION = None


# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
def get_file_to_proceed():
    filename = askopenfilename(parent=APP_WINDOW,
                               title="Выберите файл для обработки",
                               filetypes=[("Файлы Excel", "*.xlsx")], multiple=False)
    if filename == "":
        FILE_NAME.set("Файл не выбран")
        print_status("Файл не выбран")
        return
    s = filename
    r = s.rfind("/")
    if r > 0: s = s[:r] + "\n" + s[r + 1:]
    FILE_NAME.set(s)
    PROGRESS = check_file_to_proceed(filename)
    print_progress(total=PROGRESS[0], done=PROGRESS[1], message=PROGRESS[2], win_update=False)
    print_status("Файл выбран", False)
    if SEARCH_TYPE == "addr":
        SEARCH_BUTTON_CAPTION.set("Поиск по адресам")
    elif SEARCH_TYPE == "KN":
        SEARCH_BUTTON_CAPTION.set("Поиск по кад. номерам")
    else:
        SEARCH_BUTTON_CAPTION.set("...")
    APP_WINDOW.update()

# ------------------------------------------------------------ #
def check_file_to_proceed(wbName):
    global SEARCH_TYPE
    try:
        wb = load_workbook(filename=wbName)
        ws = wb.worksheets[0]
        rmax = ws.max_row
    except:
        return [None, None, "Проблема с файлом"]

## пустой файл
    if rmax == 1: SEARCH_TYPE = None; return [None, None, "Файл пустой"]

## в файле есть данные, определить тип файла - адреса или кадастровые номера
    try: v = re.match(r"^\d{1,2}:\d{1,2}:\d{6,7}:\d{1,10}$", ws.cell(row=2, column=2).value)
    except: SEARCH_TYPE = None; return [None, None, "Проблема с файлом"]

    if v is None: SEARCH_TYPE = "addr"
    else: SEARCH_TYPE = "KN"

## количество обработанных строк
    for row in range(2, rmax + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break

    if row == rmax: SEARCH_TYPE = None; return [None, None, "Файл полностью обработан"]
    else: return [rmax - 1, row - 1, None]

# ------------------------------------------------------------ #
def app_resize(event):
    with open("app.ini", "w") as conf:
        conf.write(APP_WINDOW.geometry())

# ------------------------------------------------------------ #
def print_progress(total=0, done=0, done_batch=0, duration=0, message=None, win_update=True):
    if message is not None:
        PROGRESS.set(message)
        print(message)
        APP_WINDOW.update()
        return
    msg = f"Всего в файле строк: {total}, из них обработано: {done}"
    if done_batch > 0:
        msg = msg + f"\nв т.ч. {done_batch}" + \
            (f" за {round(duration, 2)} сек." if duration > 0 else "") + \
            (f" ~{int(3600 / (duration / done_batch))} строк в час" if duration > 0 else "")
    print(msg)
    PROGRESS.set(msg)
    if msg != "":
        PROGRESSBAR['value'] = int(done / total * 100)
        PROGRESSBAR.grid()
    else:
        PROGRESSBAR.grid_remove()
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def print_status(msg, win_update=True):
    txt = msg
    if msg[:5] == "FATAL":
        if msg == "FATAL_ERROR_CAPTCHA_WRONG":
            txt = "Введена неправильная капча"
        if msg == "FATAL_ERROR_UNKNOWN_PAGE":
            txt = "Непонятная страница"
        if msg == "FATAL_ERROR_INIT_CHROME":
            txt = "Ошибка загрузки Chrome - он уже запущен или другая серъёзная ошибка"
        if msg == "FATAL_ERROR_CHROME":
            txt = "Chrome не запущен или закрыт"
        if msg == "FATAL_ERROR_INIT_START_PAGE_NOT_LOADED":
            txt = "Ошибка загрузки начальной страницы"
        if msg == "FATAL_ERROR_CAPTCHA_NOT_ENTERED":
            txt = "Капча не введена"
    elif msg[:5] == "ERROR":
        if msg == "ERROR_KN_FOUND_NOT_SHOWN":
            txt = "КН найден, но не отображается"
        if msg == "ERROR_OBJECT_NOT_SHOWN":
            txt = "Объект найден, но страница свойств не открылась"
        if msg == "ERROR_CAN_NOT_CLEAR_KN_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_CAN_NOT_CLEAR_ADDR_FIELD":
            txt = "Не получается очистить поле ввода КН"
        if msg == "ERROR_ADDR_NOT_FOUND":
            txt = "Адрес не найден"
    else: txt = msg
    print(txt)
    STATUS.set(txt)
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def main():
    global APP_WINDOW, FILE_NAME, PROGRESS, STATUS, LBL_CAPTCHA, TXT_CAPTCHA, SEARCH_BUTTON_CAPTION, PROGRESSBAR
    APP_WINDOW = Tk()
    FILE_NAME = StringVar()
    PROGRESS = StringVar()
    STATUS = StringVar()
    SEARCH_BUTTON_CAPTION = StringVar(value="...")

    APP_WINDOW.title("Росреестр - справочная информация по объектам недвижимости в режиме online")
    if os.path.isfile("app.ini"):
        with open("app.ini", "r") as conf:
            APP_WINDOW.geometry(conf.read())
    else:
        APP_WINDOW.geometry("600x300")
    APP_WINDOW.resizable(1, 1)
    APP_WINDOW.call("wm", "attributes", ".", "-topmost", True)
    APP_WINDOW.bind("<Configure>", app_resize)

    lbl_status = Label(APP_WINDOW, textvariable=STATUS)
    lbl_status.grid(row=1, column=1, columnspan=3)

    btn_1 = Button(APP_WINDOW, text="Запустить Chrome", command=RRinit, width=25)
    btn_1.grid(row=2, column=1, sticky="we")

    btn_2 = Button(APP_WINDOW, text="Выбрать файл: ", command=get_file_to_proceed)
    btn_2.grid(row=3, column=1, sticky="we")

    txt_filename = Label(APP_WINDOW, textvariable=FILE_NAME)
    txt_filename.grid(row=3, column=2, columnspan=3, sticky="w")

    lbl_progress = Label(APP_WINDOW, textvariable=PROGRESS)
    lbl_progress.grid(row=4, column=2, columnspan=3, sticky="we")

    PROGRESSBAR = Progressbar(APP_WINDOW, orient='horizontal', mode='determinate')
    PROGRESSBAR.grid(row=5, column=2, columnspan=2, sticky="we")
    PROGRESSBAR['value'] = 0
    PROGRESSBAR.grid_remove()

    TXT_CAPTCHA = Entry(APP_WINDOW, justify="center")
    TXT_CAPTCHA.grid(row=6, column=1, sticky="we")

    img_nocaptcha = ImageTk.PhotoImage(PIL.Image.open(BytesIO(b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAMgAAAAyCAMAAAAuj2TTAAAAAX"
        "NSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAADUExURQAAAKd6"
        "PdoAAAABdFJOUwBA5thmAAAACXBIWXMAAA7DAAAOwwHHb6hkAA"
        "AAIElEQVRoQ+3BAQEAAACCIP+vbkhAAAAAAAAAAAAAcKgGJ0IA"
        "AT6TWzgAAAAASUVORK5CYII="))))
    LBL_CAPTCHA = Label(APP_WINDOW, image=img_nocaptcha)
    LBL_CAPTCHA.grid(row=6, column=2, sticky="we")

    btn_captcha = Button(APP_WINDOW, text="Обновить капчу", command=RRrefresh)
    btn_captcha.grid(row=6, column=3, sticky="we")

    btn_3 = Button(APP_WINDOW, textvariable=SEARCH_BUTTON_CAPTION, command=RRgo, width=15)
    btn_3.grid(row=7, column=1, sticky="we")

    btn_4 = Button(APP_WINDOW, text="Пауза", command=RRpause, width=15)
    btn_4.grid(row=8, column=1, sticky="we")

    btn_5 = Button(APP_WINDOW, text="Выход", command=RRquit)
    btn_5.grid(row=9, column=3, sticky="we")

    lbl_about = Label(APP_WINDOW, text="Поиск по адресам и кадастровым номерам :: версия 4.0")
    lbl_about.grid(row=10, column=1, columnspan=3, sticky="w")

    Grid.columnconfigure(APP_WINDOW, 0, minsize=25)
    Grid.columnconfigure(APP_WINDOW, 1, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 2, minsize=200)
    Grid.columnconfigure(APP_WINDOW, 3, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 4, minsize=25, weight=1)
    APP_WINDOW.mainloop()

# ------------------------------------------------------------ #
def RRinit():
    global RR
    try:
        print_status("Запускается Chrome. Дождитесь появления капчи.")
        chrome_options = Options()
        if not RUN_FROM_IDLE: chrome_options.add_argument("--headless")
##        chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\rosreestr")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        RR = webdriver.Chrome(options=chrome_options)
        RR.set_page_load_timeout(180)
        RR.get("https://lk.rosreestr.ru/eservices/real-estate-objects-online")
        wait = WebDriverWait(RR, 15)
        try:
            sipono = wait.until(EC.visibility_of_element_located((By.ID, "sipono-selector")))
        except:
            sta = "FATAL_ERROR_INIT_START_PAGE_NOT_LOADED"; print_status(msg=sta)
            return sta
## первая капча никогда не срабатывает
        reloadlink = RR.find_element(By.XPATH, "//span[@class='rros-ui-lib-captcha-content-reload-btn-name']")
        reloadlink.click()
        time.sleep(1)

        img = ImageTk.PhotoImage(GetCaptcha())
        LBL_CAPTCHA.configure(image=img)
        LBL_CAPTCHA.image = img
        APP_WINDOW.update()
        print_status("Начальная страница загружена, капча получена")
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")

# ------------------------------------------------------------ #
def RRquit():
    global RR
    print_status("Программа закрывается, подождите...")
    try: RR.close()
    except: pass
    try: RR.quit()
    except: pass
    exit()

# ------------------------------------------------------------ #
def RRpause():
    global WORK_PAUSED
    WORK_PAUSED = True
    print_status("...пауза...")

# ------------------------------------------------------------ #
def RRrefresh():
    global RR
    try:
        reloadlink = RR.find_element(
            By.XPATH, "//span[@class='rros-ui-lib-captcha-content-reload-btn-name']")
        reloadlink.click()
        time.sleep(1)
        captchaimage = GetCaptcha()
        img = ImageTk.PhotoImage(captchaimage)
        LBL_CAPTCHA.configure(image=img)
        LBL_CAPTCHA.image = img
        TXT_CAPTCHA.delete(0, END)
    except:
        return

# ------------------------------------------------------------ #
def GetCaptcha():
    global RR
    try:
        captcha = RR.find_element(By.CSS_SELECTOR, "img.rros-ui-lib-captcha-content-img")
        img_base64 = RR.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 200; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", captcha)
        return HiCaptcha(PIL.Image.open(BytesIO(b64decode(img_base64))))
    except:
        print_status("Ошибка получения капчи")
        return None

# ------------------------------------------------------------ #
def HiCaptcha(img):
    mask = PIL.Image.open(BytesIO(b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAMgAAAAyCAAAAAA8Oss9AAADzE"
        "lEQVR4nNVZW7bcIAyzyNx9df9ranE/kgA2wnEyt6dzf3KYiWxs"
        "IYU88EsAKRAAuBoBEBQ5BnFAERdwnL3KO42SsCIigsUBbtROSg"
        "jDDIsDZtgOuJ6+wVDGID+K+kvCWH9JmMzYgIay5C7uLwljDNxg"
        "i87AYUVsUKQNV0coIdYfgbG8rLVrWMmRsqojqY04IKIha5FTWv"
        "/dIjWgIUVvuaMN3IUxBjisrMuN8nZYmXvUhTYeWERNwH5Sf895"
        "9W2L4JXXxlzupUVAAvA154WFebZW049dFaYNW+m7FqlOJRRW37"
        "TIfvn13RqioHPhLaA1X9eLvhkGRCmsLMtd5rWwQrRxTnWUC6nV"
        "Jpu4q7KZ/vSADdqvIvtPON6+ZReBvEgkBKLo5WqrcmWR4paaaH"
        "9rJbAqTd4nFhl29rFbFUAan2bTjC0iJ+frXUTHub7pRksGj5hw"
        "qMjJs+t0wd0xAp19tIjgH9xoCfY7/v7/ceFH70/PwZhWh1ksDc"
        "pmp9o4t5MDpu9ZJLhaHCOwOqgi2gqT/sgML5sXdnTXIn1nD9dQ"
        "IHUUyB1tOJUch7qA1YcWaR650P6+dH4WJqGwtQ7bzElto+LKjf"
        "KarkriuuFmj2B89kphPe/VlTDDVpmD/FKvNLqSkFqYKMo6IKLh"
        "jkXwst1iapQJJJLQvD2AWGRNA2ErRUOJSWGrmYQxBiwsUW6U13"
        "ZVVuG2v+hGK9MfgYlggjEakmwtntkTGk1qIw6IaLhlkXNnv9ZG"
        "YJF6RcOcV5M05Okt5pfes0j18kxapEaU6SOL9Muvm+Asd7+WLi"
        "xirxTgsxNtbLrfaSlji2wIGdGU8Vfbbnu5O9vV93chIVe4g1WB"
        "fC2XmeQl2ab+yqJbNS2XyCLTAZDj2ZdbZAtoaLvptZhsV8X86n"
        "O6Z5EacUcXnTyN7bA/IQ2Yml/w5rsq8/+mXDnzE4t4WM4i2y3t"
        "c1idsf4NH+223S1JCIv6O2E6BkQ0hDDygL58nLXrfdsirL/lyR"
        "UNcV7bn32is9H1fKNV+xsDwh1/VaW+hCNg2j0Fotr18tAiKOYX"
        "WrkQwXaON0FokR5zFIQxo+VuEGkrEgrxz1q3nAS+K6OXtdBoqI"
        "2SDeiwbQh/tIuMN41DtxDW93dYJAGTGZuwSNvZ2/++K1tHEsYY"
        "uIS9ZZFDsvEuwhZ9hi37IzCWl7V2gy3ynT2h0aQ24oCIhtsWYd"
        "/ZF4fPtsj0nZ1pAz7oCsYYuMHWE4vQF2K2259hEfLM/jMtQr6z"
        "Lw4fbhH/nZ1pA1PQB1rEf2e3pJByP9Ui9pk9J+WPtIj8BQNuUv"
        "D5KM4GAAAAAElFTkSuQmCC")))
    img = img.convert("L")
    i = img.load()
    m = mask.load()
    width, height = mask.size
    for y in range(height):
        for x in range(width):
            if i[x, y] < 20: i[x, y] = 255
            elif abs(i[x,y] - m[x, y]) < 10: i[x, y] = 255
            if i[x, y] < 250: i[x, y] = 0

    for y in range(height - 3):
        for x in range(width - 3):
            block = [i[x, y], i[x + 1, y], i[x + 2, y], i[x, y + 1], i[x + 1, y + 1], i[x + 2, y + 1], i[x, y + 2], i[x + 1, y + 2], i[x + 2, y + 2]]
            if (
                block == [255, 255, 255, 0, 255, 255, 255, 255, 255] or
                block == [255, 255, 255, 255, 0, 255, 255, 255, 255] or
                block == [255, 255, 255, 255, 255, 0, 255, 255, 255] or
                block == [255, 255, 255, 0, 0, 255, 255, 255, 255] or
                block == [255, 255, 255, 255, 0, 0, 255, 255, 255] or
                block == [255, 255, 255, 0, 0, 0, 255, 255, 255]
                ):
                i[x, y], i[x + 1, y], i[x + 2, y], i[x, y + 1], i[x + 1, y + 1], i[x + 2, y + 1], i[x, y + 2], i[x + 1, y + 2], i[x + 2, y + 2] = 255, 255, 255, 255, 255, 255, 255, 255, 255

    return img

# ------------------------------------------------------------ #
def RRgo():
    if SEARCH_TYPE == "KN": RRgoKN()
    elif SEARCH_TYPE == "addr": RRgoAddr()
    else: pass

# ------------------------------------------------------------ #
def RRgoKN():
    global WORK_PAUSED
    prev_info = None
    WORK_PAUSED = False
    print_status("...идёт обработка файла...")
    t0 = time.monotonic(); tb = t0
    if RR is None: print_status("Сайт ещё не загружен"); return
    color = {True: "00CCFFCC", False: "00CCCCFF"}; color_index = True

## капчу надо вводить два раза два раза
    try:
        captchainput = RR.find_element(By.CSS_SELECTOR, "input.rros-ui-lib-captcha-input")
        while captchainput.get_attribute("value") != "": captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
        captchainput.send_keys(TXT_CAPTCHA.get())
        while captchainput.get_attribute("value") != "": captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
        captchainput.send_keys(TXT_CAPTCHA.get())
    except:
        print_status("Пробема с сайтом или браузером")
        return
## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Пробема с файлом: невозможно открыть и т.д."); return

    rows_total=ws.max_row - 1
    if rows_total < 1: print_status("Пустой файл"); return

## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break
    rows_done=row - 1
    if rows_done == rows_total:  print_status("Файл уже полностью обработан"); return

## добавить заголовок, если в файле его нет
    if row < 2:
        h = "Статус объекта\t"\
            "Дата обновления информации\t"\
            "Вид объекта недвижимости\t"\
            "Площадь\t"\
            "Ед. изм.\t"\
            "Адрес\t"\
            "Назначение или вид разрешенного использования\t"\
            "Дата последенего перехода права".split("\t")
        for i in range(0, len(h)): ws.cell(row=1, column=i + 3).value = h[i]

    try:  # глобальная обработка исключений
## основной цикл обработки
        row = row + 1
        while True:
            if WORK_PAUSED: return
            if ws.cell(row=row, column=2).value is None:
                print_status(f"Файл полностью обработан за {round(time.monotonic() - t0, 2)} сек.")
                print_progress(message="")
                break
## обработка порциями по несколько строк
            Multi = 50
            Dict_of_KNs = dict()
            if row + Multi - 1 > rows_total + 1: Multi = rows_total + 1 - row + 1
            range_of_KNs = ws[f"B{row}": f"J{row + Multi - 1}"]
            for r in range_of_KNs:
                kn = r[0].value
                if kn is None: continue
                kn = kn.replace(" ", "")
                r[0].value = kn
                if kn == "": continue
                if re.match(r"^\d{1,2}:\d{1,2}:\d{6,7}:\d{1,10}$", kn) is None:
                    r[1].value = "это не кадастровый номер"
                else:
                    Dict_of_KNs[kn] = ""
            if len(Dict_of_KNs) == 0: print_status("Файл заполнен каким-то мусором"); return
            t = get_info_kn(RR, Dict_of_KNs)

## на выходе либо ошибка
            if t[:5] == "ERROR": wb.save(wbName); print_status(t); return
            if t[:5] == "FATAL": wb.save(wbName); print_status(t); return

## либо словарь с результатами
            for r in range_of_KNs:
                kn = r[0].value
                if kn is None: continue
                if r[1].value is not None: continue  # это не кадастровый номер (уже записано в файле)
                if Dict_of_KNs[kn] == "": r[1].value = "Объект по кадастровому номеру не найден"; continue
                t = Dict_of_KNs[kn].split("\t")
                for i in range(0, 8):
                    if i == 3:
                        try: r[i + 1].value = float(t[i])
                        except: r[i + 1].value = t[i]
                    else: r[i + 1].value = t[i]

            if time.monotonic() - tb > 600:  # backup every 10 minutes
                try: wb.save(wbName + ".backup")
                except: pass
                finally: tb = time.monotonic()

            print_progress(total=rows_total,
                          done=row - 1 + Multi - 1,
                          done_batch=row - 1 + Multi - 1 - rows_done,
                          duration=time.monotonic() - t0)
            row = row + Multi

        wb.save(wbName)
    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    finally:
        if not WORK_PAUSED: print_progress(message="")
        wb.save(wbName)

# ------------------------------------------------------------ #
def RRgoAddr():
    global WORK_PAUSED
    prev_addr = None
    WORK_PAUSED = False
    print_status("...идёт обработка файла...")
    t0 = time.monotonic(); tb = t0
    if RR is None: print_status("Сайт ещё не загружен"); return
    color = {True: "00CCFFCC", False: "00CCCCFF"}; color_index = True

## капчу надо вводить два раза два раза
    try:
        captchainput = RR.find_element(By.CSS_SELECTOR, "input.rros-ui-lib-captcha-input")
        while captchainput.get_attribute("value") != "":
            captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
        captchainput.send_keys(TXT_CAPTCHA.get())
        while captchainput.get_attribute("value") != "":
            captchainput.send_keys(Keys.BACKSPACE * len(captchainput.get_attribute("value")))
        captchainput.send_keys(TXT_CAPTCHA.get())
    except:
        print_status("Пробема с сайтом или браузером")
        return

## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Пробема с файлом: невозможно открыть и т.д."); return

    rows_total=ws.max_row - 1
    if rows_total < 1: print_status("Пустой файл"); return

## лист со списком адресов
## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=3).value is None: row = row - 1; break
    rows_done=row - 1
    if rows_done == rows_total:  print_status("Файл уже полностью обработан"); return

## добавить заголовок, если в файле его нет
    if rows_done < 0:
        h = "ЛС (может быть пустым)\t"\
            "Адрес\t"\
            "Найдено кадастровых номеров".split("\t")
        for i in range(0, len(h)): ws.cell(row=1, column=i + 1).value = h[i]

## лист с результататми текущего поиска
    try: ws2 = wb["KNaddr#results"]
    except: ws2 = wb.create_sheet("KNaddr#results")
    color_index = ws2.cell(row=ws2.max_row, column=1).fill.fgColor.rgb != "FFCCFFCC"
    row2 = ws2.max_row + 1
    h = "ЛС\t"\
        "Исходный адрес\t"\
        "Строка поиска\t"\
        "Кадастровый номер\t"\
        "Статус объекта\t"\
        "Дата обновления информации\t"\
        "Вид объекта недвижимости\t"\
        "Площадь\t"\
        "Ед. изм.\t"\
        "Адрес\t"\
        "Назначение или вид разрешенного использования\t"\
        "Дата последенего перехода права".split("\t")
    for i in range(0, len(h)): ws2.cell(row=1, column=i + 1).value = h[i]

    try:  # глобальная обработка исключений
## основной цикл обработки
        row1 = row + 1
        while True:
            if WORK_PAUSED: return

            addr_original = ws.cell(row=row1 , column=2).value
            if addr_original is None:
                print_status(f"Файл полностью обработан за {round(time.monotonic() - t0, 2)} сек.")
                print_progress(message="")
                return
            addr = re.sub(r"^\d{6}[\ ,]*", "", addr_original).strip()
            addr = clear_address_for_search(addr)
            if addr == prev_addr:
                t = "Повтор адреса"
            else:
                D = dict()
                t = get_info_addr(RR, addr, D)
            prev_addr = addr

## на выходе либо ошибка
            if t[:5] == "ERROR": wb.save(wbName); print_status(t); return
            if t[:5] == "FATAL": wb.save(wbName); print_status(t); return

## либо массив с результатами
            if t in ("Информация не найдена", "Повтор адреса"):
                ws.cell(row=row1, column=3).value = t
            else:
                ws.cell(row=row1, column=3).value = len(D)
                for key, value in D.items():
                    ws2.cell(row=row2, column=1).value = ws.cell(row=row1, column=1).value
                    ws2.cell(row=row2, column=2).value = addr_original
                    ws2.cell(row=row2, column=3).value = addr

                    t = value.split("\t")
                    ws2.cell(row=row2, column=4).value = key
                    for i in range(0, 8):
                        if i == 3:
                            try: j = float(t[i])
                            except: j = t[i]
                        else: j = t[i]
                        ws2.cell(row=row2, column=5 + i).value = j

                    for col in range(1, 13):
                        ws2.cell(row=row2, column=col).fill = PatternFill("solid", fgColor=color[color_index])

                    row2 = row2 + 1

                color_index = not color_index

            if (row - 1) % 10 == 0: wb.save(wbName)
            if time.monotonic() - tb > 600:  # backup every 10 minutes
                try: wb.save(wbName + ".backup")
                except: pass
                finally: tb = time.monotonic()

            print_progress(total=rows_total,
                          done=row1 - 1,
                          done_batch=row1 - 1 - rows_done,
                          duration=time.monotonic() - t0)
            row1 = row1 + 1

    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")
    finally:
        if not WORK_PAUSED: print_progress(message="")
        wb.save(wbName)

##########################################################################
RUN_FROM_IDLE = "idlelib" in sys.modules
main()
