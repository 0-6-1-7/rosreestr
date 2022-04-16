# Тут должен быть выбор типа выписки
defaultTimeout = 315 ## пауза между запросами в сеукндах 315

import datetime, os, re, sys, time

from openpyxl import Workbook
from openpyxl import load_workbook
from recognize import recognize
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from send_email import send_email

EGRN = None
wb = None; ws = None
AuthKey = "" # ключ доступа
KNs = None # подготовленный для запроса список КН через ";"
KNsToDo = 0 # количество необработаных КН
defaultBatchMax = 150 # максимальное количество КН, обрабатываемых за один раз (можно было бы и 199, но пусть будет 150)
defaultBatchSize = 100 # размер задания по умолчанию
BatchMax = 0 # вычисленный размер для текущего списка: если в файле меньше 150 КН, то делаем одним списком, если 150 и больше, то делим по 100: batchMax = lambda y: y if y < 150 else y // (y // 100 + 1) + 1
Reg = None # регион для запроса
RunFromIDLE = None # True если скрипт запущен из IDLE
PreviousRequestTimeStamp = None # Метка времени регистрации предыдущего запроса (системное время), string
PreviousRequestTimeMark = None # Метка времени регистрации предыдущего запроса (монотонные часы), float
CaptchaReloadingCounter = 0 # Счётчик попыток перезагрузить капчу
CaptchaReloadingCounterMax = 30 # Максимум попыток
SaveCaptcha = False # сохранение капчи в файл

##--- end of header ---##------------------------------------------------------------##

def SecStr(t):
    s1 = f"{t}"[-1]
    if len(f"{t}") > 1: s2 = f"{t}"[-2]
    else: s2 = ""
    if s2 != "1":
        if s1 == "1": return f"{t} секунда"
        elif s1 == "2" or s1 == "3" or s1 == "4" : return f"{t} секунды"
        else: return f"{t} секунд"
    else: return f"{t} секунд"
##--- end of SecStr ---##------------------------------------------------------------##
    
def DisplayErrorMessage(text):
  ts = text.split("\n")
  maxlen = 40
  for t in ts: maxlen = max(maxlen, len(t))
  print("╓─" + "─" * maxlen + "─╖")
  print("║ " + "{val:{wid}}".format(wid = maxlen, val = ts[0]) + " ║")
  if len(ts) > 1: print("╟─" + "─" * maxlen + "─╢")
  for t in range(1, len(ts)): print("║ " + "{val:{wid}}".format(wid = maxlen, val = ts[t]) + " ║")
  print("╙─" + "─" * maxlen + "─╜")
##--- end of DisplayErrorMessage ---##------------------------------------------------------------##

def printProgressBar (iteration, total):
    length = 30
    fill = '█'
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\rПауза между запросами {SecStr(total)} [{bar}] осталось {SecStr(total - iteration)}      ', end = "\r")
    if iteration == total: print(" " * 90)
##--- end of printProgressBar ---##------------------------------------------------------------##        

def Wait(TimeToWait):
    global RunFromIDLE
    if RunFromIDLE:
        print(f"     Пауза между запросами {TimeToWait} секунд [----------]")
        print(f"                             ожидание [", end = "")
        t10 = TimeToWait // 10
        for TimeToWait in range(0, 10):
            time.sleep(t10)
            print("X", end = "")
        time.sleep(TimeToWait % 10)
        print("]")
    else:
        printProgressBar(0, TimeToWait)
        for t in range(0, TimeToWait):
            time.sleep(1)
            printProgressBar(t + 1, TimeToWait)
##--- end of Wait ---##------------------------------------------------------------##        

def EGRNinit(msg = ""):
    global EGRN
    global RunFromIDLE
    SiteRestartRetriesMax = 10 # количество попыток перезапуска сайта
    SiteRestartRetriesCounter = 0
    plt = 20  # page load timeout

    if msg != "": print(msg, "\n")
# На случай если мы попали сюда после какой-нибудь ошибки - нужно начать заново
    if EGRN != None:
        EGRN.close(); EGRN.quit()
    while True:
        try:
            chrome_options = Options()
            chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            if not RunFromIDLE: chrome_options.add_argument("--headless")
            EGRN = webdriver.Chrome(options = chrome_options)
            EGRN.implicitly_wait(10)
            DemoKey = None
            while True:
                EGRN.set_page_load_timeout(plt)
                try: EGRN.get("https://rosreestr.ru/wps/portal/p/cc_present/ir_egrn")
                except: EGRN.execute_script("window.stop();")
                try: DemoKey = WebDriverWait(EGRN, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'6F9619FF-8B86-D011-B42D-00CF4FC964FF')]")))
                except: plt = plt + 10
                if DemoKey != None: print("Сайт Росреестра работает, страница авторизации загружена нормально."); return("SiteOK")
                if plt > 30: break # слишком больщой таймаут; скорее всего, сайт не работает
                print("Сайт работает медленнее, чем хотелось бы...")            
        except: pass
        SiteRestartRetriesCounter = SiteRestartRetriesCounter + 1
        print(f"Страница не загружена. Перезагрузка, попытка № {SiteRestartRetriesCounter}")
        EGRN.close(); EGRN.quit()
        if SiteRestartRetriesCounter > SiteRestartRetriesMax:
            print("\n\n\n\nСайт Росреестра не работает. Мы пытались...")
            return("SiteFailed")
##--- end of EGRNinit ##------------------------------------------------------------##
        
def EGRNauth():
    global AuthKey
    global EGRN
    AuthStatus = None
    try:
        f = open("auth.txt", "r")
        AuthKey = f.readline()[:36]
        f.close()
        if re.search("\w{8}-\w{4}-\w{4}-\w{4}-\w{12}", AuthKey) != None: AuthStatus = "GotAuthKey"
        else: return("NoAuthKey")
    except: return("NoAuthKey")
        
    if AuthStatus == "GotAuthKey":
        vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")
        AuthKeyFields = vapp.find_elements(By.XPATH, ".//input[@type='text']")
        AuthKeyFields[0].send_keys(AuthKey)
        t0 = time.monotonic()
        while True:
            if re.search(AuthKeyFields[4].get_attribute("value"), AuthKey) != None:
                AuthStatus = "KeyEntered"
                break
            if time.monotonic() - t0 > 60:
                print("За 60 секунд ключ не принят.")
                return("AuthTimeOut")
    if AuthStatus == "KeyEntered":
        SendButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Войти')]")
        SendButton.click()
        print("Авторизация выполнена.")
        return("AuthOK")
##--- end of EGRNauth ##------------------------------------------------------------##

def EGRNSearchPage():
    global EGRN
    vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")
    t0 = time.monotonic()
    while True:
        try:
            sButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Поиск объектов недвижимости')]")
            break
        except: pass
        if time.monotonic() - t0 > 60: return("SearchPageFailed") # За 60 секунд страница не загрузилась полностью
    sButton.click()
    try:
        InfoImage = vapp.find_element(By.XPATH, ".//img[contains(@src, 'i_icon.gif')]")
        print("Страница поиска готова.")
        return("SearchPageOK")
    except: return("SearchPageFailed")
##--- end of EGRNSearchPage ##------------------------------------------------------------##

def EGRNSearch():
    global KNS, Reg
    global EGRN
    try:
        vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")
        fKNs = vapp.find_element(By.XPATH, ".//input[contains(@class,'v-textfield-prompt')]")
        fReg = vapp.find_element(By.XPATH, ".//input[contains(@class,'v-filterselect-input')]")
        sButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Найти')]")

        fKNs.click()
        EGRN.execute_script(f"arguments[0].value='{KNs}';", fKNs)
        fKNs.send_keys(Keys.ENTER)
        time.sleep(1)
        fReg.send_keys(Reg)
        time.sleep(5)
        fReg.send_keys(Keys.ENTER)
        time.sleep(5)
        fReg.send_keys(Keys.TAB)
        EGRN.execute_script("arguments[0].scrollIntoView();", sButton)
        sButton.click()
        print("Поиск запущен")
        t0 = time.monotonic()
    except: return("SearchKNsError")
    EGRN.implicitly_wait(1)
    print("Ожидание реультата поиска")
    while True:
        DataTable = None
        try: DataTable = EGRN.find_element(By.CSS_SELECTOR, "table.v-table-table")
        except: pass
        NotificationWarning = None
        try:
            NotificationWarningText = EGRN.find_element(By.CLASS_NAME, "v-Notification-warning").get_attribute("innerText")
            if re.search("Не найдены данные, удовлетворяющие Вашему запросу.", NotificationWarningText) != None:
                DisplayErrorMessage(NotificationWarningText)
                return("SearchKNsNotFound")
        except: pass
        if DataTable == None and NotificationWarning == None:
            time.sleep(1)
            if time.monotonic() - t0 > 60:
                print("За 60 секунд не дождались нужной страницы. Вероятно, ошибка сайта.")
                return("SearchKNsError")   
        elif DataTable != None:
            time.sleep(1)
            t = EGRN.find_element(By.CSS_SELECTOR, "div.blockSmall").get_attribute("innerText")
            print(f"Результаты поиска получены. {t}\n")
            print(f"Проверка достоверности реультата...\n")
            if t == "Найдено более 200 записей. Необходимо уточнить параметры запроса.": 
                print("При поиске возникла какая-то ошибка, нужен перезапуск.")
                return("SearchKNsError")
            return("SearchKNsOK")
            
##--- end of EGRNSearch ##------------------------------------------------------------##

def RecognizeCaptcha(c):
    global EGRN
    global SaveCaptcha
    try:
        img_base64 = EGRN.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", c)
        captcha = recognize(img_base64, SaveCaptcha)
    except:
        print("Ошибка при обработке капчи. Скорее всего, это ошибка сайта.")
        captcha = ""
    if captcha == "44444":
        print("Ошибка при обработке капчи (44444).")
        captcha = ""
    return captcha
##--- end of RecognizeCaptcha ##------------------------------------------------------------##

def GetCaptcha(vapp):
##    print("Обработка капчи")
    OK = False
    global CaptchaReloadingCounter
    while True:
        time.sleep(1)
        try: CaptchaImage = vapp.find_element(By.XPATH, ".//img[@style='width: 180px; height: 50px;']")
        except: print("На странице не найдена капча"); break
        captcha = RecognizeCaptcha(CaptchaImage)
        if captcha != "": return(captcha)
        CaptchaReloadingCounter = CaptchaReloadingCounter + 1
        print(f"     получение другой капчи (попытка № {CaptchaReloadingCounter} из {CaptchaReloadingCounterMax})")
        CaptchaReloadButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Другую картинку')]")
        CaptchaReloadButton.click()
        if CaptchaReloadingCounter > CaptchaReloadingCounterMax:
            print("Капча не получена после нескольких попыток. Вероятно, ошибка сайта.")
            break
        time.sleep(10)
    return("00000")
##--- end of GetCaptcha ##------------------------------------------------------------##

def CheckRQfile():
    global wb, ws
    global KNs, Reg
    global KNsToDo
    global defaultBatchMax, defaultBatchSize, BatchMax

    RQStatus = None
    while True:
        try:
            wb = load_workbook(filename = 'rq.xlsx')
            ws = wb.worksheets[0]
        except:
            print("\n\n\nНет файла rq.xlsx")
            RQStatus = "NoFileRQError"
            break
        try:
            wb.save(filename = 'rq.xlsx')
        except:
            print("\n\n\nОшибка при сохранении файла rq.xlsx - вероятно, он открыт в Excel")
            RQStatus = "FileRQFailed"
            time.sleep(5)
            break
        ## Проверим структуру файла
        Reg = ws.cell(row = 1, column = 1).value # Регион: во-первых, должен быть не пустой. во-вторых - совпадать со списком
        if Reg == None:
            print("\n\n\nВ файле не указан регион")
            RQStatus = "FileRQNoRegionError"
            break
        else:
            RegOK = False
            for RegRow in wb["Список регионов"].iter_rows(1): RegOK = RegOK or RegRow[0].value == Reg
            if not RegOK:
                print("\n\n\nВ файле неправильно указан регион")
                RQStatus = "FileRQNoRegionError"
                break
        rmax = ws.max_row
        if rmax < 3:
            print("\n\n\nВ файле слишком мало данных")
            RQStatus = "FileRQTooLittleDataError"
            break
        ## Проверим все КН в списке на правильность по маске
        for row in ws.iter_rows(min_row = 3):
            kn = row[0].value
            if re.search(r"\[\d\d:\d\d:\d{1,7}:\d{1,}\]", f"[{kn}]") != None:
                row[1].value = "да"
            else:
                row[1].value = "нет"
                print(f"Неправильный кадастровый номер в строке {row[1].row}")
                RQStatus = "FileRQIncorrectKNError"
        ## Проверим  список на дубликаты КН
        KNList = []
        DupKNList = []
        for row in ws.iter_rows(min_row = 3):
            if row[0].value in KNList: DupKNList.append(row[0].row)
            else: KNList.append(row[0].value)
        if DupKNList != []:
            print(f"Найдены дубли кадастровых номеров в строках\n{DupKNList}")
            RQStatus = "FileRQDuplicateKNError"
                
        wb.save(filename = 'rq.xlsx')
        break
    if RQStatus == None:
        ## Подготовим список КН, по которым в поле Запрос пусто, т.е. запросов ещё не было
        ## Возможен вариант, когда КН не был найден - его нет в ФГИС ЕГРН
        ## KNsToDo - запроса не было; NZcount - есть номер запроса; ZZcount = КН не найден
        NZcount = 0; ZZcount = 0; KNsToDo = 0
        for row in ws.iter_rows(min_row = 3):
            if row[2].value == None:  KNsToDo = KNsToDo + 1
            elif re.search(r"80-\d{9}", f"{row[2].value}") != None:  NZcount = NZcount + 1
            elif row[2].value == "Не найден во ФГИС ЕГРН": ZZcount = ZZcount + 1
            else: print(f"Произошло что-то странное со списком в строке {row.row - 2}") ## мы не должны дойти до этого места, но вдруг...
        
        if KNsToDo == 0:
            os.system('cls')
            print("Все КН из списка обработаны.")
            if ZZcount > 0: print(f"\tЧасть КН ({ZZcount}) не найдена во ФГИС ЕГРН")
            print("\nМожно закрыть это окно.")
            RQStatus = "FileRQAllDone"
            return(RQStatus) # Выход после выполнения всех запросов
        
        getBatchMax = lambda y: y if y < defaultBatchMax else y // (y // defaultBatchSize + 1) + 1
        BatchMax = getBatchMax(KNsToDo)
        BatchKNcount = 0 # количество запросов в текущем пакете
        KNs = "" # список запросов для поиска  в текущем пакете
        print("Подготовка списка КН")
        for row in ws.iter_rows(min_row = 3):
            if  re.search(r"80-\d{9}", f"{row[2].value}") == None and row[2].value != "Не найден во ФГИС ЕГРН":
                BatchKNcount = BatchKNcount + 1
                if BatchKNcount <= BatchMax: KNs = f"{KNs};{row[0].value}" # добавим в список не больше этого количества КН
                else: break
        if KNs[0] == ";": KNs = KNs[1:]
        print(f"Всего кадастровых номеров {KNsToDo + NZcount + ZZcount}", end = "")
        if NZcount + ZZcount > 0:
            print(f", в том числе:\n\tуже выполнены: {NZcount}, ещё не выполнены: {KNsToDo}", end = "")
            if ZZcount > 0: print(f", не найдены во ФГИС ЕГРН: {ZZcount}", end = "")
        if BatchKNcount < KNsToDo: print(f", обработка частями по {BatchMax} КН", end = "")
        print()
        RQStatus = "FileRQOK"
    return(RQStatus)
##--- end of CheckRQfile ##------------------------------------------------------------##

def RQSave(KN, NZ, d):
    global wb, ws
    KNfound = False
    for r in ws.iter_rows(min_row = 3):
        if r[0].value == KN:
            ws.cell(row = r[0].row, column = 3).value = NZ
            ws.cell(row = r[0].row, column = 4).value = d
            wb.save(filename = 'rq.xlsx')
            KNfound = True
    if not KNfound:
        print(f"Кадастровый номер {KN} не найден в исходном списке, номер запроса {NZ}, дата {d}")
        print(f"{KN}\t{NZ}\t{d}")
##--- end of RQSave ##------------------------------------------------------------##

def MarkKNAsNotFound():
    global wb, ws
    global KNs
    if len(KNs) == 0: return
    for KN in KNs.split(";"):
        for r in ws.iter_rows(min_row = 3):
            if r[0].value == KN: ws.cell(row = r[0].row, column = 3).value = "Не найден во ФГИС ЕГРН"
    wb.save(filename = 'rq.xlsx')

##--- end of MarkKNAsNotFound ##------------------------------------------------------------##

def GetInfo():
    global EGRN
    global wb, ws
    global defaultTimeout
    global PreviousRequestTimeStamp, PreviousRequestTimeMark
    global CaptchaReloadingCounter

    GetInfoStatus = 0

    t0 = time.monotonic()
    while EGRN.find_element(By.CSS_SELECTOR, "div.blockGrey").is_displayed():
        if time.monotonic() - t0 > 60:
            print("За 60 секунд не дождались нужной страницы. Вероятно, ошибка сайта.")
            return(9)

    KN = re.search(r"\b\d{2}:\d{2}:\d{1,7}:\d{1,}\b", EGRN.find_element(By.CSS_SELECTOR, "div.header3").get_attribute("innerText"))[0]
    print(f"{KN} - подготовка запроса "  + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    vapp = EGRN.find_element(By.CSS_SELECTOR, "div.v-app")

    CaptchaField = vapp.find_elements(By.XPATH, ".//input[@type='text']")[0]
    SendButton = vapp.find_element(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Отправить запрос')]")

## Версия для заказа выписок о переходах права
    Radio = vapp.find_elements(By.XPATH, ".//input[@type='radio']")[1]
    EGRN.execute_script("arguments[0].scrollIntoView();", Radio)
    Radio.click()

# цикл будет работать в случае когда капча распознанна неверно
    while True:
# Обработаем капчу
        captcha = GetCaptcha(vapp)
        if captcha == "00000":
            print("Ошибка при обработке капчи")
            return(9)
        CaptchaReloadingCounter = 0 # обнулим счётчик попыток
# Ввод капчи и отправка запроса
        while True:
            try:
                CaptchaField.click()
                break
            except:
                time.sleep(1)
                print("CaptchaField.click() wait")
        time.sleep(1)
        CaptchaField.send_keys(captcha)
        time.sleep(1)
        if PreviousRequestTimeStamp != None:
            SecondsElapsed = time.monotonic() - PreviousRequestTimeMark
            if int(defaultTimeout - SecondsElapsed) in range(1, 30):
                time.sleep(defaultTimeout - SecondsElapsed)
        CurrentRequestTimeMark = time.monotonic() # метка времени отправки запроса
        # проверить что капча введена - есть подозрение, что иногда она пропадает
        # print("Captcha in field: [", CaptchaField.get_attribute("value"), "]")
        SendButton.click()

## Дождёмся PopUp
        PopUp = None
        while True:
            PopUp = vapp.find_elements(By.XPATH, "//div[@class='popupContent']")
            if len(PopUp) > 0: break
            else: time.sleep(1) #;print(".", end = " ")

## PopUp появился, разберёмся с ним
## возможнрые варианты:
##  1. с этими вариантами всё ясно
##        подтврждение успешного запроса
##        Превышен интервал между запросами
##        Ошибка при регистрации запроса
##  2. а с этим - пока не очень
##        Communication problem. Скорее всего, в зависимости от кода ошибки. можно продолжить работу - нужно проверить.
##          Иногда он продолжает нормально работать, а иногда просит авторизацию.
##  3. неожиданный вариант: невидимый PopUp с сообщением о неверной капче.
##        time.sleep(10) ### странно, зачем я вставил сюда эту паузу?
        PopUpType = None
        t = PopUp[0].get_attribute("innerText")
        if re.search("Запрос зарегистрирован", t) != None:
            PreviousRequestTimeStamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            PreviousRequestTimeMark = time.monotonic()
            PopUpType = "Normal"
        elif re.search("Превышен интервал между запросами", t) != None:
            PopUpType = "Problem_Timeout"
##            print("Превышен интервал между запросами")
        elif re.search("Ошибка при регистрации запроса", t) != None:
            PopUpType = "Problem_Registration"
        elif re.search("Произошла ошибка во время регистрации запроса", t) != None:
            PopUpType = "Problem_Registration"
        elif re.search("Ошибка ввода капчи", t) != None:
            PopUpType = "Problem_Captcha"
        elif re.search("Communication problem", t) != None:
            PopUpType = "Problem_Communication"
        else:
            PopUpType = "Problem_Unknown"

        if PopUpType != "Normal": DisplayErrorMessage(t)

        if  PopUpType == "Normal":
            OkButton = PopUp[0].find_elements(By.XPATH, ".//span[contains(@class,'v-button-caption') and contains(text(),'Продолжить работу')]")[0]
            NZ = ""
            NZ = re.search(r"\d{2}\-\d{9}", PopUp[0].get_attribute("innerText"))[0]
            OkButton.click()
            print(f"{KN} :: {NZ} - запрос выполнен в {PreviousRequestTimeStamp} за {round(PreviousRequestTimeMark - CurrentRequestTimeMark, 2)} сек.")
            send_email("EGRN bot OK", f"{KN} :: {NZ}")
            RQSave(KN, NZ, PreviousRequestTimeStamp)
            return(0)
        elif PopUpType == "Problem_Timeout" or PopUpType == "Problem_Registration":
            PopUpCloseBox = PopUp[0].find_elements(By.XPATH, ".//div[@class='v-window-closebox']")[0]
            PopUpCloseBox.click()
            time.sleep(10)
            return(1)
        elif PopUpType == "Problem_Captcha": pass
        elif PopUpType == "Problem_Communication": return(9)
        else: return(9)
##--- end of GetInfo ##------------------------------------------------------------##
        
def GetAll(tout = defaultTimeout):
    global EGRN
    global PreviousRequestTimeMark

    AllKN = []
    AllON = []
    i = 1
    RetryCount = 0
    DataTable = EGRN.find_element(By.CSS_SELECTOR, "table.v-table-table")
    AllON = DataTable.find_elements(By.CSS_SELECTOR, "tr > td:first-child")
    for ON in AllON:
        AllKN.append(ON.get_attribute("innerText"))
    for KN in AllKN:
        DataTable = EGRN.find_element(By.CSS_SELECTOR, "table.v-table-table")
        while True:
##        После каждого запроса (и попытки запроса) нужно заново определять все объекты на странице
            AllON = []
            AllON = DataTable.find_elements(By.CSS_SELECTOR, "tr > td:first-child")
            for ON in AllON:
                if KN == ON.get_attribute("innerText"): break
            EGRN.execute_script("arguments[0].scrollIntoView();", ON)
            print(f"Объект {i} из {len(AllON)} ({KNsToDo}) :: ", end = "")
            time.sleep(2)
            try: ON.click()
            except:
                print("\nНевозможно кликнуть на найденном объекте")
                print("Фатальная ошибка, придётся начинать сначала")
                send_email("EGRN bot FATAL ERROR", "Fatal error! Restart.")
                return("FatalError")
            time.sleep(2)
            GetInfoStatus = GetInfo()
            if GetInfoStatus == 0: # Запрос выполнен
                i = i + 1
                RetryCount = 0

                if PreviousRequestTimeMark != None:
                    SecondsElapsed = time.monotonic() - PreviousRequestTimeMark
                    if SecondsElapsed > tout: Wait(tout)
                    else: Wait(int(tout - SecondsElapsed - 25))
                else: Wait(tout)
                break
            elif GetInfoStatus == 1: # Превышен интервал между запросами или ошибка регистрации запроса
                  RetryCount = RetryCount + 1
                  print(f"Попытка № {RetryCount + 1}")
            elif GetInfoStatus == 9: # Фатальная ошибка, придётся начинать сначала
                  print("Фатальная ошибка, придётся начинать сначала")
                  send_email("EGRN bot FATAL ERROR", "Fatal error! Restart.")
                  return("FatalError")
    if KNsToDo == 0: # все КН обработаны
        print("Всё готово!")
        send_email("EGRN bot all done", "All done!")
        return("AllDone")
    else:
        print("Задание отработано!")
        send_email("EGRN bot batch done", "Batch done!")
        return("BatchDone")
##--- end of GetAll ##------------------------------------------------------------##    
    
##------------------------------------------------------------##
##------------------------------------------------------------##
##------------------------------------------------------------##
RunFromIDLE = "idlelib" in sys.modules

while True:
    Status = None
    Status = CheckRQfile()
    if Status == "FileRQOK":
        while True:
            Status = EGRNinit("Начало работы с сайтом")
            if Status == "SiteOK": Status = EGRNauth()
            if Status == "AuthOK": Status = EGRNSearchPage()
            if Status == "NoAuthKey": print("Нет ключа доступа. Дальнейшая работа без ключа невозможна."); quit()
            if Status == "SearchPageOK": Status = EGRNSearch()
            if Status == "SearchKNsOK": Status = GetAll()
            if Status == "SearchKNsError": print("Перезапуск для обработки очередной части списка"); break
            if Status == "AllDone": break
            if Status == "BatchDone": print("Перезапуск для обработки очередной части списка"); break
            if Status == "SearchKNsNotFound":
                print("Ни один из запрошенных КН не найден во ФГИС ЕГРН.")
                MarkKNAsNotFound()
                print("Перезапуск для обработки очередной части списка")
                break
            else: print("Перезапуск после фатальной ошибки"); break
    else: print("\n"); time.sleep(3)
    if Status == "FileRQAllDone": break
    if Status == "AllDone": break


time.sleep(14400)
