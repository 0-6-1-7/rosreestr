import re, time

from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from recognize import recognize
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from send_email import send_email

EGRN = None
wb = None
ws = None
defaulttout = 300
tout = defaulttout ## пауза между запросами в сеукндах
AuthKey = "" # ключ доступа

##------------------------------------------------------------##
def EGRNinit():
    global EGRN
    SiteRestartRetriesMax = 10 # количество попыток перезапуска сайта
    SiteRestartRetriesCounter = 0
    SiteStatusOK = False
# На случай если мы попали сюда после какой-нибудь ошибки - нужно начать заново
    if EGRN != None:
        EGRN.close()
        EGRN.quit()
    while True:
        try:
            EGRN = webdriver.Chrome()
            EGRN.set_page_load_timeout(30)
            EGRN.get("https://rosreestr.ru/wps/portal/p/cc_present/ir_egrn")
            DemoKey = WebDriverWait(EGRN, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'6F9619FF-8B86-D011-B42D-00CF4FC964FF')]")))
            print("Сайт Росреестра работает, страница авторизации загружена нормально.")
            EGRN.implicitly_wait(20)
            return("SiteOK")
        except:
            SiteRestartRetriesCounter = SiteRestartRetriesCounter + 1
            print(f"Страница не загружена. Перезагрузка, попытка № {SiteRestartRetriesCounter}")
        EGRN.close()
        EGRN.quit()
        if SiteRestartRetriesCounter > SiteRestartRetriesMax:
            print("\n\n\n\nСайт Росреестра не работает. Мы пытались...")
            return("SiteFailed")

##------------------------------------------------------------##
def EGRNauth():
    global AuthKey
    global EGRN
    Status = None
    try:
        f = open("auth.txt", "r")
        AuthKey = f.readline()
        f.close()
        if re.search("\w{8}-\w{4}-\w{4}-\w{4}-\w{12}", AuthKey) != None:
            Status = "GotAuthKey"
        else:
            return("NoAuthKey")
    except:
        return("NoAuthKey")
        
    if Status == "GotAuthKey":
        vapp = EGRN.find_element_by_css_selector("div.v-app")
        AuthKeyFields = vapp.find_elements_by_xpath(".//input[@type='text']")
        AuthKeyFields[0].send_keys(AuthKey)
        t0 = time.time()
        while True:
            if re.search(AuthKeyFields[4].get_attribute("value"), AuthKey) != None:
                Status = "KeyEntered"
                break
            if time.time() - t0 > 60:
                print("За 60 секунд ключ не принят.")
                return("AuthTimeOut")
    if Status == "KeyEntered":
        SendButton = vapp.find_element_by_xpath(".//span[contains(@class,'v-button-caption') and contains(text(),'Войти')]")
        SendButton.click()
        return("AuthOK")

def EGRNSearchPage():
    global EGRN
    vapp = EGRN.find_element_by_css_selector("div.v-app")
    t0 = time.time()
    while True:
        try:
            sButton = vapp.find_element_by_xpath(".//span[contains(@class,'v-button-caption') and contains(text(),'Поиск объектов недвижимости')]")
            break
        except:
            pass
        if time.time() - t0 > 60:
# За 60 секунд страница не загрузилась полностью
            return("SearchPageFailed")
    sButton.click()
    try:
        InfoImage = vapp.find_element_by_xpath(".//img[contains(@src, 'i_icon.gif')]")
        return("SearchPageOK")
    except:
        return("SearchPageFailed")
    
##------------------------------------------------------------##
def Wait(t):
    print(f"     Пауза между запросами {t} секунд [----------]")
    print(f"                             ожидание [", end = "")
    t10 = t // 10
    for t in range(0, 10):
        time.sleep(t10)
        print("X", end = "")
    time.sleep(t % 10)
    print("]")

##------------------------------------------------------------##
def RecognizeCaptcha(c):
    global EGRN
    try:
        img_base64 = EGRN.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", c)
        captcha = recognize(img_base64)
    except:
        print("Ошибка при обработке капчи. Скорее всего, это ошибка сайта.")
        captcha = ""
    if captcha == "44444":
        print("Ошибка при обработке капчи (44444).")
        captcha = ""
    return captcha

##------------------------------------------------------------##  
def GetCaptcha(vapp):
    print("Обработка капчи")
    OK = False
    CaptchaReloadingCounter = 100
    while True:
        time.sleep(1)
        try:
            CaptchaImage = vapp.find_element_by_xpath(".//img[@style='width: 180px; height: 50px;']")
        except:
            print("На странице не найдена капча")
            break
        captcha = RecognizeCaptcha(CaptchaImage)
        if captcha != "":
            return(captcha)
        print("     получение другой капчи")
        CaptchaReloadButton = vapp.find_element_by_xpath(".//span[contains(@class,'v-button-caption') and contains(text(),'Другую картинку')]")
        CaptchaReloadButton.click()
        CaptchaReloadingCounter = CaptchaReloadingCounter - 1
        if CaptchaReloadingCounter == 0:
            print("Капча не получена после нескольких попыток. Вероятно, ошибка сайта.")
            break
        time.sleep(10)
    return("00000")

##------------------------------------------------------------##
def CheckRQfile():
    global wb
    global ws
    try:
        wb = load_workbook(filename = 'rq.xlsx')
    except:
        wb = Workbook()
    ws = wb.worksheets[0]
    if ws.cell(row = 1, column = 1).value == None:
        ws.cell(row = 1, column = 1).value = "Кадастровый номер"
        ws.cell(row = 1, column = 2).value = "Номер запроса"
        ws.cell(row = 1, column = 3).value = "Дата запроса"
    try:
        wb.save(filename = 'rq.xlsx')
        return("FileRQOK")
    except:
        print("\n\n\n\n\nОшибка при сохранении файла rq.xlsx - вероятно, он открыт в Excel")
        return("FileRQFailed")

##------------------------------------------------------------##
def GetInfo():
    global EGRN
    global wb
    global ws
    global tout

    Status = 0

    t0 = time.time()
    while EGRN.find_element_by_css_selector("div.blockGrey").is_displayed():
        time.sleep(1)
        if time.time() - t0 > 60:
            print("За 60 секунд не дождались нужной страницы. Вероятно, ошибка сайта.")
            return(9)

    KN = re.search(r"\b\d{2}:\d{2}:\d{1,7}:\d{1,}\b", EGRN.find_element_by_css_selector("div.header3").get_attribute("innerText"))[0]
    print(f"{KN} - подготовка запроса")
    vapp = EGRN.find_element_by_css_selector("div.v-app")

    CaptchaField = vapp.find_elements_by_xpath(".//input[@type='text']")[0]
    SendButton = vapp.find_element_by_xpath(".//span[contains(@class,'v-button-caption') and contains(text(),'Отправить запрос')]")

    Radio = vapp.find_elements_by_xpath(".//input[@type='radio']")[1]
    EGRN.execute_script("arguments[0].scrollIntoView();", Radio)
    Radio.click()

# цикл будет работать в случае когда капча распознанна неверно
    while True:
# Обработаем капчу
        captcha = GetCaptcha(vapp)
        if captcha == "00000":
            print("Ошибка при обработке капчи")
            return(9)

# Ввод капчи и отправка запроса
        while True:
            try:
                CaptchaField.click()
                break
            except:
                time.sleep(1)
                print("CaptchaField.click() wait")
        time.sleep(1)
        CaptchaField.send_keys(captcha)
        time.sleep(1)
        SendButton.click()

## Дождёмся PopUp
        PopUp = None
        while True:
            PopUp = vapp.find_elements_by_xpath("//div[@class='popupContent']")
            if len(PopUp) > 0:
                break
            else:
                time.sleep(5)
                print(".", end = " ")

## PopUp появился, разберёмся с ним
## возможнрые варианты:
##  1. с этими вариантами всё ясно
##        подтврждение успешного запроса
##        Превышен интервал между запросами
##        Ошибка при регистрации запроса
##  2. а с этим - пока не очень
##        Communication problem. Скорее всего, в зависимости от кода ошибки. можно продолжить работу - нужно проверить.
##          Иногда он продолжает нормально работать, а иногда просит авторизацию.
##  3. неожиданные вариант: невидимый PopUp с сообщением о неверной капче.

        PopUpType = None
        t = PopUp[0].get_attribute("innerText")
        if re.search("Запрос зарегистрирован", t) != None:
            PopUpType = "Normal"
        elif re.search("Превышен интервал между запросами", t) != None:
            PopUpType = "Problem_Timeout"
        elif re.search("Ошибка при регистрации запроса", t) != None:
            PopUpType = "Problem_Registration"
        elif re.search("Ошибка ввода капчи", t) != None:
            PopUpType = "Problem_Captcha"
        elif re.search("Communication problem", t) != None:
            PopUpType = "Problem_Communication"
        else:
            PopUpType = "Problem_Unknown"
        print(f"всё проверили, получилось {PopUpType}")

        if  PopUpType == "Normal":
            OkButton = PopUp[0].find_elements_by_xpath(".//span[contains(@class,'v-button-caption') and contains(text(),'Продолжить работу')]")[0]
            NZ = ""
            NZ = re.search(r"\d{2}\-\d{9}", PopUp[0].get_attribute("innerText"))[0]
            OkButton.click()
            print(f"{KN} :: {NZ} - запрос выполнен " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            send_email("EGRN bot OK", f"{KN} :: {NZ}")
            row = ws.max_row + 1
            ws.cell(row = row, column = 1).value = KN
            ws.cell(row = row, column = 2).value = NZ
            ws.cell(row = row, column = 3).value = datetime.now().strftime("%d.%m.%Y")
            wb.save(filename = 'rq.xlsx')
            return(0)
        elif PopUpType == "Problem_Timeout" or PopUpType == "Problem_Registration":
            PopUpCloseBox = PopUp[0].find_elements_by_xpath(".//div[@class='v-window-closebox']")[0]
            PopUpCloseBox.click()
            time.sleep(10)
            return(1)
        elif PopUpType == "Problem_Captcha":
            pass
        elif PopUpType == "Problem_Communication":
            return(9)
        else:
            return(9)
##------------------------------------------------------------##
def GetAll(t = defaulttout):
    if not CheckRQfile():
        return
    global EGRN
    global tout
    tout = t
    AllKN = []
    AllON = []
    i = 1
    RetryCount = 0
    DataTable = EGRN.find_element_by_css_selector("table.v-table-table")
    AllON = DataTable.find_elements_by_css_selector("tr > td:first-child")
    for ON in AllON:
        AllKN.append(ON.get_attribute("innerText"))
    for KN in AllKN:
        DataTable = EGRN.find_element_by_css_selector("table.v-table-table")
        while True:
##        После каждого запроса (и попытки запроса) нужно заново определять все объекты на странице
            AllON = []
            AllON = DataTable.find_elements_by_css_selector("tr > td:first-child")
            for ON in AllON:
                if KN == ON.get_attribute("innerText"):
                    break
            EGRN.execute_script("arguments[0].scrollIntoView();", ON)
            print(f"     Объект {i} из {len(AllON)} :: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " :: ", end = "")
            ON.click()
            time.sleep(2)
            Status = GetInfo()
            if Status == 0:
                  i = i + 1
                  RetryCount = 0
                  Wait(tout)
                  break
            elif Status == 1: # Превышен интервал между запросами или ошибка регистрации запроса
                  RetryCount = RetryCount + 1
                  print(f"Попытка № {RetryCount}")
            elif Status == 9: # Фатальная ошибка - доделаю когда-нибудь позже
                  print("It looks like fatal error")
                  send_email("EGRN bot FATAL ERROR", "Fatal error!")
                  quit()

    send_email("EGRN bot all done", "All done!")
    print("Всё готово!")
    return("AllDone")
    
##------------------------------------------------------------##
##------------------------------------------------------------##
##------------------------------------------------------------##
while True:
    Status = None
    Status = CheckRQfile()
    if Status == "FileRQOK":
        Status = EGRNinit()
    if Status == "SiteOK":
        Status = EGRNauth()
    if Status == "AuthOK":
        Status = EGRNSearchPage()
    if Status == "NoAuthKey":
        break
    if Status == "SearchPageOK":
        break
print("Всё готово для работы")



