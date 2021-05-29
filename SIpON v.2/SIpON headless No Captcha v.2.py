from gatherInfo import Convert

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import openpyxl, os, sys, time

SIpON = None
localStatus = 0 # состояние сайта
RunFromIDLE = None # True если скрипт запущен из IDLE
  
##------------------------------------------------------------##
def SIpONinit():
    global SIpON, RunFromIDLE

    chrome_options = Options()
    if not RunFromIDLE: 
        chrome_options.add_argument("--headless")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    SIpON = webdriver.Chrome(options = chrome_options)

    SIpON.implicitly_wait(5)
    plt = 10
    localStatus = 0
    form = None
##  SIpON.set_page_load_timeout(plt)
    while True:
        SIpON.set_page_load_timeout(plt)
        try:
            SIpON.get("https://rosreestr.ru/wps/portal/online_request")
        except:
            SIpON.execute_script("window.stop();")
            localStatus = 999

        try:
            form = SIpON.find_element_by_id("online_request_search_form_span")
        except:
            plt = plt + 10

        if form != None:
            return 0

        if plt > 120:
            return 999

##------------------------------------------------------------##
def SIpONrestart(t):
  global SIpON
  if SIpON != None:
    SIpON.close()
    SIpON.quit()
  print(f"Restart")
  SIpONinit()

##------------------------------------------------------------##
def RecognizeCaptcha(c):
  global SIpON
  img_base64 = SIpON.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", c)
  captcha = recognize(img_base64)
  return captcha
  
##------------------------------------------------------------##
def GetInfo(KN):
  global SIpON
  i = 1
  localStatus = 1
  while localStatus != 0 and i < 10:
    localStatus = 0
    if i > 1:
      SIpONrestart(i)
    if localStatus == 0:
      try:
        SetSearchCritVisible = SIpON.find_element_by_css_selector("a[href*='SetSearchCritVisible']")
        SetSearchCritVisible.click()
      except:
        pass
    if localStatus == 0:
      try:
        SIpON.execute_script("arguments[0].click();", SIpON.find_element_by_id("cad-number"))
        cadnum = SIpON.find_element_by_css_selector("input[type='text'][name='cad_num']")
        SIpON.execute_script("arguments[0].click();", cadnum)
        cadnum.clear()
        cadnum.send_keys(KN)
        SIpON.execute_script("arguments[0].click();", SIpON.find_element_by_id("submit-button"))
      except:
        SIpON.execute_script("window.stop();")

    rc = 0
    while True:
      try:
        link = SIpON.find_element_by_css_selector("a[href*='object_data_id']").get_attribute("href")
        break
      except:
        time.sleep(2)
        rc = rc + 1
      if rc > 4:
        localStatus = "No Result"
        break

    if localStatus == "No Result":
      try:
        infomsg = SIpON.find_element_by_css_selector("td.infomsg1")
        if infomsg != None:
          return f"{KN}\tОбъект не найден"
      except:
        localStatus = "Error"
        
    if localStatus == 0:
      mainwindow = SIpON.current_window_handle
      SIpON.execute_script("window.open('" + link + "');")
      SIpON.switch_to.window(SIpON.window_handles[1])
      
      rc = 0
      sw_r_enc = None
      r_enc = None
      while True:
        try:
          sw_r_enc = SIpON.find_element_by_id("sw_r_enc")
          r_enc = SIpON.find_element_by_id("r_enc")
        except:
          pass
        if sw_r_enc == None or r_enc == None:
          time.sleep(2)
          rc = rc + 1
          print(f"Попытка {rc}")
        else:
          break
        if rc > 4:
          break
        
      try:
        if "none" in r_enc.get_attribute("style"):
          SIpON.execute_script("arguments[0].scrollIntoView();", sw_r_enc)
          sw_r_enc.click()
      except:
        pass
      try:
        content = SIpON.find_element_by_css_selector("div.portlet-body").get_attribute("innerText")
        t = Convert(content)
      except:
        localStatus = 999
      finally:
        SIpON.close()
        SIpON.switch_to.window(mainwindow)  
    if localStatus == 999:
      i = i + 1
  else:
    if localStatus == 999:
      return f"{KN}\tНе удалось выполнить за 10 попыток"
    elif localStatus == 0:
      return t # OK

##------------------------------------------------------------##
RunFromIDLE = "idlelib" in sys.modules
SIpONinit()

wbName = 'KN.xlsx'
if len(sys.argv) > 1: wbName = sys.argv[1]
else: wbName = 'KN.xlsx'

p = os.path.split(os.path.dirname(os.path.abspath(__file__)))[1] + "\\" + wbName

from openpyxl import load_workbook
wb = load_workbook(filename = wbName)
ws1 = wb.worksheets[0]
ws2 = wb.create_sheet("Sheet KN")

r = 2 # строка в исходном списке
rmax = ws1.max_row - 1 # всего строк
t = "Лицевой счёт\tКадастровый номер\tДата обновления информации\tКатегория объекта\tУчтённый\tПлощадь\tЕдиница изм.\tАдрес\tТип объекта\tФорма собств.\tДата последнего перехода права собств."
tl = t.split("\t")
for l in range(0, len(tl)):
  ws2.cell(row = 1, column = l + 1).value = tl[l]
cc = ws1.cell(row = r , column = 1).value
ls = ws1.cell(row = r , column = 2).value

while cc != None:
  t = GetInfo(cc)

  print(f"-- [ {r - 1} из {rmax} ] -- [ {p} ] --")
  print(t)
  tl = t.split("\t", 10)
  ws2.cell(row = r, column = 1).value = ls
  for l in range(0, len(tl)):
    ws2.cell(row = r, column = l + 2).value = tl[l]
  if (r - 1) % 10 == 0:
    wb.save(wbName)
  r = r + 1
  cc = ws1.cell(row = r, column = 1).value
  ls = ws1.cell(row = r, column = 2).value

wb.save(wbName)
print(f"\n\n\n\n\nВсё готово! ----- [ {p} ] ----- [ {time.strftime('%H:%M:%S', time.localtime())} ] ")
print()
