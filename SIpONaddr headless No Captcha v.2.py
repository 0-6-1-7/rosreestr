import logging
import platform
from configparser import ConfigParser
from selenium.webdriver import Chrome

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

import openpyxl, os, re, sys, time
from webdriver_manager.chrome import ChromeDriverManager

from recognize import recognize

RunFromIDLE = None  # True если скрипт запущен из IDLE
SIpON = None
status = 0  # состояние сайта

cfg = ConfigParser()
cfg.read("config.ini")
isHeadless = cfg.get("driver", "headless")
DEBUG = True

if DEBUG:
    logging.basicConfig(
        format=u'%(filename)-18s [LINE:%(lineno)-4s] %(levelname)-8s [%(asctime)s] %(message)s',
        level=logging.INFO,

    )
else:
    logging.basicConfig(
        format=u'%(filename)-18s [LINE:%(lineno)-4s] %(levelname)-8s [%(asctime)s] %(message)s',
        level=logging.WARNING
        # filename='debug.log'
    )


def SIpONinit():
    global SIpON
    # RunFromIDLE

    chrome_options = Options()

    if isHeadless:
        chrome_options.add_argument("--headless")

    try:
        if platform.system() == 'Linux':
            SIpON = Chrome('./chromedriverUnix', options=chrome_options)
        if platform.system() == 'Darwin':
            SIpON = Chrome('./chromedriverDarwin', options=chrome_options)
        else:
            SIpON = Chrome('chromedriver', options=chrome_options)

    except Exception as ex:
        logging.warning(f'Use ChromeDriverManager')
        logging.warning(f'{ex}')
        SIpON = Chrome(ChromeDriverManager().install(), options=chrome_options)

    SIpON.implicitly_wait(5)
    plt = 10
    form = None
    while True:
        SIpON.set_page_load_timeout(plt)
        try:
            SIpON.get("https://rosreestr.ru/wps/portal/online_request")
        except:
            SIpON.execute_script("window.stop();")

        try:
            form = SIpON.find_element_by_id("online_request_search_form_span")
        except:
            plt = plt + 10

        if form != None:
            return 0

        if plt > 30:
            plt = 10
            # SIpON.close()
            # SIpON.quit()
            SIpONrestart()


def SIpONrestart():
    global SIpON

    if SIpON != None:
        SIpON.close()
        SIpON.quit()

    logging.info(f"Restart")
    SIpONinit()


def RecognizeCaptcha(c):
    global SIpON
    # code from https://gist.github.com/spirkaa/4c3b8ad8fd34324bd307#gistcomment-3157744
    img_base64 = SIpON.execute_script(
        """var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""",
        c)
    captcha = recognize(img_base64)
    return captcha


def ifNone(s):
    if s == None:
        return ""
    else:
        return f"{s}"


def GetInfoByAddr(new_subject="",
                  new_region="",
                  new_settlement="",
                  new_street="",
                  new_house="",
                  new_building="",
                  new_apartment=""):
    global SIpON
    prev_subject = None
    prev_region = None
    prev_settlement = None
    prev_street = None
    prev_house = None
    prev_apartment = None

    Search_retry_count = 0  # номер попытки
    status = None
    prev_links = None

    SIpON.implicitly_wait(1)
    SIpON.set_page_load_timeout(10)

    ## начало основного цикла поиска
    while True:  # цикл крутится до получения результата или 10 попыток
        status = 0
        if Search_retry_count > 9: break
        if Search_retry_count > 1: SIpONrestart()

        # № открываем форму поиска, если она скрыта после предыдущего поиска
        try:
            SIpON.find_element_by_css_selector("a[href*='SetSearchCritVisible']").click()
        except:
            pass

        ## начало заполнения формы поиска
        ## если форма готова для ввода данных, попробуем ввести новые данные и проверить что получилось
        Fill_retry_count = 0  # номер попытки ввода данных
        while True:  # цикл крутится пока не будет заполнена форма или 10 попыток

            if Fill_retry_count > 3:
                status = 999;
                break
            else:
                time.sleep(3)

            # № пытаемся получить содержимое формы
            ## префикс
            # f_ - ссылка на элемент DOM
            # prev_ - текущее значение поля
            # new_ - новое значение для поиска
            try:
                f_subject = SIpON.find_element_by_css_selector("select[name='subject_id']");
                prev_subject = Select(
                    f_subject).first_selected_option.text
            except:
                Fill_retry_count = Fill_retry_count + 1;
                continue  # субъект не может быть пустым, повторить цикл

            try:
                f_region = SIpON.find_element_by_css_selector("select[name='region_id']");
                prev_region = Select(
                    f_region).first_selected_option.text
            except:
                prev_region = None  # регион может быть пустым

            try:
                f_settlement = SIpON.find_element_by_css_selector(
                    "select[name='settlement_id']")
                prev_settlement = Select(f_settlement).first_selected_option.text
            except:
                prev_settlement = None  # нас. пункт может быть пустым

            try:
                f_street = SIpON.find_element_by_css_selector("input[name='street']")
                prev_street = f_street.get_attribute('value')

                f_house = SIpON.find_element_by_css_selector("input[name='house']")
                prev_house = f_house.get_attribute('value')

                f_building = SIpON.find_element_by_css_selector("input[name='building']")
                prev_building = f_building.get_attribute('value')

                f_apartment = SIpON.find_element_by_css_selector("input[name='apartment']")
                prev_apartment = f_apartment.get_attribute('value')
            except:
                Fill_retry_count = Fill_retry_count + 1;
                continue

            status = 999
            # поля в форме заполнены неправильно или форма пустая
            # print("В форме\tНовое")
            # print(prev_subject,"\t",new_subject)
            # print(prev_region,"\t",new_region)
            # print(prev_settlement,"\t",new_settlement)
            # print(prev_street,"\t",new_street)
            # print(prev_house,"\t",new_house)
            # print(prev_apartment,"\t",new_apartment)

            # субъект
            if ifNone(prev_subject) != ifNone(new_subject):
                try:
                    f_subject.send_keys(Keys.NULL)
                    f_subject.send_keys(Keys.HOME)
                    f_subject.send_keys(new_subject)
                    time.sleep(3)
                except:
                    Fill_retry_count = Fill_retry_count + 1
                    continue

                # регион
            if ifNone(prev_region) != ifNone(new_region):
                try:
                    f_region.send_keys(Keys.NULL)
                    f_region.send_keys(Keys.HOME)
                    f_region.send_keys(new_region)
                    time.sleep(3)
                except:
                    Fill_retry_count = Fill_retry_count + 1
                    continue

            # населённый пункт; внезапно: выбор НП может быть не предусмотрен в данном субъекте и/или регионе
            if ifNone(prev_settlement) != ifNone(new_settlement):
                if f_settlement.get_attribute("disabled"):  # если поле не заполнено
                    if new_settlement != None:
                        Fill_retry_count = Fill_retry_count + 1
                        continue  # а должно быть
                else:  # поле заполнено
                    f_settlement.send_keys(Keys.HOME)  # непустой НП или очистка поля
                    if new_settlement != None:
                        f_settlement.send_keys(new_settlement)  # непустой НП

            try:
                if ifNone(prev_street) != ifNone(new_street):
                    f_street.click()
                    f_street.clear()
                    if new_street != None:
                        f_street.send_keys(new_street)

                if ifNone(prev_house) != ifNone(new_house):
                    f_house.click()
                    f_house.clear()
                    if new_house != None:
                        f_house.send_keys(new_house)

                if ifNone(prev_building) != ifNone(new_building):
                    f_building.click()
                    f_building.clear()
                    if new_building != None:
                        f_building.send_keys(new_building)

                if ifNone(prev_apartment) != ifNone(new_apartment):
                    f_apartment.click()
                    f_apartment.clear()
                    if new_apartment != None:
                        f_apartment.send_keys(new_apartment)
            except:
                Fill_retry_count = Fill_retry_count + 1
                continue

            # № проверка заполнения формы

            try:
                prev_subject = Select(f_subject).first_selected_option.text
            except:
                pass

            try:
                prev_region = Select(f_region).first_selected_option.text
            except:
                prev_region = None

            try:
                prev_settlement = Select(f_settlement).first_selected_option.text
            except:
                prev_settlement = None

            try:
                prev_street = f_street.get_attribute('value')
                prev_house = f_house.get_attribute('value')
                prev_apartment = f_apartment.get_attribute('value')
            except:
                Fill_retry_count = Fill_retry_count + 1;
                continue

            TestFilledData = True \
                             and re.match(f"^{ifNone(new_subject)}", ifNone(prev_subject), re.IGNORECASE) != None \
                             and re.match(f"^{ifNone(new_region)}", ifNone(prev_region), re.IGNORECASE) != None \
                             and re.match(f"^{ifNone(new_settlement)}", ifNone(prev_settlement), re.IGNORECASE) != None \
                             and ifNone(prev_street) == ifNone(new_street) \
                             and ifNone(prev_house) == ifNone(new_house) \
                             and ifNone(prev_building) == ifNone(new_building) \
                             and ifNone(prev_apartment) == ifNone(new_apartment)

            if TestFilledData:
                # все поля в форме соответствуют тому, что мы ищем
                status = 0
                break
            else:
                Fill_retry_count = Fill_retry_count + 1
                continue
            # end of if TestFilledData:
        ## конец заполнения формы поиска

        ## попытка выполнить поиск
        if status != 0:
            Search_retry_count = Search_retry_count + 1
            continue
        try:
            btn = SIpON.find_element_by_id("submit-button")
            SIpON.execute_script("arguments[0].scrollIntoView();", btn)
            SIpON.execute_script("arguments[0].click();", btn)
        except:
            Search_retry_count = Search_retry_count + 1
            continue  # не найдена кнопка поиска?

        ## ожидание результата поиска
        if status != 0: Search_retry_count = Search_retry_count + 1; continue
        Result_wait_count = 0
        while True:
            if Result_wait_count > 4:
                status = 999;
                break
            elif Result_wait_count > 0:
                time.sleep(5)
            else:
                time.sleep(1)
            try:
                links = SIpON.find_elements_by_css_selector("tr > td > a[href*='object_data_id']")
                kn = ""
                # дополнительная проверка на изменение результатов поиска, если гне изменились. ждём ещё
                if links != None and links == prev_links: Result_wait_count = Result_wait_count + 1; continue
                if links != None and links != prev_links:
                    for l in links:
                        tr = l.find_element_by_xpath("..").find_element_by_xpath("..")
                        kn = kn + ("\n" if len(kn) > 0 else "") + tr.find_elements_by_css_selector("td")[
                            1].get_attribute("innerText")
                    prev_links = links
                    return kn
            except:
                time.sleep(2)
                Result_wait_count = Result_wait_count + 1
        ## конец ожидания результата

        ## если сюда попали, значит, результата нет            
        if status != 0: Search_retry_count = Search_retry_count + 1; continue
        try:
            if SIpON.find_element_by_css_selector("td.infomsg1") != None: return f"{kn}\tОбъект не найден"
        except:
            status == 999

        ## сделали всё, что могли, финальная првоерка статуса
        if status != 0: Search_retry_count = Search_retry_count + 1
    ## конец основного цикла; если сюда попали, то всё плохо
    return "Не удалось выполнить за 10 попыток"


# RunFromIDLE = "idlelib" in sys.modules
SIpONinit()

if len(sys.argv) > 1:
    wbName = sys.argv[1]
else:
    wbName = "KNaddr.xlsx"

p = os.path.split(os.path.dirname(os.path.abspath(__file__)))[1] + "\\" + wbName

from openpyxl import load_workbook

wb = load_workbook(filename=wbName)
ws1 = wb.worksheets[0]
ws2 = wb.create_sheet("Sheet KNaddr")

r1 = 2  # строка в исходном списке
r1max = ws1.max_row - 1  # всего строк
r2 = 2  # строка в списке результатов
t = "Лицевой счёт\tКадастровый номер"
tl = t.split("\t")
for l in range(0, len(tl)):
    ws2.cell(row=1, column=l + 1).value = tl[l]
##cc = ws1.cell(row = r , column = 1).value

while ws1.cell(row=r1, column=1).value != None:
    t = GetInfoByAddr(
        ws1.cell(row=r1, column=3).value,
        ws1.cell(row=r1, column=4).value,
        ws1.cell(row=r1, column=5).value,
        ws1.cell(row=r1, column=6).value,
        ws1.cell(row=r1, column=7).value,
        ws1.cell(row=r1, column=8).value,
        ws1.cell(row=r1, column=9).value)

    logging.info(f"-- [ {r1 - 1} из {r1max} ] -- [ {p} ] --")
    logging.info(t)
    tl = t.split("\n")
    for l in range(0, len(tl)):
        ws2.cell(row=r2, column=1).value = ws1.cell(row=r1, column=1).value
        ws2.cell(row=r2, column=2).value = tl[l]
        r2 = r2 + 1
    if (r1 - 1) % 10 == 0:
        wb.save(wbName)
    r1 = r1 + 1

wb.save(wbName)
logging.info(f"Всё готово! -- [ {p} ] -- [ {time.strftime('%H:%M:%S', time.localtime())} ] ")
