from selenium import webdriver
##import openpyxl
from openpyxl import *
from datetime import datetime
import time

EGRN = None
status = 0 # состояние сайта
row = 0
wb = None
ws = None
RQ = {} # отчёт по статусам запросов
  
##------------------------------------------------------------##
def init():
  global EGRN
  EGRN = webdriver.Chrome()
  ##  EGRN.implicitly_wait(1)
  EGRN.set_page_load_timeout(10)
  EGRN.get("https://rosreestr.ru/wps/portal/p/cc_present/ir_egrn")

##------------------------------------------------------------##
def init_pyxl():
  global wb
  global ws
  try:
    wb = load_workbook(filename = 'dn.xlsx')
  except:
    wb = Workbook()
  ws = wb.worksheets[0]
  if ws.cell(row = 1, column = 1).value == None:
    ws.cell(row = 1, column = 1).value = "Номер запроса"
    ws.cell(row = 1, column = 2).value = "Дата запроса"
    ws.cell(row = 1, column = 3).value = "Статус запроса"
    wb.save(filename = 'dn.xlsx')

##------------------------------------------------------------##
def dn(d1, d2):
  global row
  global wb
  global ws
  global EGRN
  global RQ
  table = EGRN.find_element_by_css_selector("table.v-table-table")
  for tr in table.find_elements_by_css_selector("tr"):
    NZ = tr.find_elements_by_css_selector("td")[0].get_attribute("innerText")
    found = False
    for r in range(2, ws.max_row + 1):
      if ws.cell(row = r, column = 1).value == NZ:
        found = True
        row = r
        break

    DZ = tr.find_elements_by_css_selector("td")[1].get_attribute("innerText")
    dz = datetime.strptime(DZ[:10], "%d.%m.%Y")
    SZ = tr.find_elements_by_css_selector("td")[2].get_attribute("innerText")
    LZ = tr.find_elements_by_css_selector("td")[3].find_elements_by_css_selector("a")
    if (d1 <= dz <= d2):
      s = f"{NZ} {DZ} {SZ}"
      if len(LZ) > 0 and (not found or ws.cell(row = r, column = 3).value != "Завершена"):
        LZ[0].click()
        s = s + " :: download"
        RQ["Завершена и загружена"] = RQ.get("Завершена и загружена", 0) + 1
        time.sleep(1)
      if not found:
        row = ws.max_row + 1
      print(f"row = {row}")
      ws.cell(row = row, column = 1).value = NZ
      ws.cell(row = row, column = 2).value = DZ
      ws.cell(row = row, column = 3).value = SZ
      print(s)
      RQ[SZ] = RQ.get(SZ, 0) + 1
      RQ["_____Всего"] = RQ.get("_____Всего", 0) + 1
  wb.save(filename = 'dn.xlsx')
  return (d1 <= dz <= d2) or (d1 <= d2 <= dz)

def dnd(d, p1 = 1, p2 = 200):
  global EGRN
  global RQ
  RQ = {}
  init_pyxl()
  table = EGRN.find_element_by_css_selector("table.v-table-table")
  pm = EGRN.find_elements_by_css_selector("div.v-horizontallayout.v-horizontallayout-pageManagement.pageManagement > div > div")
  first_page = pm[1].find_element_by_xpath("div/div")
  is_first_page  = first_page.get_attribute("aria-pressed") != "false"
  if not is_first_page:
    first_page.click()
  while not is_first_page:
    time.sleep(1)
    is_first_page  = first_page.get_attribute("aria-pressed") != "false"
  print("Первая страница ОК")
  activepage = int(pm[3].find_element_by_css_selector("div.v-horizontallayout div.v-label.v-label-undef-w").get_attribute("innerText"))
  next_page = pm[4].find_element_by_css_selector("div")
  dd = d.split("-")
  if len(dd) == 1:
      dd.append(dd[0])
  dd[0] = datetime.strptime(dd[0], "%d.%m.%Y")
  dd[1] = datetime.strptime(dd[1], "%d.%m.%Y")
  dd.sort()

  ## skip to initial page if it is not the first one
  while activepage < p1:
      old_activepage = activepage
      next_page.click()
      while old_activepage == activepage:
        print(f'page {activepage} in range({p1}, {p2 + 1})')
        time.sleep(4)
        try:
          pm = EGRN.find_elements_by_css_selector("div.v-horizontallayout.v-horizontallayout-pageManagement.pageManagement > div > div")
          activepage = int(pm[3].find_element_by_css_selector("div.v-horizontallayout div.v-label.v-label-undef-w").get_attribute("innerText"))
          next_page = pm[4].find_element_by_css_selector("div")
          status = True
        except:
          activepage = 9999
          status = False
    
  status = True
  while dn(dd[0], dd[1]) and status and activepage in range(p1, p2):
    print(f"Обработана странца {activepage}")
    ## check if the current page is the final one
    if len(next_page.find_elements_by_css_selector("div.v-button.v-disabled.v-button-link.link")) > 0:
      break
    old_activepage = activepage
    next_page.click()
    while old_activepage == activepage:
      time.sleep(4)
      try:
        pm = EGRN.find_elements_by_css_selector("div.v-horizontallayout.v-horizontallayout-pageManagement.pageManagement > div > div")
        activepage = int(pm[3].find_element_by_css_selector("div.v-horizontallayout div.v-label.v-label-undef-w").get_attribute("innerText"))
        next_page = pm[4].find_element_by_css_selector("div")
        status = True
      except:
        activepage = 9999
        status = False

  print("===============================")
  print("Всё готово")        
  for sz in sorted(RQ):
    print("%s: %s" % (sz, RQ[sz]))
    
        

print("1. init()")
print("2. dnd(\"01.06.2020-30.06.2020\")")

