from recognize import recognize

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

import openpyxl, os, re, sys, time

RunFromIDLE = None # True если скрипт запущен из IDLE
SIpON = None
status = 0 # состояние сайта


##------------------------------------------------------------##
def SIpONinit():
  global SIpON
  global RunFromIDLE

  if RunFromIDLE:
    SIpON = webdriver.Chrome()
  else:
    chrome_options = Options()  
    chrome_options.add_argument("--headless")
    SIpON = webdriver.Chrome(options = chrome_options)

  SIpON.implicitly_wait(15)
  SIpON.set_page_load_timeout(30)
  try:
    SIpON.get("https://rosreestr.ru/wps/portal/online_request")
  except:
    print("Сайт Росреестра не работает")
    return 999

##------------------------------------------------------------##
def SIpONrestart(t):
  global SIpON
  if SIpON != None:
    SIpON.close()
    SIpON.quit()
  print(f"Restart")
  SIpONinit()

##------------------------------------------------------------##
def RecognizeCaptcha(c):
  global SIpON
  #code from https://gist.github.com/spirkaa/4c3b8ad8fd34324bd307#gistcomment-3157744
  img_base64 = SIpON.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", c)
  captcha = recognize(img_base64)
  return captcha
  
##------------------------------------------------------------##
def ifNone(s):
  if s == None:
    return ""
  else:
    return f"{s}"
  
##------------------------------------------------------------##
def GetInfoByAddr(subject_id = "", region_id = "", settlement_id = "", street = "", house = "", apartment = ""):
  global SIpON
  prev_subject_id = None
  prev_region_id = None
  prev_settlement_id = None
  prev_street = None
  prev_house = None
  prev_apartment = None

  i = 1 # номер попытки
  status = 1
  SIpON.implicitly_wait(1)
  SIpON.set_page_load_timeout(10)
  while status != 0 and i < 10:
    status = 0
    if i > 1:
      SIpONrestart(i)
    if status == 0:
      # открываем форму поиска, когда она скрыта после предыдущего поиска
      try:
        SetSearchCritVisible = SIpON.find_element_by_css_selector("a[href*='SetSearchCritVisible']")
        SetSearchCritVisible.click()
      except:
        pass
##      try:
##        captchaImage2 = SIpON.find_element_by_id("captchaImage2")
##      except:
##        status = 999 # капчи нет; вероятно, сайт не работает
##    if status == 0:    
##      try:
##        captcha = RecognizeCaptcha(captchaImage2)
##      except:
##        status = 999 # капча не проявилась
###### начало заполнения формы поиска
    if status == 0:
      # если форма готова для ввода данных, попробуем ввести новые данные и проверить что получилось
      j = 0 # номер попытки ввода данных
      while True:
        status = 0
        # пытаемся получить содержимое формы
        try:
          r_prev_subject_id = SIpON.find_element_by_css_selector("select[name='subject_id']")
          prev_subject_id = Select(r_prev_subject_id).first_selected_option.text
        except: # регион не может быть пустым
          j = j + 1
          status = 999
          break
        try:
          r_prev_region_id = SIpON.find_element_by_css_selector("select[name='region_id']")
          prev_region_id = Select(r_prev_region_id).first_selected_option.text
        except:
          prev_region_id = None
          pass #  район может быть пустым
        try:
          r_prev_settlement_id = SIpON.find_element_by_css_selector("select[name='settlement_id']")
          prev_settlement_id = Select(r_prev_settlement_id).first_selected_option.text
        except:
          prev_settlement_id = None
          pass #  нас. пункт может быть пустым
        r_prev_street = SIpON.find_element_by_css_selector("input[name='street']")
        prev_street = r_prev_street.get_attribute('value')
        r_prev_house = SIpON.find_element_by_css_selector("input[name='house']")
        prev_house = r_prev_house.get_attribute('value')
        r_prev_apartment = SIpON.find_element_by_css_selector("input[name='apartment']")
        prev_apartment = r_prev_apartment.get_attribute('value')
        
        if status == 0:
##          print(f"Try to search subject = [{subject_id}] region = [{region_id}] settlement= [{settlement_id}] street = [{street}] house = [{house}] apartment = [{apartment}]")
##          print(f"Data filled = [{prev_subject_id}] region = [{prev_region_id}] settlement= [{prev_settlement_id}] street = [{prev_street}] house = [{prev_house}] apartment = [{prev_apartment}]")
##          
          TestFilledData = True \
            and re.match(f"^{ifNone(subject_id)}", ifNone(prev_subject_id), re.IGNORECASE) != None \
            and re.match(f"^{ifNone(region_id)}", ifNone(prev_region_id), re.IGNORECASE) != None \
            and re.match(f"^{ifNone(settlement_id)}", ifNone(prev_settlement_id), re.IGNORECASE) != None \
            and ifNone(prev_street) == ifNone(street) \
            and ifNone(prev_house) == ifNone(house) \
            and ifNone(prev_apartment) == ifNone(apartment)

          
##          print(f"subject_id = {subject_id}", f"prev_subject_id = {prev_subject_id}")
##          print(re.match(f"^{ifNone(subject_id)}", ifNone(prev_subject_id), re.IGNORECASE))
##          print(".")
##          print(f"region_id = {region_id}", f"prev_region_id = {prev_region_id}")
##          print(re.match(f"^{ifNone(region_id)}", ifNone(prev_region_id), re.IGNORECASE))
##          print(".")
##          print(f"settlement_id = {settlement_id}", f"prev_settlement_id = {prev_settlement_id}")
##          print(re.match(f"^{ifNone(settlement_id)}", ifNone(prev_settlement_id), re.IGNORECASE))
##          print(".")
##          print(f"street = {street}", f"prev_street = {prev_street}")
##          print(f"house = {house}", f"prev_house = {prev_house}")
##          print(f"apartment = {apartment}", f"prev_apartment = {prev_apartment}")
##
##          print("=========================================================================")
          
          
          






















          if TestFilledData:
            break # все поля в форме соответствуют тому, что мы ищем
          else:
            status = 99
##            time.sleep(3)
            if re.match(f"^{ifNone(subject_id)}", ifNone(prev_subject_id), re.IGNORECASE) == None:
              r_prev_subject_id.send_keys(subject_id)
              time.sleep(1)

            if re.match(f"^{ifNone(region_id)}", ifNone(prev_region_id), re.IGNORECASE) == None:
              try:
                if region_id == None:
                  r_prev_region_id.send_keys(Keys.HOME)
                else:
                  r_prev_region_id.send_keys(region_id)
                  time.sleep(1)
              except:
                status = 999 # не удалось выбрать регион, вероятно, не загрузился список
                break
              
            if re.match(f"^{ifNone(settlement_id)}", ifNone(prev_settlement_id), re.IGNORECASE) == None:
  ##  внезапно: выбор НП может быть не предусмотрен в данном субъекте и/или регионе
              if r_prev_settlement_id.get_attribute("disabled"):
                if settlement_id == None: # НП пустой, всё ОК
                  pass
                else: # НП непустой, а выбор заблокирован - ошибка, на выход
                  status = 999
                  break
              else:
                if settlement_id == None: 
                  r_prev_settlement_id.send_keys(Keys.HOME)
                else:
                  r_prev_settlement_id.send_keys(settlement_id)
                  time.sleep(1)

  ##          print(f)
              
            r_prev_street.click()
            r_prev_street.clear()
            if street != None:
              r_prev_street.send_keys(street)
            
            r_prev_house.click()
            r_prev_house.clear()
            if house != None:
              r_prev_house.send_keys(house)

            r_prev_apartment.click()
            r_prev_apartment.clear()
            if apartment != None:
              r_prev_apartment.send_keys(apartment)
            j = j + 1
            
        if status == 99:
          if j > 10: # терпение лопнуло
            status = 999 # за 10 попыток не смогли ввести данные в форму и сверить содержимое формы
            break
          if j > 1:
            time.sleep(3)
######      конец заполнения формы поиска

######      попытка выполнить поиск
    if status == 0:
##      print(f"Try to search subject = [{subject_id}] region = [{region_id}] settlement= [{settlement_id}] street = [{street}] house = [{house}] apartment = [{apartment}]")
      try:
##        captchaText = SIpON.find_element_by_css_selector("input[type='text'][name='captchaText']")
##        SIpON.execute_script("arguments[0].click();", captchaText)
##        captchaText.clear()
##        captchaText.send_keys(captcha)
        SIpON.execute_script("arguments[0].click();", SIpON.find_element_by_id("submit-button"))
      except:
        status = 999
        
      if status == 0:
        links = SIpON.find_elements_by_css_selector("tr > td > a[href*='object_data_id']")
        kn = ""
        if links != None:
          for l in links:
            tr = l.find_element_by_xpath("..").find_element_by_xpath("..")
            kn = kn + ("\n" if len(kn) > 0 else "") + tr.find_elements_by_css_selector("td")[1].get_attribute("innerText")
        else:
          infomsg = SIpON.find_element_by_css_selector("td.infomsg1")
          if infomsg != None:
            kn = "Объект не найден"
        return kn

    if status == 999:
      i = i + 1
  else:
    if status == 999:
      return "Не удалось выполнить за 10 попыток"
    elif status == 0:
      return t # OK

##------------------------------------------------------------##
RunFromIDLE = "idlelib" in sys.modules
SIpONinit()

p = os.path.split(os.path.dirname(os.path.abspath(__file__)))[1]

from openpyxl import load_workbook
wb = load_workbook(filename = 'KNaddr.xlsx')
ws1 = wb.worksheets[0]
ws2 = wb.create_sheet("Sheet KNaddr")

r1 = 2 # строка в исходном списке
r1max = ws1.max_row - 1 # всего строк
r2 = 2 # строка в списке результатов
t = "Лицевой счёт\tКадастровый номер"
tl = t.split("\t")
for l in range(0, len(tl)):
  ws2.cell(row = 1, column = l + 1).value = tl[l]
##cc = ws1.cell(row = r , column = 1).value

while ws1.cell(row = r1, column=1).value != None:
  t = GetInfoByAddr( 
    ws1.cell(row = r1 , column = 3).value,
    ws1.cell(row = r1 , column = 4).value,
    ws1.cell(row = r1 , column = 5).value,
    ws1.cell(row = r1 , column = 6).value,
    ws1.cell(row = r1 , column = 7).value,
    ws1.cell(row = r1 , column = 8).value)

  print(f"----- [ {r1 - 1} из {r1max} ] ----- [ {p} ] ----------------------------------------------")
  print(t)
  tl = t.split("\n")
  for l in range(0, len(tl)):
    ws2.cell(row = r2, column = 1).value = ws1.cell(row = r1 , column = 1).value
    ws2.cell(row = r2, column = 2).value = tl[l]
    r2 = r2 + 1
  if (r1 - 1) % 10 == 0:
    wb.save("KNaddr.xlsx")
  r1 = r1 + 1

wb.save("KNaddr.xlsx")
print(f"\n\n\n\n\nВсё готово! ----- [ {p} ] ----- [ {time.strftime('%H:%M:%S', time.localtime())} ] ")
print()
