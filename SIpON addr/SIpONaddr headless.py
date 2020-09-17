from recognize import recognize

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

import openpyxl, os, re, time

SIpON = None
status = 0 # состояние сайта

prev_subject_id = "" 
prev_region_id = ""
prev_settlement_id = ""
prev_street = ""
prev_house = ""
prev_apartment = ""

subject_id = ""
region_id = ""
settlement_id = ""
street = ""
house = ""
apartment = ""
  
##------------------------------------------------------------##
def SIpONinit():
  global SIpON
  chrome_options = Options()  
  chrome_options.add_argument("--headless")
  SIpON = webdriver.Chrome(options = chrome_options)
  SIpON.implicitly_wait(15)
  SIpON.set_page_load_timeout(30)
  try:
    SIpON.get("https://rosreestr.gov.ru/wps/portal/online_request")
  except:
    print("Сайт Росреестра не работает")
    return 999

##------------------------------------------------------------##
def SIpONrestart(t):
  global SIpON
  if SIpON != None:
    SIpON.close()
    SIpON.quit()
  print(f"Restart")
  SIpONinit()

##------------------------------------------------------------##
def RecognizeCaptcha(c):
  global SIpON
  #code from https://gist.github.com/spirkaa/4c3b8ad8fd34324bd307#gistcomment-3157744
  img_base64 = SIpON.execute_script("""var ele = arguments[0];var cnv = document.createElement('canvas');cnv.width = 180; cnv.height = 50;cnv.getContext('2d').drawImage(ele, 0, 0);return cnv.toDataURL('image/png').substring(22);""", c)
  captcha = recognize(img_base64)
  return captcha
  
##------------------------------------------------------------##
def GetInfoByAddr(subject_id = "", region_id = "", settlement_id = "", street = "", house = "", apartment = ""):
  global SIpON
  global prev_subject_id
  global prev_region_id
  global prev_settlement_id
  global prev_street
  global prev_house
  global prev_apartment

  i = 1
  status = 1
  SIpON.implicitly_wait(1)
  SIpON.set_page_load_timeout(10)
  while status != 0 and i < 10:
    status = 0
    if i > 1:
      SIpONrestart(i)
      prev_subject_id = "" 
      prev_region_id = ""
      prev_settlement_id = ""
      prev_street = ""
      prev_house = ""
      prev_apartment = ""
    if status == 0:
      try:
        SetSearchCritVisible = SIpON.find_element_by_css_selector("a[href*='SetSearchCritVisible']")
        SetSearchCritVisible.click()
      except:
        pass
      try:
        captchaImage2 = SIpON.find_element_by_id("captchaImage2")
      except:
        status = 999 # капчи нет, вероятно, сайт не работает
    if status == 0:    
      try:
        captcha = RecognizeCaptcha(captchaImage2)
      except:
        status = 999 # капча не проявилась
        pass
    if status == 0:
      try:
        SIpON.execute_script("arguments[0].click();", SIpON.find_element_by_id("adress"))
        l = 1
        while True:
          r = SIpON.find_element_by_css_selector("select[name='subject_id']")
          if prev_subject_id != subject_id:
            time.sleep(3)
            r.send_keys(subject_id)
            rs = Select(r)
            rst = rs.first_selected_option.text
            if re.match(subject_id, rst, re.IGNORECASE) != None:
              prev_subject_id = subject_id
              break
            else:
              print(f"Ошибка при выборе региона, попытка {l}")
              l = l + 1
              if l > 3:
                print("Три попытки, на выход")
                raise
          else:
            break
          
        r = SIpON.find_element_by_css_selector("select[name='region_id']")           
        l = 1
        while True:
          if prev_region_id != region_id:
            time.sleep(3)
            r.send_keys(region_id)
            rs = Select(r)
            rst = rs.first_selected_option.text
            if re.match(region_id, rst, re.IGNORECASE) != None:
              prev_region_id = region_id
              break
            else:
              print(f"Ошибка при выборе района, попытка {l}")
              l = l + 1
              if l > 3:
                print("Три попытки, на выход")
                raise
          else:
            break

        r = SIpON.find_element_by_css_selector("select[name='settlement_id']")
        l = 1
        while True:
          if prev_settlement_id != settlement_id:
            time.sleep(3)
            if settlement_id == None:
              r.send_keys(Keys.HOME)
              prev_settlement_id = settlement_id
              break
            else:
              r.send_keys(settlement_id)
              rs = Select(r)
              rst = rs.first_selected_option.text
              if re.match(settlement_id, rst, re.IGNORECASE) != None:
                prev_settlement_id = settlement_id
                break
              else:
                print(f"Ошибка при выборе населённого пункта, попытка {l}")
                l = l + 1
                if l > 3:
                  print("Три попытки, на выход")
                  raise
          else:
            break
          
        r = SIpON.find_element_by_css_selector("input[name='street']")
        if street == None:
          street = ""
        if prev_street != street:
          r.click()
          r.clear()
          r.send_keys(street)
          prev_street = street
        r = SIpON.find_element_by_css_selector("input[name='house']")           
        if prev_house != house:
          r.click()
          r.clear()
          r.send_keys(house)
          prev_house = house
        r = SIpON.find_element_by_css_selector("input[name='apartment']")
        if apartment == None:
          apartment = ""
        if prev_apartment != apartment:
          r.click()
          r.clear()
          r.send_keys(apartment)
          prev_apartment = apartment
        captchaText = SIpON.find_element_by_css_selector("input[type='text'][name='captchaText']")
        SIpON.execute_script("arguments[0].click();", captchaText)
        captchaText.clear()
        captchaText.send_keys(captcha)
        SIpON.execute_script("arguments[0].click();", SIpON.find_element_by_id("submit-button"))
      except:
        status = 999
    if status == 0:
      links = SIpON.find_elements_by_css_selector("tr > td > a[href*='object_data_id']")
      kn = ""
      if links != None:
        for l in links:
          tr = l.find_element_by_xpath("..").find_element_by_xpath("..")
          kn = kn + ("\n" if len(kn) > 0 else "") + tr.find_elements_by_css_selector("td")[1].get_attribute("innerText")
      else:
        print("links not found")
        infomsg = SIpON.find_element_by_css_selector("td.infomsg1")
        if infomsg != None:
          kn = "Объект не найден"
      return kn
    if status == 999:
      i = i + 1
  else:
    if status == 999:
      return "Не удалось выполнить за 10 попыток"
    elif status == 0:
      return t # OK

##------------------------------------------------------------##
SIpONinit()

p = os.path.split(os.path.dirname(os.path.abspath(__file__)))[1]

from openpyxl import load_workbook
wb = load_workbook(filename = 'KNaddr.xlsx')
ws1 = wb.worksheets[0]
ws2 = wb.create_sheet("Sheet KNaddr")

r1 = 2 # строка в исходном списке
r1max = ws1.max_row - 1 # всего строк
r2 = 2 # строка в списке результатов
t = "Лицевой счёт\tКадастровый номер"
tl = t.split("\t")
for l in range(0, len(tl)):
  ws2.cell(row = 1, column = l + 1).value = tl[l]
##cc = ws1.cell(row = r , column = 1).value

while ws1.cell(row = r1, column=1).value != None:
  t = GetInfoByAddr( 
    ws1.cell(row = r1 , column = 3).value,
    ws1.cell(row = r1 , column = 4).value,
    ws1.cell(row = r1 , column = 5).value,
    ws1.cell(row = r1 , column = 6).value,
    ws1.cell(row = r1 , column = 7).value,
    ws1.cell(row = r1 , column = 8).value)

  print(f"----- [ {r1 - 1} из {r1max} ] ----- [ {p} ] ----------------------------------------------")
  print(t)
  tl = t.split("\n")
  for l in range(0, len(tl)):
    ws2.cell(row = r2, column = 1).value = ws1.cell(row = r1 , column = 1).value
    ws2.cell(row = r2, column = 2).value = tl[l]
    r2 = r2 + 1
  if (r1 - 1) % 10 == 0:
    wb.save("KNaddr.xlsx")
  r1 = r1 + 1

wb.save("KNaddr.xlsx")
print(f"\n\n\n\n\nВсё готово! ----- [ {p} ] ----- [ {time.strftime('%H:%M:%S', time.localtime())} ] ")
print()
